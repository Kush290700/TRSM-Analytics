# app/blueprints/products.py
from __future__ import annotations

import copy
import hashlib
import json
import logging
import math
import os
import csv
import threading
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone, timedelta
from functools import lru_cache, wraps
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from pathlib import Path
from urllib.parse import urlencode, unquote
from werkzeug.routing import BuildError

import numpy as np
import pandas as pd
from flask import Blueprint, abort, current_app, jsonify, make_response, redirect, render_template, request, session, url_for
from flask_login import current_user, login_required as _login_required
from io import BytesIO, StringIO
from werkzeug.datastructures import MultiDict

try:
    from app.core.exports import dataframes_to_xlsx_response, dataframe_to_csv_response, sanitize_filename
except Exception:  # pragma: no cover
    dataframes_to_xlsx_response = None  # type: ignore
    dataframe_to_csv_response = None  # type: ignore

    def sanitize_filename(name: str, default: str = "export") -> str:  # type: ignore[override]
        safe = str(name or default).replace("/", "_").replace("\\", "_")
        return safe or default

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
except Exception:  # pragma: no cover - optional dependency
    pa = None
    pq = None

# Optional: if your project already has Flask-Caching wired as "..cache import cache"
try:
    from ..cache import cache  # type: ignore
except Exception:  # pragma: no cover
    cache = None

# RBAC imports
try:
    from ..core.rbac import can_view_costs, roles_for as rbac_roles_for, route_permission_override_allows_request  # type: ignore
except Exception:
    rbac_roles_for = None  # type: ignore
    route_permission_override_allows_request = None  # type: ignore

    def can_view_costs(user=None):  # pragma: no cover
        return False

from app.core.sensitive_data import mask_dataframe

# Shared/global filters (other blueprints)
try:
    from app.services import filters as filter_svc  # type: ignore
except Exception:  # pragma: no cover
    filter_svc = None

# Fact store for canonical sales data
try:
    from app.services import fact_store  # type: ignore
except Exception:  # pragma: no cover
    fact_store = None

try:
    from app.services import analytics_utils as au  # type: ignore
except Exception:  # pragma: no cover
    au = None

try:
    from app.services.data_access import get_fact_context, current_data_version  # type: ignore
except Exception:  # pragma: no cover
    get_fact_context = None  # type: ignore
    current_data_version = lambda: str(int(time.time()))  # type: ignore

logger = logging.getLogger(__name__)
bp = Blueprint("products", __name__, url_prefix="/products")

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_PRODUCTS_DIR = (PROJECT_ROOT / "cache").resolve()
DEFAULT_PRODUCTS_PATH = (DEFAULT_PRODUCTS_DIR / "products.parquet").resolve()
_BOOL_TRUE = {"1", "true", "yes", "on", "y", "t"}
_BOOL_FALSE = {"0", "false", "no", "off", "n"}
PARQUET_REBUILD_TTL_MIN = int(os.getenv("PRODUCTS_PARQUET_TTL_MIN", "240"))
PARQUET_LOCK_TIMEOUT = int(os.getenv("PRODUCTS_PARQUET_LOCK_TIMEOUT", "30"))


def _coerce_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    normalized = str(value).strip().lower()
    if normalized in _BOOL_TRUE:
        return True
    if normalized in _BOOL_FALSE:
        return False
    try:
        return bool(int(normalized))
    except Exception:
        return default


def _default_products_dir(cfg: Any = None) -> Path:
    """Best-effort resolution for where to keep products parquet files."""
    # Prefer configured DATA_DIR if present
    candidates: List[Path] = []
    data_dir_raw = getattr(cfg, "DATA_DIR", None) if cfg else None
    data_dir_raw = data_dir_raw or os.getenv("DATA_DIR")
    if data_dir_raw:
        try:
            candidates.append(Path(data_dir_raw))
        except Exception:
            candidates = candidates

    # Fall back to instance/data/parquet if available
    try:
        from flask import current_app as _cur_app  # type: ignore

        inst = getattr(_cur_app, "instance_path", None)
        if inst:
            candidates.append(Path(inst) / "data" / "parquet")
    except Exception:
        pass

    # Finally use the repo cache dir
    candidates.append(DEFAULT_PRODUCTS_DIR)

    for cand in candidates:
        try:
            return cand.expanduser().resolve()
        except Exception:
            continue
    return DEFAULT_PRODUCTS_DIR


def _normalize_product_id(raw: str) -> str:
    """Normalize product identifiers from URL path segments."""
    try:
        return unquote(raw or "").strip()
    except Exception:
        return str(raw or "").strip()


def _product_drilldown_v2_enabled() -> bool:
    raw = current_app.config.get("PRODUCT_DRILLDOWN_V2")
    if raw is None:
        raw = os.getenv("PRODUCT_DRILLDOWN_V2", "0")
    return str(raw).strip().lower() in {"1", "true", "yes", "on", "y"}


def _product_forecast_v1_enabled() -> bool:
    raw = current_app.config.get("PRODUCT_FORECAST_V1")
    if raw is None:
        raw = os.getenv("PRODUCT_FORECAST_V1", "0")
    if raw is None:
        raw = current_app.config.get("FEATURE_FORECAST_ENABLED")
    return str(raw).strip().lower() in {"1", "true", "yes", "on", "y"}


def _redact_product_drilldown_cost_fields(context: Dict[str, Any]) -> Dict[str, Any]:
    """
    Hide cost-sensitive values for users without cost permissions.
    """
    if not isinstance(context, dict):
        return {}
    safe = copy.deepcopy(context)

    kpis = safe.get("kpis")
    if isinstance(kpis, dict):
        for key in (
            "cost",
            "profit",
            "margin_pct",
            "profit_per_lb",
            "margin_risk_exposure_pct",
            "margin_uplift_to_target",
            "cost_coverage_pct",
            "data_confidence_score",
        ):
            kpis[key] = None
        mom = kpis.get("mom")
        if isinstance(mom, dict):
            mom["margin_delta_pp"] = None
        delta_28 = kpis.get("delta_28d")
        if isinstance(delta_28, dict):
            delta_28["margin_delta_pp"] = None
            delta_28["current_margin_pct"] = None
            delta_28["prior_margin_pct"] = None

    trend = safe.get("trend")
    if isinstance(trend, dict):
        trend["cost"] = [None for _ in (trend.get("months") or [])]
        trend["profit"] = [None for _ in (trend.get("months") or [])]
        trend["margin_pct"] = [None for _ in (trend.get("months") or [])]
        trend["profit_per_lb"] = [None for _ in (trend.get("months") or [])]

    distributions = safe.get("distributions")
    if isinstance(distributions, dict):
        distributions["margin"] = {"samples": [], "p10": None, "p50": None, "p90": None}
        distributions["profit_per_lb"] = {"samples": [], "p10": None, "p50": None, "p90": None}

    sensitive_fields = {
        "cost",
        "profit",
        "margin_pct",
        "profit_per_lb",
        "margin_delta_pp",
        "cost_coverage_pct",
        "missing_cost_revenue",
        "missing_cost_rows",
        "uplift_to_target",
        "target_gap_pp",
        "revenue_exposure_pct",
        "data_confidence_score",
    }

    datasets = safe.get("datasets")
    if isinstance(datasets, dict):
        for key, rows in list(datasets.items()):
            if not isinstance(rows, list):
                continue
            cleaned_rows: List[Dict[str, Any]] = []
            for row in rows:
                if not isinstance(row, dict):
                    continue
                clean = dict(row)
                for field in sensitive_fields:
                    if field in clean:
                        clean[field] = None
                cleaned_rows.append(clean)
            datasets[key] = cleaned_rows

    for section_key in ("customers", "regions", "suppliers", "ship_methods", "margin_risk"):
        section = safe.get(section_key)
        if not isinstance(section, dict):
            continue
        for list_key in ("rows", "top_rows", "drop_signals", "top_gainers", "top_decliners"):
            rows = section.get(list_key)
            if not isinstance(rows, list):
                continue
            cleaned_rows = []
            for row in rows:
                if not isinstance(row, dict):
                    continue
                clean = dict(row)
                for field in sensitive_fields:
                    if field in clean:
                        clean[field] = None
                cleaned_rows.append(clean)
            section[list_key] = cleaned_rows

    quality = safe.get("quality")
    if isinstance(quality, dict):
        quality["cost_coverage_pct"] = None
        quality["missing_cost_rows"] = None
        quality["missing_cost_revenue"] = None

    story = safe.get("performance_story")
    if isinstance(story, dict):
        decomp = story.get("decomposition")
        if isinstance(decomp, dict):
            decomp["price_effect"] = None
            decomp["volume_effect"] = None
            decomp["mix_effect"] = None
            decomp["delta_profit"] = None
        current_window = story.get("current_window")
        prior_window = story.get("prior_window")
        if isinstance(current_window, dict):
            current_window["profit"] = None
            current_window["margin_pct"] = None
            current_window["profit_per_lb"] = None
        if isinstance(prior_window, dict):
            prior_window["profit"] = None
            prior_window["margin_pct"] = None
            prior_window["profit_per_lb"] = None

    price_volume = safe.get("price_volume")
    if isinstance(price_volume, dict):
        points = price_volume.get("points")
        if isinstance(points, list):
            for point in points:
                if isinstance(point, dict):
                    point["revenue"] = None
                    point["margin_pct"] = None

    weight_analytics = safe.get("weight_analytics")
    if isinstance(weight_analytics, dict):
        summary = weight_analytics.get("summary")
        if isinstance(summary, dict):
            summary["profit_per_lb"] = None

    margin_risk = safe.get("margin_risk")
    if isinstance(margin_risk, dict):
        summary = margin_risk.get("summary")
        if isinstance(summary, dict):
            summary["uplift_to_target"] = None
            summary["revenue_exposure_pct"] = None

    risk_opportunity = safe.get("risk_opportunity")
    if isinstance(risk_opportunity, dict):
        def _is_margin_signal(row: Dict[str, Any]) -> bool:
            title = str(row.get("title") or "").lower()
            detail = str(row.get("detail") or "").lower()
            return "margin" in title or "margin" in detail or "pricing recovery" in title

        for list_key in ("risks", "opportunities"):
            rows = risk_opportunity.get(list_key)
            if isinstance(rows, list):
                risk_opportunity[list_key] = [row for row in rows if isinstance(row, dict) and not _is_margin_signal(row)]
        primary_risk = risk_opportunity.get("primary_risk")
        if isinstance(primary_risk, dict) and _is_margin_signal(primary_risk):
            risk_opportunity["primary_risk"] = None
        primary_opp = risk_opportunity.get("primary_opportunity")
        if isinstance(primary_opp, dict) and _is_margin_signal(primary_opp):
            risk_opportunity["primary_opportunity"] = None

    header_summary = safe.get("header_summary")
    if isinstance(header_summary, dict):
        if "margin" in str(header_summary.get("primary_risk") or "").lower():
            header_summary["primary_risk"] = None
            header_summary["primary_risk_detail"] = None
        if "pricing recovery" in str(header_summary.get("primary_opportunity") or "").lower():
            header_summary["primary_opportunity"] = None
            header_summary["primary_opportunity_detail"] = None

    decision_panel = safe.get("decision_panel")
    if isinstance(decision_panel, dict):
        actions = decision_panel.get("actions")
        if isinstance(actions, list):
            decision_panel["actions"] = [
                row
                for row in actions
                if isinstance(row, dict)
                and "margin" not in str(row.get("title") or "").lower()
                and "margin" not in str(row.get("detail") or "").lower()
            ]

    return safe


def resolve_products_parquet_path(explicit_path: Optional[str] = None) -> str:
    try:
        cfg = current_app.config  # type: ignore[attr-defined]
    except Exception:
        cfg = {}
    data_dir = _default_products_dir(cfg)
    default_path = (data_dir / "products.parquet").resolve()
    path = (
        explicit_path
        or getattr(cfg, "PRODUCTS_PARQUET_PATH", None)
        or os.getenv("PRODUCTS_PARQUET_PATH")
        or getattr(cfg, "PRODUCTS_SALES_PARQUET", None)
        or os.getenv("PRODUCTS_SALES_PARQUET")
        or default_path
    )
    try:
        return Path(path).expanduser().resolve().as_posix()
    except Exception:
        return default_path.as_posix()


@dataclass
class ProductsParquetStatus:
    path: str
    exists: bool
    auto_create: bool
    valid_schema: bool = True
    created_dir: bool = False
    created_placeholder: bool = False
    used_sentinel: bool = False
    warning: Optional[str] = None
    error: Optional[str] = None
    columns_missing: List[str] = field(default_factory=list)
    schema_version: Optional[str] = None
    meta_path: Optional[str] = None
    rebuilt: bool = False
    rows: Optional[int] = None

    @property
    def available(self) -> bool:
        return self.exists and self.error is None and self.valid_schema

# -----------------------------
# Tuning knobs
# -----------------------------
DEFAULT_MONTHS = 6           # default filter window
API_TTL_SEC = 180            # API cache TTL
PAYLOAD_LRU_TTL_SEC = 180    # in-process compute cache TTL
TOP_N_DEFAULT = 15
TOP_N_MAX = 5000
TABLE_PAGE_SIZE_DEFAULT = 50
TABLE_PAGE_SIZE_MAX = 100
BUBBLE_LIMIT_MAX = 250
SEGMENTS_PAGE_SIZE_MAX = 200
HISTOGRAM_SAMPLE_MAX = 10_000
CO_PURCHASE_SAMPLE_ORDERS = 4000
AGGREGATE_CACHE_TTL_SEC = int(os.getenv("PRODUCTS_AGG_CACHE_TTL_SEC", str(1030 * 60)))

# -----------------------------
# Auth helpers (minimal + safe)
# -----------------------------
# Allow all standard roles to view Products + drilldown pages.
VIEW_ROLES = ("production", "gm", "owner", "sales_manager", "admin", "sales", "analyst", "viewer")


def login_required(fn):
    """Respects LOGIN_DISABLED / AUTHZ_DISABLED for demo/local."""
    protected = _login_required(fn)

    @wraps(fn)
    def wrapper(*args, **kwargs):
        if current_app.config.get("LOGIN_DISABLED") or current_app.config.get("AUTHZ_DISABLED"):
            return fn(*args, **kwargs)
        return protected(*args, **kwargs)

    return wrapper


def requires_roles(*roles: str):
    """Lightweight role gate; supports current_user.roles (list) or current_user.role (string)."""
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if current_app.config.get("AUTHZ_DISABLED"):
                return fn(*args, **kwargs)
            if callable(route_permission_override_allows_request):
                try:
                    allow_override = route_permission_override_allows_request()
                except Exception:
                    allow_override = None
                if allow_override is not None:
                    if not allow_override:
                        abort(403)
                    return fn(*args, **kwargs)

            wanted = {str(r).strip().lower() for r in roles if r}
            user_roles = set()
            try:
                if rbac_roles_for:
                    user_roles |= {str(r).strip().lower() for r in rbac_roles_for(current_user)}
                else:
                    if hasattr(current_user, "roles") and current_user.roles:
                        for role in current_user.roles:
                            name = getattr(role, "name", None)
                            user_roles.add(str(name or role).strip().lower())
                    if hasattr(current_user, "role") and current_user.role:
                        user_roles.add(str(current_user.role).strip().lower())
            except Exception:
                user_roles = set()

            if wanted and not user_roles.intersection(wanted):
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return decorator


# -----------------------------
# JSON / caching helpers
# -----------------------------
def _json_default(o: Any):
    if isinstance(o, (np.integer,)):
        return int(o)
    if isinstance(o, (np.floating,)):
        return float(o)
    if isinstance(o, (np.bool_,)):
        return bool(o)
    if isinstance(o, (pd.Timestamp,)):
        return o.isoformat()
    if isinstance(o, (pd.Period,)):
        return str(o)
    return str(o)


def _json_dumps(payload: Dict[str, Any]) -> str:
    # Replace NaN/Inf with None-friendly values
    def _clean(x):
        if isinstance(x, dict):
            return {k: _clean(v) for k, v in x.items()}
        if isinstance(x, list):
            return [_clean(v) for v in x]
        if isinstance(x, float) and (np.isnan(x) or np.isinf(x)):
            return None
        return x

    return json.dumps(_clean(payload), sort_keys=True, separators=(",", ":"), default=_json_default)


def _etag(body: str) -> str:
    return hashlib.sha256(body.encode("utf-8")).hexdigest()


def _user_cache_key() -> str:
    try:
        uid = getattr(current_user, "id", None) or getattr(current_user, "email", None) or "anon"
        roles = getattr(current_user, "roles", None) or getattr(current_user, "role", None) or ""
        scope_hash = ""
        try:
            from app.core.access_policy import scope_for_user  # type: ignore

            scope = scope_for_user(current_user, use_cache=True)
            scope_hash = getattr(scope, "scope_hash", "") or ""
        except Exception:
            scope_hash = ""
        return f"{uid}:{roles}:{scope_hash}"
    except Exception:
        return "anon"


def _cache_version() -> str:
    try:
        return current_data_version()
    except Exception:
        return str(int(time.time() // max(1, CACHE_TTL_SECONDS)))


def _ttl_bucket(ttl: int) -> int:
    return int(time.time() // max(1, ttl))


def _filters_fingerprint(filters: Dict[str, Any]) -> str:
    blob = json.dumps(filters or {}, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def _api_cache_key(endpoint: str, filters: Dict[str, Any], extra: Optional[Dict[str, Any]] = None) -> str:
    raw = {
        "ep": endpoint,
        "u": _user_cache_key(),
        "f": _filters_fingerprint(filters),
        "x": extra or {},
        "v": _cache_version(),
    }
    key = "products:" + hashlib.sha256(json.dumps(raw, sort_keys=True, default=str).encode("utf-8")).hexdigest()
    try:
        if filter_svc and hasattr(filter_svc, "bind_filter_cache_key"):
            filter_svc.bind_filter_cache_key(key)
    except Exception:
        pass
    return key


def _log_timing(event: str, started: float, filters: Dict[str, Any], extra: Optional[Dict[str, Any]] = None) -> None:
    try:
        payload = {
            "event": event,
            "duration_ms": int((time.perf_counter() - started) * 1000),
            "filters_hash": _filters_fingerprint(filters),
        }
        if extra:
            payload.update(extra)
        logger.info("products.timing", extra=payload)
    except Exception:
        logger.debug("products.timing_failed", exc_info=True)


def _json_response(payload: Dict[str, Any], *, status: int = 200, cache_key: Optional[str] = None, ttl: int = API_TTL_SEC):
    """ETag + optional server cache (Flask-Caching)."""
    inm = request.headers.get("If-None-Match")

    body_payload: Dict[str, Any] = payload
    if isinstance(payload, dict) and "ok" not in payload:
        body_payload = {"ok": True, "data": payload}
        # Preserve top-level keys for backward compatibility
        body_payload.update(payload)

    if cache is not None and cache_key and not current_app.config.get("TESTING"):
        cached = cache.get(cache_key)
        if cached:
            et = cached.get("etag")
            body = cached.get("body", "")
            if inm and et and inm == et:
                resp = make_response("", 304)
            else:
                resp = make_response(body, status)
                resp.mimetype = "application/json"
            resp.headers["ETag"] = et
            resp.headers["Cache-Control"] = f"private, max-age={ttl}"
            return resp

    body = _json_dumps(body_payload)
    et = _etag(body)

    if cache is not None and cache_key and not current_app.config.get("TESTING"):
        cache.set(cache_key, {"etag": et, "body": body}, timeout=ttl)

    if inm and inm == et:
        resp = make_response("", 304)
    else:
        resp = make_response(body, status)
        resp.mimetype = "application/json"
    resp.headers["ETag"] = et
    resp.headers["Cache-Control"] = f"private, max-age={ttl}" if status < 400 else "no-store"
    return resp


def _error_response(message: str, *, status: int = 400):
    return _json_response({"ok": False, "error": {"message": str(message)}}, status=status)


# -----------------------------
# Parquet warnings
# -----------------------------
def _attach_parquet_warning(payload: Dict[str, Any], status: ProductsParquetStatus, *, add_empty_data: bool = False) -> Dict[str, Any]:
    if not status or not status.warning:
        return payload

    wrapped = copy.deepcopy(payload)
    if add_empty_data and "data" not in wrapped:
        wrapped["data"] = []
    meta = wrapped.setdefault("meta", {})
    meta.setdefault("products_parquet_path", status.path)
    meta.setdefault("products_parquet_available", status.available)
    meta.setdefault("products_parquet_auto_create", status.auto_create)
    meta.setdefault("products_parquet_created_placeholder", status.created_placeholder)
    meta.setdefault("products_parquet_schema_version", status.schema_version)
    meta.setdefault("products_parquet_rebuilt", status.rebuilt)
    meta.setdefault("products_parquet_rows", status.rows)
    if status.columns_missing:
        meta.setdefault("products_parquet_missing_columns", status.columns_missing)
    wrapped.setdefault("warning", status.warning)
    return wrapped


# -----------------------------
# Parquet store (loads once, reload on change)
# -----------------------------
@dataclass(frozen=True)
class CanonCols:
    date: str = "date"
    product_id: str = "product_id"
    product_name: str = "product_name"
    sku: str = "sku"
    cost: str = "cost"
    customer_id: str = "customer_id"
    customer_name: str = "customer_name"
    region: str = "region"
    region_id: str = "region_id"
    supplier: str = "supplier"
    supplier_id: str = "supplier_id"
    order_id: str = "order_id"
    qty: str = "qty"
    weight: str = "weight_lb"
    revenue: str = "revenue"
    discount: str = "discount"
    margin: str = "margin"
    margin_pct: str = "margin_pct"
    unit_price: str = "unit_price"
    unit_cost: str = "unit_cost"
    qty_basis: str = "qty_basis"
    qty_basis_label: str = "qty_basis_label"


CAN = CanonCols()

REQUIRED_PRODUCTS_COLUMNS: List[str] = [
    CAN.date,
    CAN.product_id,
    CAN.product_name,
    CAN.sku,
    CAN.customer_id,
    CAN.customer_name,
    CAN.region,
    CAN.supplier,
    CAN.order_id,
    CAN.qty,
    CAN.weight,
    CAN.revenue,
    CAN.discount,
    CAN.qty_basis,
    CAN.qty_basis_label,
    CAN.unit_price,
    CAN.cost,
    CAN.margin,
    CAN.margin_pct,
    CAN.unit_cost,
]


def _empty_products_frame() -> pd.DataFrame:
    """Create an empty frame with the required columns and stable dtypes."""
    return pd.DataFrame(
        {
            CAN.date: pd.Series(dtype="datetime64[ns]"),
            CAN.product_id: pd.Series(dtype="string"),
            CAN.product_name: pd.Series(dtype="string"),
            CAN.sku: pd.Series(dtype="string"),
            CAN.customer_id: pd.Series(dtype="string"),
            CAN.customer_name: pd.Series(dtype="string"),
            CAN.region: pd.Series(dtype="string"),
            CAN.region_id: pd.Series(dtype="string"),
            CAN.supplier: pd.Series(dtype="string"),
            CAN.supplier_id: pd.Series(dtype="string"),
            CAN.order_id: pd.Series(dtype="string"),
            CAN.qty: pd.Series(dtype="float64"),
            CAN.weight: pd.Series(dtype="float64"),
            CAN.qty_basis: pd.Series(dtype="float64"),
            CAN.qty_basis_label: pd.Series(dtype="string"),
            CAN.revenue: pd.Series(dtype="float64"),
            CAN.cost: pd.Series(dtype="float64"),
            CAN.margin: pd.Series(dtype="float64"),
            CAN.margin_pct: pd.Series(dtype="float64"),
            CAN.unit_price: pd.Series(dtype="float64"),
            CAN.unit_cost: pd.Series(dtype="float64"),
            CAN.discount: pd.Series(dtype="float64"),
        }
    )


def _resolve_col(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _to_numeric(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def _to_dt(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s, errors="coerce").dt.tz_localize(None)


def normalize_products_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Produce canonical product analytics columns with defensive defaults.
    This is the single normalization layer for all product endpoints.
    """
    if df is None or df.empty:
        return _empty_products_frame()

    out = pd.DataFrame(index=df.index)

    # Resolve source columns (best-effort, tolerant of schema drift)
    src_date = _resolve_col(
        df,
        (
            "Date",
            "date",
            "ShipDate",
            "DateShipped",
            "DateExpected",
            "DateOrdered",
            "OrderDate",
            "InvoiceDate",
            "created_at",
        ),
    )
    src_pid = _resolve_col(df, ("ProductId", "product_id", "SKU", "sku", "ItemId", "ProductCode"))
    src_sku = _resolve_col(df, ("SKU", "Sku", "SkuName", "product_sku", "ItemNumber", "ItemId", "ProductCode"))
    src_pname = _resolve_col(
        df,
        ("SkuName", "SKUName", "ProductName", "product_name", "Name", "Description", "Product_Label", "Product"),
    )
    src_cid = _resolve_col(df, ("CustomerId", "customer_id", "ClientId"))
    src_cname = _resolve_col(df, ("CustomerName", "customer_name", "ClientName", "AccountName"))
    src_region_id = _resolve_col(df, ("RegionId", "region_id", "RegionID"))
    src_region = _resolve_col(df, ("RegionName", "region", "Region", "region_name"))
    src_supplier_id = _resolve_col(df, ("SupplierId", "supplier_id", "VendorId", "VendorID"))
    src_supplier = _resolve_col(df, ("SupplierName", "supplier", "Supplier", "Supplier_Name"))
    src_order = _resolve_col(df, ("OrderId", "order_id", "OrderNumber", "OrderNo", "InvoiceId"))
    src_qty = _resolve_col(
        df,
        (
            "QuantityShipped",
            "QuantityOrdered",
            "QtyShipped",
            "QtyOrdered",
            "Quantity",
            "ItemCount",
            "pack_item_count_sum",
            "QtyNative",
            CAN.qty,
            CAN.qty_basis,
        ),
    )
    src_weight = _resolve_col(df, ("WeightLb", "pack_weight_lb_sum", "weight_lb", "weight", "ShippedLb", CAN.weight))
    src_rev = _resolve_col(df, ("Revenue", "revenue_shipped", "revenue_ordered", "TotalSales", "ExtPrice", "LinesTotalPrice", CAN.revenue))
    src_cost = _resolve_col(
        df,
        (
            "Cost",
            "total_cost",
            "TotalCost",
            "Total Cost",
            "cost_shipped",
            "cost_ordered",
            "COGS",
            "cogs",
            "ExtCost",
            "ExtendedCost",
            "CostAmount",
            "CostLine",
            CAN.cost,
        ),
    )
    src_margin = _resolve_col(df, ("Margin", "Profit", "gross_margin_shipped", "gross_margin_ordered", "ProfitDollars", CAN.margin))
    src_unit_cost = _resolve_col(
        df,
        (
            "CostPerUnit",
            "unit_cost_effective",
            "UnitCost",
            "CostPrice",
            "CostPrice_x",
            "CostPrice_y",
            "CostPrice_line",
            "StandardCost",
            "LandedCost",
            CAN.unit_cost,
        ),
    )
    src_disc = _resolve_col(df, ("Discount", "discount", "discount_amount", "discount_rate", CAN.discount))

    out[CAN.date] = _to_dt(df[src_date]) if src_date else pd.NaT

    # IDs as strings (stable joins + filtering)
    out[CAN.product_id] = df[src_pid].astype("string") if src_pid else pd.Series(pd.NA, index=df.index, dtype="string")
    out[CAN.product_name] = df[src_pname].astype("string") if src_pname else pd.Series(pd.NA, index=df.index, dtype="string")
    out[CAN.sku] = df[src_sku].astype("string") if src_sku else pd.Series(pd.NA, index=df.index, dtype="string")
    out[CAN.customer_id] = df[src_cid].astype("string") if src_cid else pd.Series(pd.NA, index=df.index, dtype="string")
    out[CAN.customer_name] = df[src_cname].astype("string") if src_cname else pd.Series(pd.NA, index=df.index, dtype="string")
    out[CAN.region_id] = df[src_region_id].astype("string") if src_region_id else pd.Series(pd.NA, index=df.index, dtype="string")
    out[CAN.region] = df[src_region].astype("string") if src_region else pd.Series(pd.NA, index=df.index, dtype="string")
    out[CAN.supplier_id] = df[src_supplier_id].astype("string") if src_supplier_id else pd.Series(pd.NA, index=df.index, dtype="string")
    out[CAN.supplier] = df[src_supplier].astype("string") if src_supplier else pd.Series(pd.NA, index=df.index, dtype="string")
    out[CAN.order_id] = df[src_order].astype("string") if src_order else pd.Series(pd.NA, index=df.index, dtype="string")

    # Base numerics (only additive measures are filled to zero)
    revenue = _to_numeric(df[src_rev]) if src_rev else pd.Series(0.0, index=df.index, dtype="float64")
    qty_units = _to_numeric(df[src_qty]) if src_qty else pd.Series(0.0, index=df.index, dtype="float64")
    weight_lb = _to_numeric(df[src_weight]) if src_weight else pd.Series(0.0, index=df.index, dtype="float64")

    rev_clean = revenue.fillna(0.0)
    out[CAN.revenue] = rev_clean
    out[CAN.qty] = qty_units.fillna(0.0)
    out[CAN.weight] = weight_lb.fillna(0.0)

    # Determine quantity basis (weight first if available)
    weight_total = float(weight_lb.fillna(0.0).abs().sum())
    qty_total = float(qty_units.fillna(0.0).abs().sum())
    use_weight = weight_total > 0.0
    qty_label = "Weight (lb)" if use_weight else "Quantity"
    qty_basis = weight_lb if use_weight else qty_units
    out[CAN.qty_basis] = qty_basis.fillna(0.0)
    out[CAN.qty_basis_label] = pd.Series(qty_label, index=df.index, dtype="string")

    # Cost resolution rules
    margin_series = _to_numeric(df[src_margin]) if src_margin else None
    cost = au.resolve_cost(
        df,
        cost_col=src_cost,
        units_col=src_qty,
        weight_col=src_weight,
        cost_rate_cols=[c for c in [src_unit_cost] if c],
    )
    if margin_series is not None and margin_series.notna().any():
        alt_cost = rev_clean - margin_series
        missing = cost.isna() | (cost == 0)
        fill_mask = missing & alt_cost.notna() & (alt_cost != 0)
        cost = cost.where(~fill_mask, alt_cost)

    cost_valid = cost.notna() & (cost > 0)
    out[CAN.cost] = cost.where(cost_valid, np.nan)
    margin = rev_clean - cost
    out[CAN.margin] = margin.where(cost_valid, np.nan)
    out[CAN.margin_pct] = np.where(
        (rev_clean > 0) & cost_valid,
        out[CAN.margin] / rev_clean * 100.0,
        np.nan,
    )
    out[CAN.margin_pct] = out[CAN.margin_pct].clip(lower=-200.0, upper=200.0)

    out[CAN.unit_price] = np.where(out[CAN.qty_basis] > 0, rev_clean / out[CAN.qty_basis], np.nan)
    out[CAN.unit_cost] = np.where(
        (out[CAN.qty_basis] > 0) & cost_valid,
        cost / out[CAN.qty_basis],
        np.nan,
    )

    out[CAN.discount] = _to_numeric(df[src_disc]).fillna(0.0) if src_disc else 0.0

    # Clean strings
    for c in (CAN.product_name, CAN.customer_name, CAN.region_id, CAN.region, CAN.supplier_id, CAN.supplier, CAN.sku):
        out[c] = out[c].astype("string").str.strip()
        out[c] = out[c].where(out[c].notna() & (out[c].str.len() > 0))

    # Fallback name/sku
    out[CAN.product_name] = out[CAN.product_name].fillna(out[CAN.product_id]).fillna("Unknown")
    out[CAN.sku] = out[CAN.sku].fillna(out[CAN.product_id])

    # Drop rows with no date or no product_id (keeps intelligence reliable)
    out = out.dropna(subset=[CAN.date, CAN.product_id])

    # Small perf wins
    out[CAN.region] = out[CAN.region].astype("category")
    out[CAN.supplier] = out[CAN.supplier].astype("category")

    # Alias for compatibility with legacy code if present
    out["weight"] = out[CAN.weight]

    return out


def _standardize_sales_df(df: pd.DataFrame) -> pd.DataFrame:
    """Backward-compatible alias for normalization."""
    return normalize_products_df(df)


def _parquet_lock(lock_path: Path, timeout: int = PARQUET_LOCK_TIMEOUT):
    """Return a context manager that locks the parquet path across processes."""
    try:
        from filelock import FileLock  # type: ignore

        return FileLock(lock_path.as_posix(), timeout=timeout)
    except Exception:
        class _SimpleFileLock:
            def __init__(self, path: Path, timeout: int = 30):
                self.path = path
                self.timeout = timeout
                self.fd: Optional[int] = None

            def __enter__(self):
                start = time.time()
                while True:
                    try:
                        self.fd = os.open(self.path.as_posix(), os.O_CREAT | os.O_EXCL | os.O_RDWR)
                        os.write(self.fd, str(os.getpid()).encode("utf-8"))
                        return self
                    except FileExistsError:
                        if (time.time() - start) > self.timeout:
                            raise TimeoutError(f"Timed out acquiring lock {self.path}")
                        time.sleep(0.1)
                    except Exception:
                        raise

            def __exit__(self, exc_type, exc, tb):
                try:
                    if self.fd:
                        os.close(self.fd)
                finally:
                    try:
                        if self.path.exists():
                            self.path.unlink()
                    except Exception:
                        pass

        return _SimpleFileLock(lock_path, timeout)


def _needs_rebuild(path_obj: Path, ttl_min: int) -> bool:
    try:
        stat = path_obj.stat()
        if stat.st_size <= 0:
            return True
        if ttl_min and ttl_min > 0:
            age_min = (time.time() - stat.st_mtime) / 60.0
            if age_min > ttl_min:
                return True
    except FileNotFoundError:
        return True
    except Exception:
        return True
    return False


def _parquet_ttl_minutes(cfg: Any = None) -> int:
    raw = None
    if cfg:
        raw = getattr(cfg, "PRODUCTS_PARQUET_TTL_MIN", None)
    if raw is None:
        raw = os.getenv("PRODUCTS_PARQUET_TTL_MIN")
    try:
        return int(raw) if raw is not None else PARQUET_REBUILD_TTL_MIN
    except Exception:
        return PARQUET_REBUILD_TTL_MIN


def _load_products_source_df() -> pd.DataFrame:
    """Load the latest fact snapshot (parquet or live) and standardize columns."""
    t0 = time.perf_counter()
    df: Optional[pd.DataFrame] = None
    source = "none"

    try:
        import data_loader as loader  # type: ignore

        df = loader.load_snapshot()
        source = "snapshot"
    except Exception:
        df = None

    if df is None or df.empty:
        try:
            import data_loader as loader  # type: ignore

            df = loader.get_dataframe()
            source = "direct_sql"
        except Exception:
            df = None

    if df is None or not isinstance(df, pd.DataFrame):
        df = pd.DataFrame()

    try:
        df_std = _standardize_sales_df(df) if not df.empty else _empty_products_frame()
    except Exception:
        logger.exception("products.parquet.source_standardize_failed")
        df_std = _empty_products_frame()

    duration_ms = (time.perf_counter() - t0) * 1000
    try:
        logger.info(
            "products.parquet.source_loaded",
            extra={
                "event": "products_parquet_source_loaded",
                "source": source,
                "rows": len(df_std),
                "duration_ms": round(duration_ms, 2),
            },
        )
    except Exception:
        pass
    return df_std


def _materialize_products_parquet(path_obj: Path, schema_version: Optional[str], auto_flag: bool, ttl_min: int, *, force: bool = False) -> tuple[bool, Optional[int], Optional[str]]:
    """
    Ensure the parquet file exists and is fresh enough.
    Returns (rebuilt, rows, error_message).
    """
    if not auto_flag:
        return False, None, None

    if not force and not _needs_rebuild(path_obj, ttl_min):
        return False, None, None

    lock_path = path_obj.with_suffix(path_obj.suffix + ".lock")
    t0 = time.perf_counter()
    try:
        with _parquet_lock(lock_path, PARQUET_LOCK_TIMEOUT):
            if not force and not _needs_rebuild(path_obj, ttl_min):
                return False, None, None

            frame = _load_products_source_df()
            rows = len(frame)

            try:
                frame.to_parquet(path_obj.as_posix(), index=False)
            except Exception:
                # Retry with an engine hint if available
                engine = "pyarrow" if pa is not None else None
                frame.to_parquet(path_obj.as_posix(), index=False, engine=engine)

            _write_schema_meta(path_obj, schema_version)
            duration_ms = (time.perf_counter() - t0) * 1000
            logger.info(
                "products.parquet.rebuilt",
                extra={
                    "event": "products_parquet_rebuilt",
                    "path": path_obj.as_posix(),
                    "rows": rows,
                    "duration_ms": round(duration_ms, 2),
                },
            )
            return True, rows, None
    except TimeoutError as exc:
        logger.warning(
            "products.parquet.lock_timeout",
            extra={"event": "products_parquet_lock_timeout", "path": path_obj.as_posix(), "timeout": PARQUET_LOCK_TIMEOUT},
        )
        return False, None, str(exc)
    except Exception as exc:  # pragma: no cover - defensive
        logger.exception(
            "products.parquet.rebuild_failed",
            extra={"event": "products_parquet_rebuild_failed", "path": path_obj.as_posix(), "error": str(exc)},
        )
        return False, None, str(exc)


def _read_parquet_columns(parquet_path: Path) -> List[str]:
    if pa is not None and pq is not None:
        # Fast metadata-only path
        meta = pq.read_schema(parquet_path)
        return list(meta.names)

    df_cols = pd.read_parquet(parquet_path, engine="pyarrow" if pa is not None else None, columns=None)
    return list(df_cols.columns)


def _write_schema_meta(parquet_path: Path, schema_version: Optional[str]) -> Optional[str]:
    try:
        meta_path = parquet_path.with_suffix(parquet_path.suffix + ".meta.json")
        meta_path.write_text(
            json.dumps(
                {
                    "schema_version": str(schema_version) if schema_version is not None else None,
                    "required_columns": REQUIRED_PRODUCTS_COLUMNS,
                    "parquet": parquet_path.as_posix(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                },
                indent=2,
                sort_keys=True,
            ),
            encoding="utf-8",
        )
        return meta_path.as_posix()
    except Exception:
        logger.debug(
            "products.parquet.meta_write_failed",
            exc_info=True,
            extra={"event": "products_parquet_meta_write_failed", "path": parquet_path.as_posix()},
        )
        return None


def _create_placeholder_parquet(parquet_path: Path, schema_version: Optional[str]) -> tuple[bool, bool, Optional[str]]:
    created = False
    used_sentinel = False
    meta_path: Optional[str] = None
    try:
        if pa is not None and pq is not None:
            schema_fields = []
            for col_name in REQUIRED_PRODUCTS_COLUMNS:
                if col_name == CAN.date:
                    schema_fields.append(pa.field(col_name, pa.timestamp("ns")))
                elif col_name in {CAN.qty, CAN.weight, CAN.revenue, CAN.discount}:
                    schema_fields.append(pa.field(col_name, pa.float64()))
                else:
                    schema_fields.append(pa.field(col_name, pa.string()))
            table = pa.Table.from_arrays([pa.array([], type=f.type) for f in schema_fields], schema=pa.schema(schema_fields))
            pq.write_table(table, parquet_path)
        else:
            frame = _empty_products_frame()
            frame.to_parquet(parquet_path, index=False)
        created = True
        meta_path = _write_schema_meta(parquet_path, schema_version)
    except Exception as exc:
        logger.exception(
            "products.parquet.placeholder_create_failed",
            extra={"event": "products_parquet_placeholder_failed", "path": parquet_path.as_posix()},
        )
        try:
            sentinel = parquet_path.with_suffix(parquet_path.suffix + ".missing")
            sentinel.write_text(
                (
                    "Products parquet placeholder could not be created. "
                    f"Expected columns: {', '.join(REQUIRED_PRODUCTS_COLUMNS)}\nError: {exc}"
                ),
                encoding="utf-8",
            )
            used_sentinel = True
        except Exception:
            logger.exception(
                "products.parquet.sentinel_failed",
                extra={"event": "products_parquet_sentinel_failed", "path": parquet_path.as_posix()},
            )
    return created, used_sentinel, meta_path


def ensure_products_parquet_available(
    path: Optional[str] = None,
    *,
    auto_create: Optional[bool] = None,
    schema_version: Optional[str] = None,
) -> ProductsParquetStatus:
    resolved_path = resolve_products_parquet_path(path)
    try:
        cfg = current_app.config  # type: ignore[attr-defined]
    except Exception:
        cfg = {}
    auto_flag = _coerce_bool(
        auto_create
        if auto_create is not None
        else (getattr(cfg, "AUTO_CREATE_PRODUCTS_PARQUET", None) if cfg else os.getenv("AUTO_CREATE_PRODUCTS_PARQUET")),
        default=True,
    )
    version = str(
        schema_version
        or getattr(cfg, "PRODUCTS_PARQUET_SCHEMA_VERSION", None)
        or os.getenv("PRODUCTS_PARQUET_SCHEMA_VERSION")
        or "1"
    )
    ttl_min = _parquet_ttl_minutes(cfg)
    status = ProductsParquetStatus(
        path=resolved_path,
        exists=os.path.exists(resolved_path),
        auto_create=auto_flag,
        schema_version=version,
    )

    path_obj = Path(resolved_path)
    try:
        if not path_obj.parent.exists():
            os.makedirs(path_obj.parent, exist_ok=True)
            status.created_dir = True
    except Exception:
        status.warning = f"Unable to create products parquet directory {path_obj.parent}"
        status.error = "parent_dir_unavailable"
        logger.exception(
            "products.parquet.ensure.dir_failed",
            extra={
                "event": "products_parquet_dir_failed",
                "path": resolved_path,
                "auto_create": auto_flag,
            },
        )
        return status

    rebuilt, rows, rebuild_error = _materialize_products_parquet(path_obj, version, auto_flag, ttl_min)
    status.rebuilt = rebuilt
    status.rows = rows
    if rebuild_error and not status.warning:
        status.warning = rebuild_error

    status.exists = path_obj.exists()

    if not path_obj.exists():
        if auto_flag:
            created, used_sentinel, meta_path = _create_placeholder_parquet(path_obj, version)
            status.created_placeholder = created
            status.used_sentinel = used_sentinel
            status.meta_path = meta_path
            status.exists = created
            status.valid_schema = created
            status.warning = (
                f"Products parquet missing; placeholder created at {resolved_path}"
                if created
                else (
                    f"Products parquet missing at {resolved_path}. "
                    "Install a parquet engine (pyarrow/fastparquet) or create the file via ETL."
                )
            )
        else:
            status.exists = False
            status.valid_schema = False
            status.warning = (
                f"Products parquet missing at {resolved_path}. "
                f"Generate it via `python manage.py build-products-parquet --output \"{resolved_path}\"`."
            )
        logger.info(
            "products.parquet.ensure.missing",
            extra={
                "event": "products_parquet_missing",
                "path": resolved_path,
                "created_dir": status.created_dir,
                "created_placeholder": status.created_placeholder,
                "auto_create": auto_flag,
                "warning": status.warning,
            },
        )
        return status

    # If auto-create is disabled and the file is stale, we continue with the existing file but note it.
    if not auto_flag and _needs_rebuild(path_obj, ttl_min):
        status.warning = status.warning or f"Products parquet is older than {ttl_min} minutes; using existing snapshot."

    try:
        cols = _read_parquet_columns(path_obj)
    except Exception as exc:
        status.exists = True
        status.valid_schema = False
        status.error = str(exc)
        logger.exception(
            "products.parquet.ensure.read_failed",
            extra={
                "event": "products_parquet_read_failed",
                "path": resolved_path,
                "auto_create": auto_flag,
            },
        )
        if auto_flag:
            created, used_sentinel, meta_path = _create_placeholder_parquet(path_obj, version)
            status.created_placeholder = created
            status.used_sentinel = used_sentinel
            status.meta_path = meta_path
            status.exists = created
            status.valid_schema = created
            status.warning = (
                f"Products parquet at {resolved_path} was unreadable; placeholder rebuilt."
                if created
                else f"Products parquet at {resolved_path} is unreadable ({exc.__class__.__name__})."
            )
        else:
            status.warning = (
                f"Products parquet at {resolved_path} is unreadable. "
                "Rebuild it via your ETL or `python manage.py build-products-parquet`."
            )
        return status

    missing_cols = [c for c in REQUIRED_PRODUCTS_COLUMNS if c not in cols]
    if missing_cols:
        status.columns_missing = missing_cols
        status.valid_schema = False
        status.warning = (
            "Products parquet is missing required columns: "
            f"{', '.join(missing_cols)}. Expected: {', '.join(REQUIRED_PRODUCTS_COLUMNS)}."
        )
        if auto_flag and len(cols) == 0:
            created, used_sentinel, meta_path = _create_placeholder_parquet(path_obj, version)
            status.created_placeholder = created
            status.used_sentinel = used_sentinel
            status.meta_path = meta_path
            status.exists = created
            status.valid_schema = created
        return status

    meta_path_obj = path_obj.with_suffix(path_obj.suffix + ".meta.json")
    if meta_path_obj.exists():
        status.meta_path = meta_path_obj.as_posix()

    logger.info(
        "products.parquet.ensure",
        extra={
            "event": "products_parquet_ensure",
            "path": resolved_path,
            "created_dir": status.created_dir,
            "created_placeholder": status.created_placeholder,
            "auto_create": auto_flag,
            "valid_schema": status.valid_schema,
            "schema_version": status.schema_version,
            "warning": status.warning,
        },
    )
    return status


def rebuild_products_parquet(path: Optional[str] = None, *, force: bool = False) -> ProductsParquetStatus:
    """
    Explicit rebuild entry-point (used by CLI). Forces a rebuild when force=True.
    """
    resolved = resolve_products_parquet_path(path)
    try:
        cfg = current_app.config  # type: ignore[attr-defined]
    except Exception:
        cfg = {}
    version = str(
        getattr(cfg, "PRODUCTS_PARQUET_SCHEMA_VERSION", None)
        or os.getenv("PRODUCTS_PARQUET_SCHEMA_VERSION")
        or "1"
    )
    target = Path(resolved)
    target.parent.mkdir(parents=True, exist_ok=True)
    _materialize_products_parquet(
        target,
        version,
        True,
        0 if force else _parquet_ttl_minutes(cfg),
        force=force,
    )
    return ensure_products_parquet_available(resolved, auto_create=True, schema_version=version)


class SalesParquetStore:
    def __init__(self, path: str):
        self.path = resolve_products_parquet_path(path)
        self._lock = threading.RLock()
        self._mtime: Optional[float] = None
        self._df: Optional[pd.DataFrame] = None
        self._last_status: Optional[ProductsParquetStatus] = None

    def _current_mtime(self) -> Optional[float]:
        try:
            return os.path.getmtime(self.path)
        except Exception:
            return None

    def get_df(self) -> pd.DataFrame:
        with self._lock:
            status = ensure_products_parquet_available(self.path)
            self.path = status.path
            self._last_status = status
            mtime = self._current_mtime()
            start = time.perf_counter()

            if self._df is not None and self._mtime == mtime and status.available:
                return self._df

            if not status.available:
                self._df = _load_products_source_df()
                self._mtime = mtime
                try:
                    status.warning = status.warning or "products_parquet_unavailable_fallback"
                    logger.info(
                        "products.parquet.fallback",
                        extra={
                            "event": "products_parquet_fallback",
                            "path": self.path,
                            "rows": len(self._df),
                            "duration_ms": round((time.perf_counter() - start) * 1000, 2),
                        },
                    )
                except Exception:
                    pass
                return self._df

            try:
                raw = pd.read_parquet(self.path)  # engine auto (pyarrow/fastparquet)
            except Exception as exc:
                status.error = str(exc)
                status.valid_schema = False
                status.warning = status.warning or (
                    f"Products parquet at {self.path} is unreadable ({exc.__class__.__name__})."
                )
                logger.exception(
                    "products.parquet.read_failed",
                    extra={"event": "products_parquet_read_failed", "path": self.path},
                )
                self._df = _load_products_source_df()
                self._mtime = self._current_mtime()
                return self._df

            df = _standardize_sales_df(raw)
            missing_after_std = [c for c in REQUIRED_PRODUCTS_COLUMNS if c not in df.columns]
            if missing_after_std:
                status.valid_schema = False
                status.columns_missing = missing_after_std
                status.warning = status.warning or (
                    "Products dataset missing required columns after standardization: "
                    f"{', '.join(missing_after_std)}."
                )
                if status.auto_create and raw.shape[1] == 0:
                    created, used_sentinel, meta_path = _create_placeholder_parquet(Path(self.path), status.schema_version)
                    status.created_placeholder = status.created_placeholder or created
                    status.used_sentinel = status.used_sentinel or used_sentinel
                    status.meta_path = status.meta_path or meta_path
                self._df = _empty_products_frame()
                self._mtime = self._current_mtime()
                return self._df

            if CAN.date in df.columns:
                df = df.sort_values(CAN.date, kind="mergesort")  # stable & fast
            self._df = df
            self._mtime = self._current_mtime()
            try:
                logger.info(
                    "products.parquet.loaded",
                    extra={
                        "event": "products_parquet_loaded",
                        "path": self.path,
                        "rows": len(df),
                        "duration_ms": round((time.perf_counter() - start) * 1000, 2),
                    },
                )
            except Exception:
                pass
            return df

    @property
    def last_status(self) -> ProductsParquetStatus:
        if self._last_status is None:
            try:
                cfg = current_app.config  # type: ignore[attr-defined]
            except Exception:
                cfg = {}
            self._last_status = ProductsParquetStatus(
                path=self.path,
                exists=os.path.exists(self.path),
                auto_create=_coerce_bool(
                    getattr(cfg, "AUTO_CREATE_PRODUCTS_PARQUET", None) if cfg else None,
                    default=False,
                ),
                valid_schema=False,
                warning="products_parquet_not_initialized",
            )
        return self._last_status


def _store() -> SalesParquetStore:
    path = resolve_products_parquet_path()
    return SalesParquetStore(path)


# Lazy singleton per worker
_STORE_SINGLETON: Optional[SalesParquetStore] = None


def sales_df(filters: Any = None) -> pd.DataFrame:
    # Prefer the canonical fact store; fallback to legacy parquet if unavailable
    try:
        from flask import current_app  # type: ignore
        cfg = current_app.config if current_app else {}
    except Exception:
        cfg = {}

    override_path = os.getenv("PRODUCTS_SALES_PARQUET") or (
        cfg.get("PRODUCTS_SALES_PARQUET") if isinstance(cfg, dict) else getattr(cfg, "PRODUCTS_SALES_PARQUET", None)
    )
    testing = False
    try:
        testing = cfg.get("TESTING") if isinstance(cfg, dict) else getattr(cfg, "TESTING", False)
    except Exception:
        testing = False
    if testing:
        test_override = os.getenv("PRODUCTS_PARQUET_PATH")
        if test_override:
            override_path = test_override
        elif not os.getenv("PRODUCTS_SALES_PARQUET"):
            cfg_path = cfg.get("PRODUCTS_PARQUET_PATH") if isinstance(cfg, dict) else getattr(cfg, "PRODUCTS_PARQUET_PATH", None)
            cfg_sales = cfg.get("PRODUCTS_SALES_PARQUET") if isinstance(cfg, dict) else getattr(cfg, "PRODUCTS_SALES_PARQUET", None)
            if cfg_path and cfg_path != cfg_sales:
                override_path = cfg_path
    if override_path:
        try:
            df_override = pd.read_parquet(Path(override_path))
            df_norm = normalize_products_df(df_override)
            if cfg.get("TESTING", False) or (df_norm is not None and not df_norm.empty):
                return df_norm
        except Exception:
            logger.exception("products.sales_df.override_failed", extra={"path": override_path})

    if get_fact_context is not None:
        try:
            user_obj = current_user if current_user else None
        except Exception:
            user_obj = None
        try:
            ctx = get_fact_context(user=user_obj, filters=filters)
            df_live = normalize_products_df(ctx.df)
            if df_live is not None and not df_live.empty:
                return df_live
        except Exception:
            logger.debug("products.sales_df.live_failed", exc_info=True)

    try:
        from flask import current_app  # type: ignore
        if current_app and current_app.config.get("TESTING"):
            from app.services import products as svc_products

            return normalize_products_df(svc_products._load_fact())
    except Exception:
        pass

    if fact_store is not None:
        try:
            base = fact_store.get_sales_fact()
            df_norm = normalize_products_df(base)
            if df_norm is not None and not df_norm.empty:
                return df_norm
        except Exception:
            logger.exception("products.sales_df_fact_store_failed")

    global _STORE_SINGLETON
    if _STORE_SINGLETON is None:
        _STORE_SINGLETON = _store()
    return _STORE_SINGLETON.get_df()


def get_products_parquet_status() -> ProductsParquetStatus:
    if fact_store is not None:
        try:
            meta = fact_store.get_meta() if hasattr(fact_store, "get_meta") else {}
        except Exception:
            meta = {}
        path_obj = getattr(fact_store, "FACT_PATH", Path(resolve_products_parquet_path()))
        status = ProductsParquetStatus(
            path=path_obj.as_posix(),
            exists=path_obj.exists(),
            auto_create=True,
            valid_schema=path_obj.exists(),
            rows=meta.get("row_count"),
            schema_version=str(meta.get("source") or "fact"),
        )
        meta_path_obj = getattr(fact_store, "META_PATH", None)
        if meta_path_obj is not None:
            try:
                status.meta_path = meta_path_obj.as_posix()  # type: ignore[attr-defined]
            except Exception:
                status.meta_path = None
        status.warning = meta.get("warning") if not status.exists else meta.get("warning")
        return status

    global _STORE_SINGLETON
    if _STORE_SINGLETON is None:
        _STORE_SINGLETON = _store()
    return _STORE_SINGLETON.last_status


def sales_df_with_status() -> tuple[pd.DataFrame, ProductsParquetStatus]:
    df = sales_df()
    status = get_products_parquet_status()
    return df, status


def _apply_filters_with_recent_fallback(base_df: pd.DataFrame, filters: Dict[str, Any]) -> tuple[pd.DataFrame, Dict[str, Any]]:
    filtered = apply_filters(base_df, filters)
    if filtered.empty and not base_df.empty:
        try:
            date_series = pd.to_datetime(base_df.get(CAN.date), errors="coerce")
            recent_end = date_series.max()
            if pd.notna(recent_end):
                recent_start = (recent_end - pd.DateOffset(months=DEFAULT_MONTHS)).normalize()
                mask = date_series >= recent_start
                filtered = base_df.loc[mask].copy()
                filters = {**filters, "start_date": recent_start.date().isoformat(), "end_date": recent_end.date().isoformat()}
                logger.info(
                    "products.filters.expanded_window",
                    extra={"start": filters.get("start_date"), "end": filters.get("end_date")},
                )
        except Exception:
            logger.debug("products.filters.recent_fallback_failed", exc_info=True)
    return filtered, filters


# -----------------------------
# Filters
# -----------------------------
def _safe_int(v: Any, default: int) -> int:
    try:
        return int(v)
    except Exception:
        return default


def _recent_window(months: int) -> Tuple[pd.Timestamp, pd.Timestamp]:
    months = max(1, int(months))
    now = pd.Timestamp.utcnow().tz_localize(None).normalize()
    end = now
    start = (end - pd.DateOffset(months=months)).normalize().replace(day=1)
    return start, end


def _list_arg(name: str) -> List[str]:
    # supports ?regions=A&regions=B and ?regions=A,B
    vals = request.args.getlist(name)
    out: List[str] = []
    for v in vals:
        if not v:
            continue
        parts = [p.strip() for p in str(v).split(",") if p.strip()]
        out.extend(parts)
    # treat "All" as no filter
    out = [x for x in out if x.lower() != "all"]
    return out


def _to_multidict(source: Any) -> MultiDict:
    md = MultiDict()
    if source is None:
        return md
    if hasattr(source, "lists"):
        try:
            for k, vals in source.lists():
                for v in vals:
                    if v not in (None, ""):
                        md.add(k, v)
            return md
        except Exception:
            md = MultiDict()
    if isinstance(source, dict):
        for k, v in source.items():
            if isinstance(v, (list, tuple, set)):
                for item in v:
                    if item not in (None, ""):
                        md.add(k, item)
            elif v not in (None, ""):
                md.add(k, v)
    return md


def _to_ts(val: Any) -> Optional[pd.Timestamp]:
    if val is None:
        return None
    try:
        ts = pd.to_datetime(val, errors="coerce")
    except Exception:
        return None
    if ts is None or pd.isna(ts):
        return None
    try:
        return pd.Timestamp(ts).tz_localize(None).normalize()
    except Exception:
        return pd.Timestamp(ts)


def _ts_to_datestr(ts: Any) -> Optional[str]:
    if ts is None:
        return None
    try:
        return pd.Timestamp(ts).date().isoformat()
    except Exception:
        try:
            return ts.date().isoformat()  # type: ignore[attr-defined]
        except Exception:
            return str(ts)


def _make_filter_params(
    start: Optional[pd.Timestamp],
    end: Optional[pd.Timestamp],
    regions: Iterable[str] = (),
    methods: Iterable[str] = (),
    customers: Iterable[str] = (),
    suppliers: Iterable[str] = (),
    products: Iterable[str] = (),
    sales_reps: Iterable[str] = (),
    protein_min: Any = None,
    protein_max: Any = None,
    protein_name_like: Any = None,
    complete_months_only: bool = True,
):
    if filter_svc and hasattr(filter_svc, "FilterParams"):
        try:
            return filter_svc.FilterParams(
                start=start,
                end=end,
                regions=tuple(regions or ()),
                methods=tuple(methods or ()),
                customers=tuple(customers or ()),
                suppliers=tuple(suppliers or ()),
                products=tuple(products or ()),
                sales_reps=tuple(sales_reps or ()),
                protein_min=protein_min,
                protein_max=protein_max,
                protein_name_like=protein_name_like,
                complete_months_only=complete_months_only,
            )
        except Exception:
            pass
    return type("FilterParams", (), {
        "start": start,
        "end": end,
        "regions": tuple(regions or ()),
        "methods": tuple(methods or ()),
        "customers": tuple(customers or ()),
        "suppliers": tuple(suppliers or ()),
        "products": tuple(products or ()),
        "sales_reps": tuple(sales_reps or ()),
        "protein_min": protein_min,
        "protein_max": protein_max,
        "protein_name_like": protein_name_like,
        "complete_months_only": complete_months_only,
    })()


def parse_filter_params(params: Any, fallback_months: int = DEFAULT_MONTHS):
    """Wrapper around shared filter parser with safe defaults (test-friendly)."""
    parsed = None
    explicit_date = False
    try:
        explicit_date = any(
            k in params
            for k in (
                "start",
                "start_date",
                "startDate",
                "end",
                "end_date",
                "endDate",
                "date_preset",
                "preset",
                "range_preset",
            )
        )
    except Exception:
        explicit_date = False
    if filter_svc and hasattr(filter_svc, "parse_filters"):
        try:
            parsed = filter_svc.parse_filters(params)
        except Exception:
            parsed = None
    if parsed is None:
        start, end = _recent_window(fallback_months)
        return _make_filter_params(start=start, end=end)

    start = _to_ts(getattr(parsed, "start", None))
    end = _to_ts(getattr(parsed, "end", None))
    preset = str(getattr(parsed, "preset", None) or "").strip().lower()
    explicit_all_time = preset in {"all", "__all__", "*"}
    if not explicit_date:
        start, end = _recent_window(fallback_months)
    elif start is None and end is None and not explicit_all_time:
        start, end = _recent_window(fallback_months)

    return _make_filter_params(
        start=start,
        end=end,
        regions=getattr(parsed, "regions", ()),
        methods=getattr(parsed, "methods", ()),
        customers=getattr(parsed, "customers", ()),
        suppliers=getattr(parsed, "suppliers", ()),
        products=getattr(parsed, "products", ()),
        sales_reps=getattr(parsed, "sales_reps", ()),
        protein_min=getattr(parsed, "protein_min", None),
        protein_max=getattr(parsed, "protein_max", None),
        protein_name_like=getattr(parsed, "protein_name_like", None),
        complete_months_only=bool(getattr(parsed, "complete_months_only", True)),
    )


def parse_global_filters(arg: Any = None, fallback_months: int = DEFAULT_MONTHS) -> Dict[str, Any]:
    """
    Unified global filter parser used across products endpoints.
    Returns normalized dict with canonical keys.
    """
    if isinstance(arg, (int, float)):
        fallback_months = int(arg)
        arg = None

    try:
        v2_enabled = bool(current_app.config.get("FILTERS_CANONICAL_V2", False))
    except Exception:
        v2_enabled = False
    if v2_enabled and filter_svc and hasattr(filter_svc, "resolve_filters"):
        if arg is None:
            source = request.args
        elif hasattr(arg, "args"):
            source = arg.args
        elif isinstance(arg, dict):
            source = arg
        else:
            source = {}
        try:
            params, _meta = filter_svc.resolve_filters(
                request,
                current_user,
                session_obj=session,
                source=source,
                sticky_enabled=bool(current_app.config.get("STICKY_FILTERS", True)),
                update_session=False,
            )
            start = _to_ts(getattr(params, "start", None))
            end = _to_ts(getattr(params, "end", None))
            return {
                "start_date": _ts_to_datestr(start),
                "end_date": _ts_to_datestr(end),
                "date_preset": getattr(params, "preset", None),
                "regions": list(getattr(params, "regions", ()) or []),
                "customers": list(getattr(params, "customers", ()) or []),
                "suppliers": list(getattr(params, "suppliers", ()) or []),
                "products": list(getattr(params, "products", ()) or []),
                "sales_reps": list(getattr(params, "sales_reps", ()) or []),
                "shipping_methods": list(getattr(params, "methods", ()) or []),
                "methods": list(getattr(params, "methods", ()) or []),
                "include_current_month": _coerce_bool(
                    (source.get("include_current") if hasattr(source, "get") else None)
                    or (source.get("include_current_month") if hasattr(source, "get") else None),
                    default=False,
                ),
                "forecast": _coerce_bool(source.get("forecast") if hasattr(source, "get") else None, default=False),
                "complete_months_only": bool(getattr(params, "complete_months_only", True)),
                "statuses": list(getattr(params, "statuses", ()) or []),
                "protein_min": getattr(params, "protein_min", None),
                "protein_max": getattr(params, "protein_max", None),
                "protein_name_like": getattr(params, "protein_name_like", None),
            }
        except Exception:
            logger.exception("products.resolve_filters_v2_failed")

    stored: Dict[str, Any] = {}
    for key in ("filters", "global_filters", "products_filters"):
        val = session.get(key)
        if isinstance(val, dict):
            stored.update(val)

    if arg is None:
        source = request.args
    elif hasattr(arg, "args"):
        source = arg.args
    elif isinstance(arg, dict):
        source = arg
    else:
        source = {}

    merged = _to_multidict(stored)
    for k, vals in _to_multidict(source).lists():
        merged.setlist(k, list(vals))

    try:
        params = parse_filter_params(merged, fallback_months=fallback_months)
    except TypeError:
        params = parse_filter_params(merged)
    start = _to_ts(getattr(params, "start", None))
    end = _to_ts(getattr(params, "end", None))
    date_preset = None
    try:
        date_preset = (
            source.get("date_preset")
            or source.get("preset")
            or stored.get("date_preset")
            or stored.get("preset")
        )
    except Exception:
        date_preset = None
    preset_token = str(date_preset).strip().lower() if date_preset else None
    explicit_all_time = preset_token in {"all", "__all__", "*"}
    if start is None and end is None and not explicit_all_time:
        start, end = _recent_window(fallback_months)

    include_current = _coerce_bool(
        (source.get("include_current") if hasattr(source, "get") else None)
        or (source.get("include_current_month") if hasattr(source, "get") else None)
        or stored.get("include_current_month"),
        default=False,
    )
    forecast_flag = _coerce_bool(
        (source.get("forecast") if hasattr(source, "get") else None)
        or stored.get("forecast"),
        default=False,
    )

    out = {
        "start_date": _ts_to_datestr(start),
        "end_date": _ts_to_datestr(end),
        "date_preset": str(date_preset).strip() if date_preset else None,
        "regions": list(getattr(params, "regions", ()) or []),
        "customers": list(getattr(params, "customers", ()) or []),
        "suppliers": list(getattr(params, "suppliers", ()) or []),
        "products": list(getattr(params, "products", ()) or []),
        "sales_reps": list(getattr(params, "sales_reps", ()) or []),
        "shipping_methods": list(getattr(params, "methods", ()) or []),
        "methods": list(getattr(params, "methods", ()) or []),  # compatibility
        "include_current_month": include_current,
        "forecast": forecast_flag,
        "complete_months_only": bool(getattr(params, "complete_months_only", True)),
        "statuses": list(getattr(params, "statuses", ()) or []),
        "protein_min": getattr(params, "protein_min", None),
        "protein_max": getattr(params, "protein_max", None),
        "protein_name_like": getattr(params, "protein_name_like", None),
    }
    return out


def parse_filters(arg: Any = None, fallback_months: int = DEFAULT_MONTHS) -> Dict[str, Any]:
    """
    Parse filters from a request-like source or dict (backward compatible).
    """
    gf = parse_global_filters(arg, fallback_months=fallback_months)
    return {
        "start": gf.get("start_date"),
        "end": gf.get("end_date"),
        "start_date": gf.get("start_date"),
        "end_date": gf.get("end_date"),
        "date_preset": gf.get("date_preset"),
        "regions": gf.get("regions") or [],
        "customers": gf.get("customers") or [],
        "suppliers": gf.get("suppliers") or [],
        "products": gf.get("products") or [],
        "methods": gf.get("methods") or [],
        "shipping_methods": gf.get("shipping_methods") or [],
        "sales_reps": gf.get("sales_reps") or [],
        "include_current_month": gf.get("include_current_month", False),
        "forecast": gf.get("forecast", False),
        "complete_months_only": gf.get("complete_months_only", True),
        "statuses": gf.get("statuses") or [],
        "protein_min": gf.get("protein_min"),
        "protein_max": gf.get("protein_max"),
        "protein_name_like": gf.get("protein_name_like"),
    }


def build_querystring(filters: Dict[str, Any], include_prefix: bool = True) -> str:
    """Build a stable querystring for drilldown links/export preserving current filters."""
    if not filters:
        return ""
    params: List[Tuple[str, Any]] = []

    def _append_list(key: str, out_key: Optional[str] = None):
        vals = filters.get(key) or []
        if isinstance(vals, (list, tuple, set)):
            for v in vals:
                params.append((out_key or key, v))

    for k in ("start_date", "end_date", "date_preset"):
        v = filters.get(k) or filters.get(k.replace("_date", ""))
        if v:
            params.append((k, v))

    _append_list("regions")
    _append_list("customers")
    _append_list("suppliers")
    _append_list("products")
    _append_list("shipping_methods", "methods")
    _append_list("statuses")
    _append_list("sales_reps")

    for key in ("protein_min", "protein_max", "protein_name_like"):
        value = filters.get(key)
        if value not in (None, ""):
            params.append((key, value))

    if _coerce_bool(filters.get("include_current_month"), default=False):
        params.append(("include_current", "1"))
    if _coerce_bool(filters.get("forecast"), default=False):
        params.append(("forecast", "1"))
    if filters.get("complete_months_only") is not None:
        params.append(("complete_months_only", "1" if _coerce_bool(filters.get("complete_months_only"), default=False) else "0"))

    qs = urlencode(params, doseq=True)
    if not qs:
        return ""
    return f"?{qs}" if include_prefix else qs


def apply_filters(df: pd.DataFrame, filters: Dict[str, Any]) -> pd.DataFrame:
    if df.empty:
        return df

    start_raw = filters.get("start") or filters.get("start_date")
    end_raw = filters.get("end") or filters.get("end_date")
    start = pd.to_datetime(start_raw, errors="coerce") if start_raw is not None else pd.NaT
    end = pd.to_datetime(end_raw, errors="coerce") if end_raw is not None else pd.NaT
    try:
        start = pd.Timestamp(start).tz_localize(None) if pd.notna(start) else pd.NaT
        end = pd.Timestamp(end).tz_localize(None) if pd.notna(end) else pd.NaT
    except Exception:
        start, end = pd.NaT, pd.NaT

    dates = pd.to_datetime(df[CAN.date], errors="coerce")
    try:
        dates = dates.dt.tz_localize(None)
    except Exception:
        dates = dates

    mask = pd.Series(True, index=df.index, dtype=bool)
    if pd.notna(start):
        mask &= dates >= start
    if pd.notna(end):
        end_value = pd.Timestamp(end)
        inclusive_end = end_value
        use_strict_lt = False
        try:
            if (
                end_value.hour == 0
                and end_value.minute == 0
                and end_value.second == 0
                and end_value.microsecond == 0
                and getattr(end_value, "nanosecond", 0) == 0
            ):
                inclusive_end = end_value + pd.Timedelta(days=1)
                use_strict_lt = True
        except Exception:
            inclusive_end = end_value
        mask &= dates < inclusive_end if use_strict_lt else dates <= inclusive_end

    include_current = _coerce_bool(filters.get("include_current_month"), default=False)
    if not include_current:
        try:
            current_month_start = pd.Timestamp.utcnow().tz_localize(None).normalize().replace(day=1)
            mask &= dates < current_month_start
        except Exception:
            pass

    def _entity_mask(columns: Iterable[str], values: Iterable[Any]) -> pd.Series:
        tokens = {str(x).strip().lower() for x in (values or []) if str(x).strip()}
        if not tokens:
            return pd.Series(True, index=df.index, dtype=bool)
        entity_mask = pd.Series(False, index=df.index, dtype=bool)
        for col in columns:
            if col in df.columns:
                entity_mask |= df[col].astype("string").str.strip().str.lower().isin(tokens)
        return entity_mask

    regions = filters.get("regions") or []
    if regions:
        mask &= _entity_mask((CAN.region, CAN.region_id), regions)

    suppliers = filters.get("suppliers") or []
    if suppliers:
        mask &= _entity_mask((CAN.supplier, CAN.supplier_id), suppliers)

    customers = filters.get("customers") or []
    if customers:
        mask &= _entity_mask((CAN.customer_id, CAN.customer_name), customers)

    products = filters.get("products") or []
    if products:
        mask &= _entity_mask((CAN.product_id, CAN.product_name, CAN.sku), products)

    shipping_methods = filters.get("shipping_methods") or filters.get("methods") or []
    if shipping_methods:
        ship_col = None
        for cand in ("Method", "ShippingMethodName", "shipping_method", "ShippingMethodLabel"):
            if cand in df.columns:
                ship_col = cand
                break
        if ship_col:
            mask &= df[ship_col].astype("string").isin([str(x) for x in shipping_methods])

    statuses = [str(x).strip().lower() for x in (filters.get("statuses") or []) if str(x).strip()]
    if statuses:
        for cand in ("OrderStatus", "order_status", "Status"):
            if cand in df.columns:
                mask &= df[cand].astype("string").str.strip().str.lower().isin(statuses)
                break

    sales_reps = [str(x).strip().lower() for x in (filters.get("sales_reps") or []) if str(x).strip()]
    if sales_reps:
        rep_mask = pd.Series(False, index=df.index)
        for cand in ("SalesRepId", "SalesRepName", "ERPUserId", "SalesRepERPUserId"):
            if cand in df.columns:
                rep_mask |= df[cand].astype("string").str.strip().str.lower().isin(sales_reps)
        mask &= rep_mask

    protein_min = filters.get("protein_min")
    protein_max = filters.get("protein_max")
    if "Protein" in df.columns and (protein_min is not None or protein_max is not None):
        prot = pd.to_numeric(df["Protein"], errors="coerce")
        if protein_min is not None:
            try:
                mask &= prot >= float(protein_min)
            except Exception:
                pass
        if protein_max is not None:
            try:
                mask &= prot <= float(protein_max)
            except Exception:
                pass

    protein_name_like = str(filters.get("protein_name_like") or "").strip()
    if protein_name_like and CAN.product_name in df.columns:
        mask &= df[CAN.product_name].astype("string").str.contains(protein_name_like, case=False, na=False)

    return df.loc[mask]


# -----------------------------
# Core metrics helpers/builders (fast)
# -----------------------------
def _qty_basis_series(df: pd.DataFrame) -> pd.Series:
    col = CAN.qty_basis if CAN.qty_basis in df.columns else CAN.qty
    if col in df.columns:
        return pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    return pd.Series(0.0, index=df.index if not df.empty else pd.Index([]), dtype="float64")


def _weight_series(df: pd.DataFrame) -> pd.Series:
    col = CAN.weight if CAN.weight in df.columns else "weight"
    if col in df.columns:
        return pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    return pd.Series(0.0, index=df.index if not df.empty else pd.Index([]), dtype="float64")


def _qty_label(df: pd.DataFrame) -> str:
    if CAN.qty_basis_label in df.columns:
        try:
            label = df[CAN.qty_basis_label].dropna().astype("string").str.strip()
            if not label.empty and label.iloc[0]:
                return str(label.iloc[0])
        except Exception:
            pass
    has_weight = bool(_weight_series(df).abs().sum() > 0)
    return "Weight (lb)" if has_weight else "Quantity"


def _has_cost_data(df: pd.DataFrame) -> bool:
    if CAN.cost not in df.columns:
        return False
    try:
        cost_series = pd.to_numeric(df[CAN.cost], errors="coerce")
        return bool((cost_series > 0).any())
    except Exception:
        return False


def _strip_costs_for_view(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    safe = df.copy()
    for col in (CAN.cost, CAN.margin, CAN.margin_pct, CAN.unit_cost):
        if col in safe.columns:
            safe[col] = np.nan
    return safe


def _strip_costs_from_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Remove cost-dependent fields from a payload in-place."""
    if not payload:
        return payload

    kpis = payload.get("kpis")
    if isinstance(kpis, dict):
        for key in ("avg_margin", "avg_margin_pct", "margin_total"):
            if key in kpis:
                kpis[key] = None

    for collection_key in ("top_products", "top_movers"):
        rows = payload.get(collection_key)
        if isinstance(rows, list):
            for row in rows:
                if not isinstance(row, dict):
                    continue
                for key in (
                    "cost",
                    "margin",
                    "margin_pct",
                    "profit",
                    "unit_cost",
                    "target_price",
                    "target_price_27",
                    "target_price_21",
                    "target_unit_price",
                    "uplift_pct",
                    "recommendation_label",
                ):
                    if key in row:
                        row[key] = None

    trend = payload.get("trend")
    if isinstance(trend, dict) and "margin" in trend:
        trend["margin"] = []

    return payload


def _unit_price(df: pd.DataFrame) -> pd.Series:
    qty = _qty_basis_series(df).replace(0, np.nan)
    rev = pd.to_numeric(df[CAN.revenue], errors="coerce")
    return (rev / qty).replace([np.inf, -np.inf], np.nan)


def _monthly_trend(df: pd.DataFrame) -> Dict[str, List[Any]]:
    if df.empty:
        return {"labels": [], "revenue": [], "values": [], "qty": [], "weight": [], "margin": [], "asp": []}

    tmp = df[[CAN.date, CAN.revenue]].copy(deep=False)
    tmp["qty_basis"] = _qty_basis_series(df)
    tmp["weight"] = _weight_series(df)
    if CAN.margin in df.columns:
        tmp["margin"] = pd.to_numeric(df[CAN.margin], errors="coerce")
    tmp["month"] = tmp[CAN.date].dt.to_period("M").dt.to_timestamp()
    g = tmp.groupby("month", observed=True)

    rev = g[CAN.revenue].sum().astype(float)
    qty = g["qty_basis"].sum().astype(float)
    w = g["weight"].sum().astype(float)
    margin = g["margin"].sum(min_count=1).astype(float) if "margin" in tmp.columns else pd.Series(dtype=float)

    labels = [d.strftime("%Y-%m") for d in rev.index.to_pydatetime()]
    revenue = [float(round(x, 2)) for x in rev.values]
    qty_v = [float(round(x, 2)) for x in qty.values]
    w_v = [float(round(x, 2)) for x in w.values]
    margin_v = [float(round(x, 2)) if not pd.isna(x) else None for x in margin.values] if not margin.empty else []
    asp = [(float(round(r / q, 2)) if q else None) for r, q in zip(revenue, qty_v)]

    return {
        "labels": labels,
        "revenue": revenue,
        "values": revenue,  # backward compatibility
        "qty": qty_v,
        "weight": w_v,
        "margin": margin_v,
        "asp": asp,
    }


def _price_dist(df: pd.DataFrame) -> Dict[str, Any]:
    up = _unit_price(df).dropna()
    if up.empty:
        return {"prices": [], "p10": None, "p50": None, "p90": None}

    sample = up.sample(min(len(up), HISTOGRAM_SAMPLE_MAX), random_state=42).round(4).tolist()
    return {
        "prices": sample,
        "p10": float(round(up.quantile(0.10), 2)),
        "p50": float(round(up.quantile(0.50), 2)),
        "p90": float(round(up.quantile(0.90), 2)),
    }


def _top_products(df: pd.DataFrame, limit: int) -> List[Dict[str, Any]]:
    if df.empty:
        return []
    limit = max(1, min(int(limit), TOP_N_MAX))

    working = df.copy(deep=False)
    working["_qty_basis"] = _qty_basis_series(df)
    working["_weight"] = _weight_series(df)

    g = working.groupby([CAN.product_id, CAN.product_name], observed=True, sort=False)
    rev = g[CAN.revenue].sum().astype(float)
    qty_basis = g["_qty_basis"].sum().astype(float)
    weight_sum = g["_weight"].sum().astype(float)
    cost = g[CAN.cost].sum(min_count=1).astype(float) if CAN.cost in df.columns else None
    first_sold = g[CAN.date].min()
    last_sold = g[CAN.date].max()

    out = pd.DataFrame(
        {
            "product_id": rev.index.get_level_values(0).astype(str),
            "desc": rev.index.get_level_values(1).astype(str),
            "revenue": rev.values,
            "qty": qty_basis.values,
            "weight": weight_sum.values,
            "first_sold": first_sold.values,
            "last_sold": last_sold.values,
        }
    )
    out["avg_price"] = np.where(out["qty"] > 0, out["revenue"] / out["qty"], np.nan)
    out["current_unit_price"] = out["avg_price"]
    if cost is not None:
        out["cost"] = cost.values
        out["cost_available"] = (out["cost"] > 0) & out["cost"].notna()
        out["margin_valid"] = out["cost_available"] & (out["revenue"] > 0)
        out["cost"] = out["cost"].where(out["cost_available"], np.nan)
        out["profit"] = np.where(out["cost_available"], out["revenue"] - out["cost"], np.nan)
        out["margin"] = out["profit"]
        if au is not None and hasattr(au, "safe_margin_pct"):
            out["margin_pct"] = au.safe_margin_pct(out["revenue"], out["cost"])
        else:
            out["margin_pct"] = np.where(out["margin_valid"], out["profit"] / out["revenue"] * 100.0, np.nan)
            out["margin_pct"] = np.clip(out["margin_pct"], -200.0, 200.0)
        out["unit_cost"] = np.where(out["qty"] > 0, out["cost"] / out["qty"], np.nan)
        out["target_price_27"] = np.where(out["unit_cost"].notna(), out["unit_cost"] / (1 - 0.27), np.nan)
        out["target_price_21"] = np.where(out["unit_cost"].notna(), out["unit_cost"] / (1 - 0.21), np.nan)
        out["target_unit_price"] = out["target_price_27"]
        if au is not None and hasattr(au, "safe_uplift_pct"):
            out["uplift_pct"] = au.safe_uplift_pct(out["current_unit_price"], out["target_price_27"])
        else:
            out["uplift_pct"] = np.where(
                (out["current_unit_price"] > 0) & out["target_price_27"].notna(),
                (out["target_price_27"] - out["current_unit_price"]) / out["current_unit_price"] * 100.0,
                np.nan,
            )
            out["uplift_pct"] = np.clip(out["uplift_pct"], -100.0, 200.0)
    else:
        out["cost"] = np.nan
        out["margin"] = np.nan
        out["margin_pct"] = np.nan
        out["profit"] = np.nan
        out["unit_cost"] = np.nan
        out["target_price_27"] = np.nan
        out["target_price_21"] = np.nan
        out["target_unit_price"] = np.nan
        out["uplift_pct"] = np.nan
        out["cost_available"] = False
        out["margin_valid"] = False

    out["recommendation_label"] = [
        _recommendation_label(bool(mv), None if pd.isna(up) else float(up))
        for mv, up in zip(out.get("margin_valid", pd.Series(False, index=out.index)), out.get("uplift_pct", pd.Series(np.nan, index=out.index)))
    ]
    out = out.sort_values("revenue", ascending=False).head(limit)

    total_rev = float(out["revenue"].sum()) if not out.empty else 0.0
    total_qty = float(out["qty"].sum()) if not out.empty else 0.0
    out["revenue_share"] = (out["revenue"] / total_rev) * 100.0 if total_rev else np.nan
    out["qty_share"] = (out["qty"] / total_qty) * 100.0 if total_qty else np.nan

    rows: List[Dict[str, Any]] = []
    for r in out.itertuples(index=False):
        first_sold = pd.to_datetime(r.first_sold) if hasattr(r, "first_sold") else None
        last_sold = pd.to_datetime(r.last_sold) if hasattr(r, "last_sold") else None
        rows.append({
            "product_id": r.product_id,
            "sku": r.product_id,   # keep old UI expectations
            "desc": r.desc,
            "revenue": float(round(r.revenue, 2)),
            "qty": float(round(r.qty, 2)),
            "qty_basis": float(round(r.qty, 2)),
            "qty_share": float(round(r.qty_share, 4)) if hasattr(r, "qty_share") and not pd.isna(r.qty_share) else None,
            "avg_price": float(round(r.avg_price, 2)) if not pd.isna(r.avg_price) else None,
            "unit_price": float(round(r.avg_price, 2)) if not pd.isna(r.avg_price) else None,
            "current_unit_price": float(round(r.current_unit_price, 2)) if hasattr(r, "current_unit_price") and not pd.isna(r.current_unit_price) else None,
            "revenue_share": float(round(r.revenue_share, 4)) if not pd.isna(r.revenue_share) else None,
            "margin": float(round(r.margin, 2)) if not pd.isna(r.margin) else None,
            "margin_pct": float(round(r.margin_pct, 2)) if not pd.isna(r.margin_pct) else None,
            "profit": float(round(r.profit, 2)) if hasattr(r, "profit") and not pd.isna(r.profit) else None,
            "cost": float(round(r.cost, 2)) if not pd.isna(r.cost) else None,
            "unit_cost": float(round(r.unit_cost, 4)) if hasattr(r, "unit_cost") and not pd.isna(r.unit_cost) else None,
            "target_price": float(round(r.target_price_27, 2)) if hasattr(r, "target_price_27") and not pd.isna(r.target_price_27) else None,
            "target_price_27": float(round(r.target_price_27, 2)) if hasattr(r, "target_price_27") and not pd.isna(r.target_price_27) else None,
            "target_price_21": float(round(r.target_price_21, 2)) if hasattr(r, "target_price_21") and not pd.isna(r.target_price_21) else None,
            "target_unit_price": float(round(r.target_unit_price, 2)) if hasattr(r, "target_unit_price") and not pd.isna(r.target_unit_price) else None,
            "uplift_pct": float(round(r.uplift_pct, 2)) if hasattr(r, "uplift_pct") and not pd.isna(r.uplift_pct) else None,
            "recommendation_label": r.recommendation_label if hasattr(r, "recommendation_label") else None,
            "cost_available": bool(getattr(r, "cost_available", False)),
            "margin_valid": bool(getattr(r, "margin_valid", False)),
            "first_sold": first_sold.strftime("%Y-%m-%d") if pd.notna(first_sold) else None,
            "last_sold": last_sold.strftime("%Y-%m-%d") if pd.notna(last_sold) else None,
        })
    return rows


def _top_movers(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """Delta of last month vs previous month (revenue)."""
    if df.empty:
        return []
    tmp = df[[CAN.product_id, CAN.product_name, CAN.date, CAN.revenue]].copy(deep=False)
    tmp["month"] = tmp[CAN.date].dt.to_period("M")
    m = tmp.groupby([CAN.product_id, CAN.product_name, "month"], observed=True)[CAN.revenue].sum().reset_index()
    if m.empty:
        return []

    # last two months per product
    m = m.sort_values("month")
    last_two = m.groupby([CAN.product_id, CAN.product_name], observed=True).tail(2)
    movers = []
    for (pid, pname), grp in last_two.groupby([CAN.product_id, CAN.product_name], observed=True):
        if len(grp) != 2:
            continue
        a, b = grp.iloc[0], grp.iloc[1]
        prev = float(a[CAN.revenue])
        curr = float(b[CAN.revenue])
        delta = curr - prev
        pct = (delta / prev * 100.0) if prev else None
        movers.append({
            "product_id": str(pid),
            "sku": str(pid),
            "desc": str(pname),
            "product": str(pname),
            "delta_revenue": float(round(delta, 2)),
            "growth_pct": float(round(pct, 2)) if pct is not None else None,
            "current_revenue": float(round(curr, 2)),
            "previous_revenue": float(round(prev, 2)),
        })

    movers.sort(key=lambda x: x.get("delta_revenue") or 0.0, reverse=True)
    return movers[:TOP_N_DEFAULT]


def _build_also_bought(df: pd.DataFrame, product_id: str, limit: int = 20) -> Dict[str, Any]:
    """Compute co-purchased products within the filtered frame."""
    if df is None or df.empty or CAN.product_id not in df.columns:
        return {"rows": [], "approximate": False, "base_orders": 0, "total_orders": 0}

    pid = str(product_id)
    working = df.copy()
    working[CAN.product_id] = working[CAN.product_id].astype("string")

    order_col = CAN.order_id if CAN.order_id in working.columns else None
    approximate = False

    if not order_col or working[order_col].isna().all():
        if CAN.customer_id in working.columns and CAN.date in working.columns:
            approximate = True
            tmp_dates = pd.to_datetime(working[CAN.date], errors="coerce")
            working["_order_key"] = working[CAN.customer_id].astype("string").str.strip() + "|" + tmp_dates.dt.to_period("W").astype("string")
            order_col = "_order_key"
        else:
            return {"rows": [], "approximate": True, "base_orders": 0, "total_orders": 0}
    else:
        working[order_col] = working[order_col].astype("string")

    base_orders_raw = working.loc[working[CAN.product_id] == pid, order_col].dropna().astype("string").tolist()
    base_orders = list(dict.fromkeys(base_orders_raw))
    if not base_orders:
        return {"rows": [], "approximate": approximate, "base_orders": 0, "total_orders": int(working[order_col].dropna().nunique())}

    if len(base_orders) > CO_PURCHASE_SAMPLE_ORDERS:
        base_orders = list(pd.Series(base_orders).sample(CO_PURCHASE_SAMPLE_ORDERS, random_state=42))
    target_orders = set(base_orders)

    total_orders = int(working[order_col].dropna().nunique())
    base_orders_count = len(target_orders)

    scoped = working[working[order_col].isin(target_orders)].copy()
    scoped["_qty_basis"] = _qty_basis_series(scoped)
    co_products = scoped[scoped[CAN.product_id] != pid]

    if co_products.empty:
        return {"rows": [], "approximate": approximate, "base_orders": base_orders_count, "total_orders": total_orders}

    co_orders = co_products.groupby(CAN.product_id, observed=True)[order_col].nunique()
    co_revenue = co_products.groupby(CAN.product_id, observed=True)[CAN.revenue].sum().astype(float)
    co_qty = co_products.groupby(CAN.product_id, observed=True)["_qty_basis"].sum().astype(float)
    last_seen = co_products.groupby(CAN.product_id, observed=True)[CAN.date].max()

    overall_orders = working.groupby(CAN.product_id, observed=True)[order_col].nunique()

    name_map = working.groupby(CAN.product_id, observed=True)[CAN.product_name].agg(lambda s: s.dropna().iloc[0] if not s.dropna().empty else None)
    sku_map = working.groupby(CAN.product_id, observed=True)[CAN.sku].agg(lambda s: s.dropna().iloc[0] if not s.dropna().empty else None)

    rows: List[Dict[str, Any]] = []
    for other_pid, co_order_count in co_orders.items():
        try:
            base_freq = co_order_count / base_orders_count if base_orders_count else 0.0
            support = co_order_count / total_orders if total_orders else 0.0
            other_freq = overall_orders.get(other_pid, 0) / total_orders if total_orders else 0.0
            lift = (base_freq / other_freq) if other_freq else None
            rows.append({
                "product_id": str(other_pid),
                "product_name": name_map.get(other_pid) or str(other_pid),
                "sku": sku_map.get(other_pid) or str(other_pid),
                "co_orders": int(co_order_count),
                "support": float(round(support * 100.0, 2)) if support else 0.0,
                "confidence": float(round(base_freq * 100.0, 2)) if base_freq else 0.0,
                "lift": float(round(lift, 3)) if lift is not None else None,
                "paired_revenue": float(round(co_revenue.get(other_pid, 0.0), 2)),
                "paired_qty": float(round(co_qty.get(other_pid, 0.0), 2)),
                "last_seen": pd.to_datetime(last_seen.get(other_pid)).strftime("%Y-%m-%d") if pd.notna(last_seen.get(other_pid)) else None,
            })
        except Exception:
            continue

    rows.sort(key=lambda r: (r.get("confidence") or 0.0, r.get("lift") or 0.0), reverse=True)
    rows = rows[: max(1, min(limit, TOP_N_MAX))]
    return {"rows": rows, "approximate": approximate, "base_orders": base_orders_count, "total_orders": total_orders}


def _kpis(df: pd.DataFrame) -> Dict[str, Any]:
    if df.empty:
        return {
            "total_revenue": 0.0,
            "total_qty": 0.0,
            "total_weight": 0.0,
            "unique_products": 0,
            "avg_margin": None,
            "avg_unit_price": None,
            "median_unit_price": None,
            "revenue_per_product": None,
            "revenue_per_customer": None,
            "customer_count": None,
            "avg_qty_per_product": None,
            "unit_price_p50": None,
            "unit_price_p10": None,
            "unit_price_p90": None,
            "avg_margin": None,
            "avg_margin_pct": None,
            "margin_total": None,
        }

    qty_series = _qty_basis_series(df)
    weight_series = _weight_series(df)

    total_rev = float(pd.to_numeric(df[CAN.revenue], errors="coerce").sum())
    total_qty = float(round(qty_series.sum(), 2))
    total_w = float(round(weight_series.sum(), 2)) if not weight_series.empty else 0.0
    uniq_prod = int(df[CAN.product_id].nunique())
    cust_count = int(df[CAN.customer_id].nunique()) if CAN.customer_id in df.columns else 0
    total_cost = None
    if CAN.cost in df.columns:
        cost_series = pd.to_numeric(df[CAN.cost], errors="coerce")
        total_cost = cost_series.sum(min_count=1)
        if pd.isna(total_cost) or total_cost <= 0:
            total_cost = None

    up = _unit_price(df).dropna()
    avg_up = float(up.mean()) if not up.empty else None
    med_up = float(up.median()) if not up.empty else None
    qty_label = _qty_label(df)

    avg_margin_pct = None
    margin_total = None
    if total_rev and total_cost is not None:
        if au is not None and hasattr(au, "safe_profit"):
            margin_total = au.safe_profit(total_rev, total_cost)
        else:
            margin_total = total_rev - total_cost if total_cost > 0 else None
        if margin_total is not None:
            margin_total = float(round(margin_total, 2))
            if au is not None and hasattr(au, "safe_margin_pct"):
                avg_margin_pct = au.safe_margin_pct(total_rev, total_cost)
                if avg_margin_pct is not None:
                    avg_margin_pct = float(round(avg_margin_pct, 2))
            else:
                avg_margin_pct = float(round((margin_total / total_rev) * 100.0, 2)) if total_rev else None

    revenue_per_product = (total_rev / uniq_prod) if uniq_prod else None
    revenue_per_customer = (total_rev / cust_count) if cust_count else None
    avg_qty_per_product = (total_qty / uniq_prod) if uniq_prod else None

    unit_price_p10 = float(round(up.quantile(0.10), 2)) if not up.empty else None
    unit_price_p50 = float(round(up.quantile(0.50), 2)) if not up.empty else None
    unit_price_p90 = float(round(up.quantile(0.90), 2)) if not up.empty else None

    return {
        "total_revenue": float(round(total_rev, 2)),
        "total_qty": float(round(total_qty, 2)),
        "total_weight": float(round(total_w, 2)),
        "unique_products": uniq_prod,
        "avg_margin": avg_margin_pct,
        "avg_margin_pct": avg_margin_pct,
        "margin_total": margin_total,
        "avg_unit_price": float(round(avg_up, 2)) if avg_up is not None else None,
        "median_unit_price": float(round(med_up, 2)) if med_up is not None else None,
        "revenue_per_product": revenue_per_product,
        "revenue_per_customer": revenue_per_customer,
        "customer_count": cust_count or None,
        "avg_qty_per_product": avg_qty_per_product,
        "unit_price_p10": unit_price_p10,
        "unit_price_p50": unit_price_p50,
        "unit_price_p90": unit_price_p90,
        "qty_label": qty_label,
    }


def _build_velocity_pulse(
    df: pd.DataFrame,
    kpis: Dict[str, Any],
    top_movers: Optional[List[Dict[str, Any]]] = None,
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    velocity = {
        "avg_weekly": None,
        "w13_trend": None,
        "weekly_revenue": None,
        "rev_per_product": kpis.get("revenue_per_product"),
        "active_skus": kpis.get("unique_products"),
        "roi_pct": None,
        "retail_velocity": None,
        "retail_velocity_basis": None,
        "top_mover": None,
    }
    health: Dict[str, Any] = {"missing_columns": [], "retail_velocity_unavailable": False}

    if df is None or df.empty:
        return velocity, health

    required = [CAN.date, CAN.product_id, CAN.revenue]
    missing = [c for c in required if c not in df.columns]
    if missing:
        health["missing_columns"] = missing

    working = df.copy()
    try:
        working[CAN.date] = pd.to_datetime(working[CAN.date], errors="coerce")
        working = working.dropna(subset=[CAN.date])
    except Exception:
        working = working.iloc[0:0]
    if working.empty:
        return velocity, health

    qty_series = _qty_basis_series(working)
    rev_series = pd.to_numeric(working.get(CAN.revenue, 0.0), errors="coerce").fillna(0.0)
    working["_qty_basis"] = qty_series
    working["_revenue"] = rev_series

    weekly = working.set_index(CAN.date).resample("W-MON").agg({"_qty_basis": "sum", "_revenue": "sum"})
    if not weekly.empty:
        velocity["avg_weekly"] = float(round(weekly["_qty_basis"].mean(), 2))
        velocity["w13_trend"] = float(round(weekly["_qty_basis"].tail(min(13, len(weekly))).mean(), 2))
        velocity["weekly_revenue"] = float(round(weekly["_revenue"].mean(), 2))

    if velocity["rev_per_product"] is None:
        prod_count = int(working[CAN.product_id].nunique()) if CAN.product_id in working.columns else 0
        velocity["rev_per_product"] = float(round(rev_series.sum() / prod_count, 2)) if prod_count else None
    velocity["active_skus"] = int(working[CAN.product_id].nunique()) if CAN.product_id in working.columns else None

    if CAN.cost in working.columns:
        cost_series = pd.to_numeric(working[CAN.cost], errors="coerce")
        cost_total = float(cost_series.sum(min_count=1)) if hasattr(cost_series, "sum") else None
        revenue_total = float(rev_series.sum()) if not rev_series.empty else 0.0
        if cost_total and cost_total > 0:
            roi = ((revenue_total - cost_total) / cost_total) * 100.0
            velocity["roi_pct"] = float(round(roi, 2))
        else:
            health["roi_unavailable"] = True
    else:
        health["roi_unavailable"] = True

    # Retail velocity: prefer explicit store/location, fallback to customers
    store_series = None
    store_basis = None
    for cand in ("StoreId", "store_id", "Store", "store", "LocationId", "location_id", "Location"):
        if cand in working.columns:
            s = working[cand].astype("string").str.strip()
            if s.notna().any():
                store_series = s
                store_basis = "store"
                break
    if store_series is None and CAN.customer_id in working.columns:
        store_series = working[CAN.customer_id].astype("string").str.strip()
        store_basis = "customer"
    store_count = int(store_series.replace("", pd.NA).dropna().nunique()) if store_series is not None else 0

    dates = working[CAN.date]
    date_span_days = float((dates.max() - dates.min()).days) + 1.0 if not dates.empty else 0.0
    week_span = date_span_days / 7.0 if date_span_days > 0 else 0.0
    if store_count > 0 and week_span > 0:
        velocity["retail_velocity"] = float(round(qty_series.sum() / (store_count * week_span), 2))
        velocity["retail_velocity_basis"] = store_basis
    else:
        health["retail_velocity_unavailable"] = True

    mover_rows = top_movers if top_movers is not None else _top_movers(working)
    if mover_rows:
        lead = mover_rows[0]
        velocity["top_mover"] = {
            "label": lead.get("desc") or lead.get("product") or lead.get("sku"),
            "sku": lead.get("sku") or lead.get("product_id"),
            "delta_revenue": lead.get("delta_revenue"),
            "growth_pct": lead.get("growth_pct"),
        }

    return velocity, health


def get_fact_df(filters: Optional[Dict[str, Any]] = None, fallback_months: int = DEFAULT_MONTHS) -> pd.DataFrame:
    """
    Lightweight accessor that applies filters to the cached sales parquet.
    Safe for tests and falls back to an empty schema-preserving frame.
    """
    try:
        parsed = parse_filters(filters or fallback_months)
    except Exception:
        parsed = parse_filters(fallback_months)

    try:
        df = sales_df(filters)
    except Exception:
        logger.exception("products.sales_df_failed")
        return _empty_products_frame()

    try:
        return apply_filters(df, parsed)
    except Exception:
        logger.exception("products.apply_filters_failed", exc_info=True)
        return _empty_products_frame()


@lru_cache(maxsize=64)
def _cached_df_overview(filters_kv: Tuple[Tuple[str, Any], ...], ttl_bucket: int, version: str) -> pd.DataFrame:
    _ = ttl_bucket
    _ = version
    filters = _tuple_to_filters(filters_kv)
    return get_fact_df(filters)


def _get_cached_df(filters: Dict[str, Any]) -> Tuple[pd.DataFrame, bool]:
    info_before = _cached_df_overview.cache_info()
    df = _cached_df_overview(_filters_to_tuple(filters), _ttl_bucket(PAYLOAD_LRU_TTL_SEC), _cache_version())
    info_after = _cached_df_overview.cache_info()
    cache_hit = info_after.hits > info_before.hits
    return df, cache_hit


def _normalize_trend(trend_obj: Any) -> Dict[str, Any]:
    """
    Accepts either the internal trend dict or a list[dict] with period/revenue keys.
    Returns a normalized dict that products.js expects.
    """
    if isinstance(trend_obj, dict) and "labels" in trend_obj:
        return {
            "labels": list(trend_obj.get("labels") or []),
            "revenue": list(trend_obj.get("revenue") or trend_obj.get("values") or []),
            "values": list(trend_obj.get("values") or trend_obj.get("revenue") or []),
            "qty": list(trend_obj.get("qty") or []),
            "weight": list(trend_obj.get("weight") or []),
            "margin": list(trend_obj.get("margin") or []),
            "asp": list(trend_obj.get("asp") or []),
            "forecast": trend_obj.get("forecast") or {},
        }

    if isinstance(trend_obj, list):
        labels = []
        revenue = []
        qty = []
        for item in trend_obj:
            try:
                labels.append(str(item.get("period")))
                revenue.append(float(item.get("revenue") or 0.0))
                qty.append(float(item.get("qty") or 0.0))
            except Exception:
                continue
        return {"labels": labels, "revenue": revenue, "values": revenue, "qty": qty, "weight": [], "margin": [], "asp": []}

    return {"labels": [], "revenue": [], "values": [], "qty": [], "weight": [], "margin": [], "asp": []}


def _build_insights(
    trend: Dict[str, Any],
    top_products: List[Dict[str, Any]],
    kpis: Dict[str, Any],
    forecast: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    insights: List[Dict[str, Any]] = []

    labels = trend.get("labels") or []
    revenue = trend.get("revenue") or trend.get("values") or []
    if len(labels) >= 2 and len(revenue) >= 2:
        curr = float(revenue[-1] or 0.0)
        prev = float(revenue[-2] or 0.0)
        delta = curr - prev
        delta_pct = (delta / prev * 100.0) if prev else None
        insights.append(
            {
                "metric": "revenue_momentum",
                "current": curr,
                "delta": delta,
                "delta_pct": delta_pct,
                "periods": labels[-2:],
            }
        )
        insights.append(
            {
                "metric": "mom_delta",
                "current": curr,
                "previous": prev,
                "delta": delta,
                "delta_pct": delta_pct,
                "period": labels[-1],
                "previous_period": labels[-2],
            }
        )

    if top_products:
        lead = top_products[0]
        insights.append(
            {
                "metric": "top_product",
                "label": lead.get("desc") or lead.get("product") or lead.get("sku"),
                "sku": lead.get("sku") or lead.get("product_id"),
                "share_pct": lead.get("revenue_share"),
                "margin_pct": lead.get("margin_pct"),
            }
        )

    if forecast:
        ds = forecast.get("ds") or forecast.get("dates") or forecast.get("labels") or []
        yhat = forecast.get("yhat") or forecast.get("values") or []
        lower = forecast.get("lower") or forecast.get("yhat_lower") or []
        upper = forecast.get("upper") or forecast.get("yhat_upper") or []
        if yhat:
            insights.append(
                {
                    "metric": "projected_next_month",
                    "label": ds[0] if ds else None,
                    "value": yhat[0],
                    "lower": lower[0] if lower else None,
                    "upper": upper[0] if upper else None,
                }
            )

    # Simple Pareto/80-20 note
    try:
        total_rev = float(kpis.get("total_revenue") or 0.0)
        if total_rev > 0 and top_products:
            sorted_rev = [float(p.get("revenue") or 0.0) for p in top_products]
            cum = 0.0
            count = 0
            for v in sorted_rev:
                cum += v
                count += 1
                if cum / total_rev >= 0.8:
                    break
            insights.append(
                {
                    "metric": "pareto",
                    "sku_count": count,
                    "share_pct": (cum / total_rev) * 100.0 if total_rev else None,
                }
            )
    except Exception:
        pass

    return insights


def _get_service_payload(filters: Dict[str, Any]) -> Dict[str, Any]:
    """
    Thin indirection layer so tests can monkeypatch service calls.
    """
    cache_key = None
    try:
        if cache is not None:
            cache_key = f"products:agg:{_api_cache_key('overview_payload', filters)}"
            if not current_app.config.get("TESTING"):
                cached = cache.get(cache_key)
                if cached is not None:
                    return copy.deepcopy(cached)
    except Exception:
        cache_key = None

    try:
        payload = build_overview_payload(filters=filters)
        if cache_key and cache is not None and not current_app.config.get("TESTING"):
            try:
                cache.set(cache_key, copy.deepcopy(payload), timeout=AGGREGATE_CACHE_TTL_SEC)
            except Exception:
                pass
        return payload
    except Exception:
        logger.exception("products.service_payload_failed")
        return {
            "kpis": _kpis(_empty_products_frame()),
            "trend": {"labels": [], "revenue": [], "values": [], "qty": [], "weight": [], "margin": [], "asp": []},
            "price_dist": {"prices": [], "p10": None, "p50": None, "p90": None},
            "top_products": [],
            "breakdowns": {},
            "top_movers": [],
            "insights": [],
            "velocity": {},
            "forecast": {},
            "data_health": {},
        }


def _sanitize_cost_sensitive_fields(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Remove cost/margin fields from payloads when the caller should not see them.
    """
    sanitized = copy.deepcopy(payload or {})
    top_rows = sanitized.get("top_products") or []
    for row in top_rows:
        if isinstance(row, dict):
            for key in ("cost", "cost_price", "margin", "margin_pct", "profit", "unit_cost", "unit_margin"):
                row.pop(key, None)
    return sanitized


def _build_overview_from_service(filters: Dict[str, Any]) -> Dict[str, Any]:
    payload = copy.deepcopy(_get_service_payload(filters) or {})
    meta = payload.setdefault("meta", {})

    qty_label = payload.get("qty_label") or meta.get("qty_label") or (payload.get("kpis") or {}).get("qty_label") or "Quantity"
    has_cost_data = bool(meta.get("has_cost_data", False))
    if not has_cost_data:
        top_rows = payload.get("top_products") or []
        has_cost_data = any((isinstance(r, dict) and r.get("cost") is not None) for r in top_rows)
    allow_costs = can_view_costs(current_user)
    has_cost = bool(has_cost_data and allow_costs)

    kpis = payload.get("kpis") or _kpis(_empty_products_frame())
    trend = _normalize_trend(payload.get("trend") or {})
    price_dist = payload.get("price_dist") or {"prices": [], "p10": None, "p50": None, "p90": None}
    top_products = payload.get("top_products") or []
    forecast = payload.get("forecast") or {}
    velocity = payload.get("velocity") or {}
    data_health = payload.get("data_health") or {}
    insights = payload.get("insights") or _build_insights(trend, top_products, kpis, forecast)

    payload["kpis"] = kpis
    payload["trend"] = trend
    payload["price_dist"] = price_dist
    payload["top_products"] = top_products
    payload.setdefault("breakdowns", {})
    payload["top_movers"] = payload.get("top_movers") or []
    payload["insights"] = insights
    payload["forecast"] = forecast
    payload["velocity"] = velocity
    payload["data_health"] = data_health
    payload["qty_label"] = qty_label
    payload["has_cost"] = has_cost
    payload.setdefault("meta", {}).update({"qty_label": qty_label, "has_cost_data": has_cost_data, "has_cost": has_cost})
    if "kpis" in payload and isinstance(payload["kpis"], dict):
        payload["kpis"]["qty_label"] = qty_label
    if not has_cost:
        if isinstance(payload.get("velocity"), dict):
            payload["velocity"]["roi_pct"] = None
        _strip_costs_from_payload(payload)

    return payload


def _segment_label(revenue_share: float, recency_days: float) -> str:
    if revenue_share >= 0.20:
        return "Stars" if recency_days <= 60 else "At Risk"
    if revenue_share >= 0.12:
        return "Steady Sellers" if recency_days <= 120 else "At Risk"
    if recency_days > 180:
        return "Dormant"
    if recency_days > 120:
        return "At Risk"
    return "Long Tail"


def _recommendation_label(margin_valid: bool, uplift_pct: Optional[float]) -> str:
    if not margin_valid:
        return "Cost missing"
    if uplift_pct is None:
        return "No price data"
    if uplift_pct >= 5.0:
        return "Increase price"
    if uplift_pct <= -5.0:
        return "Decrease price / investigate"
    return "On target"


def _build_sales_segments(filters: Dict[str, Any]) -> Dict[str, Any]:
    df = _cached_df_overview(_filters_to_tuple(filters), _ttl_bucket(PAYLOAD_LRU_TTL_SEC), _cache_version())
    if CAN.product_id not in df.columns and not df.empty:
        try:
            df = _standardize_sales_df(df)
        except Exception:
            logger.debug("products.segments.standardize_failed", exc_info=True)
    if df.empty:
        return {"products": [], "summary": [], "top_movers": [], "recommendations": []}

    df = df.copy()
    df["_qty_basis"] = _qty_basis_series(df)
    df["month"] = pd.to_datetime(df[CAN.date], errors="coerce").dt.to_period("M")
    last_sold_map = df.groupby(CAN.product_id, observed=True)[CAN.date].max()
    now = pd.Timestamp.utcnow().tz_localize(None).normalize()

    revenue_totals = df.groupby(CAN.product_id, observed=True)[CAN.revenue].sum().astype(float)
    qty_totals = df.groupby(CAN.product_id, observed=True)["_qty_basis"].sum().astype(float)
    revenue_sum_all = float(revenue_totals.sum())

    # ABC by cumulative revenue share
    revenue_sorted = revenue_totals.sort_values(ascending=False)
    cum_share = (revenue_sorted.cumsum() / revenue_sum_all) if revenue_sum_all else pd.Series(dtype=float)
    abc_map: Dict[str, str] = {}
    for pid, share in cum_share.items():
        if share <= 0.80:
            abc = "A"
        elif share <= 0.95:
            abc = "B"
        else:
            abc = "C"
        abc_map[str(pid)] = abc

    # XYZ by coefficient of variation of monthly qty_basis
    monthly_qty = df.groupby([CAN.product_id, "month"], observed=True)["_qty_basis"].sum()
    xyz_map: Dict[str, str] = {}
    cv_map: Dict[str, float] = {}
    for pid, grp in monthly_qty.groupby(level=0):
        vals = grp.values.astype(float)
        mean = float(np.nanmean(vals)) if len(vals) else 0.0
        std = float(np.nanstd(vals)) if len(vals) else 0.0
        cv = (std / mean) if mean else float("inf")
        cv_map[str(pid)] = cv
        if cv <= 0.5:
            xyz = "X"
        elif cv <= 1.0:
            xyz = "Y"
        else:
            xyz = "Z"
        xyz_map[str(pid)] = xyz

    # Velocity: average qty_basis per month over the most recent 3 months
    months_sorted = sorted(monthly_qty.index.get_level_values("month").unique())
    recent_months = set(months_sorted[-3:]) if months_sorted else set()
    recent_qty = monthly_qty[monthly_qty.index.get_level_values("month").isin(recent_months)]
    velocity_map = recent_qty.groupby(level=0).mean() if not recent_qty.empty else pd.Series(dtype=float)

    rows: List[Dict[str, Any]] = []
    for pid, rev in revenue_totals.items():
        pid_str = str(pid)
        qty_val = float(round(qty_totals.get(pid, 0.0), 2))
        share_fraction = (rev / revenue_sum_all) if revenue_sum_all else 0.0
        share_pct = (share_fraction * 100.0) if revenue_sum_all else None
        abc = abc_map.get(pid_str, "C")
        xyz = xyz_map.get(pid_str, "Z" if qty_val else "C")
        cv_val = cv_map.get(pid_str, float("inf"))
        velocity = float(round(velocity_map.get(pid, np.nan), 2)) if not isinstance(velocity_map, float) else None
        segment_code = f"{abc}{xyz}"
        last_sold = last_sold_map.get(pid)
        recency_days = None
        if pd.notna(last_sold):
            try:
                last_sold_ts = pd.Timestamp(last_sold).tz_localize(None).normalize()
                recency_days = int((now - last_sold_ts).days)
            except Exception:
                recency_days = None
        if recency_days is None or not revenue_sum_all:
            segment_label = "No Signal"
        else:
            segment_label = _segment_label(float(share_fraction), float(recency_days))

        rows.append(
            {
                "product_id": pid_str,
                "sku": pid_str,
                "segment": segment_label,
                "segment_code": segment_code,
                "segment_reasons": [
                    f"Revenue share {round(share_pct, 2)}%" if share_pct is not None else "No revenue",
                    f"Velocity {velocity} / mo" if velocity is not None else "No recent velocity",
                    f"CV {round(cv_val, 2)} ({xyz})" if not np.isinf(cv_val) else "No qty variation data",
                ],
                "key_metrics": {
                    "revenue": float(round(rev, 2)),
                    "qty": qty_val,
                    "revenue_share": round(share_pct, 2) if share_pct is not None else None,
                    "velocity": velocity,
                    "cv": float(round(cv_val, 3)) if not np.isinf(cv_val) else None,
                },
                "flags": {"unstable": cv_val > 1.0},
            }
        )

    rows.sort(key=lambda r: r["key_metrics"].get("revenue", 0.0), reverse=True)

    # Summary by segment code
    summary_map: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        seg = r["segment"]
        summary_map.setdefault(seg, {"segment": seg, "count": 0, "share_pct": 0.0})
        summary_map[seg]["count"] += 1
        try:
            summary_map[seg]["share_pct"] += r["key_metrics"].get("revenue", 0.0)
        except Exception:
            pass

    summary: List[Dict[str, Any]] = []
    for seg, payload in summary_map.items():
        share_pct = (payload["share_pct"] / revenue_sum_all * 100.0) if revenue_sum_all else None
        summary.append({"segment": seg, "count": payload["count"], "share_pct": round(share_pct, 2) if share_pct is not None else None})
    summary.sort(key=lambda s: s["segment"])

    movers = _top_movers(df)
    recs = []
    if any((r.get("segment_code") or "").startswith("C") for r in rows):
        recs.append({"title": "Trim long-tail SKUs", "explanation": "C-segment items contribute little revenue; consider consolidation."})
    if any((r.get("segment_code") or "").startswith("A") and r["flags"].get("unstable") for r in rows):
        recs.append({"title": "Stabilize A-segment volatility", "explanation": "High-value SKUs show unstable demand; review pricing and supply."})

    return {"products": rows, "summary": summary, "top_movers": movers, "recommendations": recs}


@lru_cache(maxsize=64)
def _cached_sales_segments(filters_kv: Tuple[Tuple[str, Any], ...], ttl_bucket: int, version: str) -> Dict[str, Any]:
    _ = ttl_bucket
    _ = version
    filters = _tuple_to_filters(filters_kv)
    return _build_sales_segments(filters)


def _bubble_payload(df: pd.DataFrame, filters: Dict[str, Any], limit: int) -> Dict[str, Any]:
    limit = max(1, min(int(limit), BUBBLE_LIMIT_MAX))
    if df is None or df.empty:
        return {"rows": [], "has_cost": False, "limit": limit, "total": 0}

    working = df.copy()
    working["_qty_basis"] = _qty_basis_series(working)
    working["month"] = pd.to_datetime(working[CAN.date], errors="coerce").dt.to_period("M")

    g = working.groupby([CAN.product_id, CAN.product_name], observed=True, sort=False)
    revenue = g[CAN.revenue].sum().astype(float)
    qty_total = g["_qty_basis"].sum().astype(float)
    sku_map = working.groupby(CAN.product_id, observed=True)[CAN.sku].agg(lambda s: s.dropna().iloc[0] if not s.dropna().empty else None)
    months_active = working.groupby(CAN.product_id, observed=True)["month"].nunique()

    cost_sum = g[CAN.cost].sum(min_count=1).astype(float) if CAN.cost in working.columns else pd.Series(dtype=float)
    has_cost = _has_cost_data(working)

    total_rev = float(revenue.sum()) if not revenue.empty else 0.0

    data = pd.DataFrame(
        {
            "product_id": revenue.index.get_level_values(0).astype(str),
            "product_name": revenue.index.get_level_values(1).astype(str),
            "revenue": revenue.values,
            "qty_basis": qty_total.values,
        }
    )
    data["sku"] = data["product_id"].map(sku_map)
    data["months_active"] = data["product_id"].map(months_active).fillna(0)
    data["velocity"] = np.where(data["months_active"] > 0, data["qty_basis"] / data["months_active"], np.nan)
    data["unit_price"] = np.where(data["qty_basis"] > 0, data["revenue"] / data["qty_basis"], np.nan)
    data["current_unit_price"] = data["unit_price"]
    data["velocity_month"] = data["velocity"]

    if has_cost and not cost_sum.empty:
        cost_map = {str(pid): val for pid, val in zip(cost_sum.index.get_level_values(0).astype(str), cost_sum.values)}
        data["cost"] = data["product_id"].map(cost_map)
        data["cost_available"] = (data["cost"] > 0) & data["cost"].notna()
        data["margin_valid"] = data["cost_available"] & (data["revenue"] > 0)
        data["cost"] = data["cost"].where(data["cost_available"], np.nan)
        data["profit"] = np.where(data["cost_available"], data["revenue"] - data["cost"], np.nan)
        data["unit_cost"] = np.where((data["qty_basis"] > 0) & data["cost_available"], data["cost"] / data["qty_basis"], np.nan)
        if au is not None and hasattr(au, "safe_margin_pct"):
            data["margin_pct"] = au.safe_margin_pct(data["revenue"], data["cost"])
        else:
            data["margin_pct"] = np.where(
                data["margin_valid"],
                (data["revenue"] - data["cost"]) / data["revenue"] * 100.0,
                np.nan,
            )
            data["margin_pct"] = np.clip(data["margin_pct"], -200.0, 200.0)
        data["target_price_27"] = np.where(data["unit_cost"].notna(), data["unit_cost"] / (1 - 0.27), np.nan)
        data["target_unit_price"] = data["target_price_27"]
        if au is not None and hasattr(au, "safe_uplift_pct"):
            data["uplift_pct"] = au.safe_uplift_pct(data["current_unit_price"], data["target_price_27"])
        else:
            data["uplift_pct"] = np.where(
                (data["current_unit_price"] > 0) & data["target_price_27"].notna(),
                (data["target_price_27"] - data["current_unit_price"]) / data["current_unit_price"] * 100.0,
                np.nan,
            )
            data["uplift_pct"] = np.clip(data["uplift_pct"], -100.0, 200.0)
        data["uplift_abs"] = np.where(
            (data["target_price_27"].notna()) & (data["current_unit_price"] > 0),
            (data["target_price_27"] - data["current_unit_price"].fillna(0)) * data["qty_basis"].fillna(0),
            np.nan,
        )
    else:
        data["cost"] = np.nan
        data["unit_cost"] = np.nan
        data["margin_pct"] = np.nan
        data["profit"] = np.nan
        data["target_price_27"] = np.nan
        data["target_unit_price"] = np.nan
        data["uplift_pct"] = np.nan
        data["uplift_abs"] = np.nan
        data["cost_available"] = False
        data["margin_valid"] = False

    data["revenue_share"] = (data["revenue"] / total_rev * 100.0) if total_rev else np.nan

    # Attach segment labels if available
    try:
        seg_payload = _cached_sales_segments(_filters_to_tuple(filters), _ttl_bucket(PAYLOAD_LRU_TTL_SEC), _cache_version())
        seg_map = {str(r.get("product_id")): r.get("segment") for r in seg_payload.get("products", []) if isinstance(r, dict)}
    except Exception:
        seg_map = {}
    data["segment"] = data["product_id"].map(seg_map)

    data = data.sort_values("revenue", ascending=False)
    total_rows = len(data)
    limit_cap = limit
    data = data.head(limit_cap)

    rows: List[Dict[str, Any]] = []
    for r in data.itertuples(index=False):
        rows.append(
            {
                "product_id": r.product_id,
                "product_name": r.product_name,
                "sku": r.sku or r.product_id,
                "segment": getattr(r, "segment", None),
                "revenue": float(round(r.revenue, 2)) if not pd.isna(r.revenue) else 0.0,
                "cost": float(round(r.cost, 2)) if hasattr(r, "cost") and not pd.isna(r.cost) else None,
                "profit": float(round(r.profit, 2)) if hasattr(r, "profit") and not pd.isna(r.profit) else None,
                "qty_basis": float(round(r.qty_basis, 2)) if not pd.isna(r.qty_basis) else 0.0,
                "velocity": float(round(r.velocity, 3)) if not pd.isna(r.velocity) else None,
                "velocity_month": float(round(r.velocity_month, 3)) if hasattr(r, "velocity_month") and not pd.isna(r.velocity_month) else None,
                "unit_price": float(round(r.unit_price, 4)) if not pd.isna(r.unit_price) else None,
                "current_unit_price": float(round(r.current_unit_price, 4)) if hasattr(r, "current_unit_price") and not pd.isna(r.current_unit_price) else None,
                "unit_cost": float(round(r.unit_cost, 4)) if not pd.isna(r.unit_cost) else None,
                "target_price_27": float(round(r.target_price_27, 4)) if hasattr(r, "target_price_27") and not pd.isna(r.target_price_27) else None,
                "target_unit_price": float(round(r.target_unit_price, 4)) if hasattr(r, "target_unit_price") and not pd.isna(r.target_unit_price) else None,
                "uplift_abs": float(round(r.uplift_abs, 2)) if hasattr(r, "uplift_abs") and not pd.isna(r.uplift_abs) else None,
                "uplift_pct": float(round(r.uplift_pct, 2)) if hasattr(r, "uplift_pct") and not pd.isna(r.uplift_pct) else None,
                "margin_pct": float(round(r.margin_pct, 2)) if hasattr(r, "margin_pct") and not pd.isna(r.margin_pct) else None,
                "cost_available": bool(getattr(r, "cost_available", False)),
                "margin_valid": bool(getattr(r, "margin_valid", False)),
                "revenue_share": float(round(r.revenue_share, 4)) if hasattr(r, "revenue_share") and not pd.isna(r.revenue_share) else None,
            }
        )

    return {"rows": rows, "has_cost": has_cost, "limit": limit, "total": total_rows}


# -----------------------------
# Cached payload builders (in-process)
# -----------------------------
def _filters_to_tuple(filters: Dict[str, Any]) -> Tuple[Tuple[str, Any], ...]:
    def _norm(v):
        if isinstance(v, list):
            return tuple(v)
        return v
    return tuple(sorted((k, _norm(v)) for k, v in (filters or {}).items()))


def _tuple_to_filters(t: Tuple[Tuple[str, Any], ...]) -> Dict[str, Any]:
    out = {}
    for k, v in t:
        if isinstance(v, tuple):
            out[k] = list(v)
        else:
            out[k] = v
    return out


def _log_debug_metrics(df: pd.DataFrame, filters: Dict[str, Any]) -> None:
    """Lightweight debug log for current filter scope."""
    try:
        cost_val = None
        if CAN.cost in df.columns:
            cost_series = pd.to_numeric(df[CAN.cost], errors="coerce")
            if cost_series.notna().any():
                cost_val = float(cost_series.sum())
        margin_100_count = 0
        margin_100_sample: List[str] = []
        try:
            if CAN.product_id in df.columns and CAN.revenue in df.columns and CAN.cost in df.columns:
                grouped = df.groupby(CAN.product_id, observed=True)[[CAN.revenue, CAN.cost]].sum(min_count=1)
                if au is not None and hasattr(au, "safe_margin_pct"):
                    grouped["margin_pct"] = au.safe_margin_pct(grouped[CAN.revenue], grouped[CAN.cost])
                else:
                    grouped["margin_pct"] = np.where(
                        (grouped[CAN.revenue] > 0) & (grouped[CAN.cost] > 0),
                        (grouped[CAN.revenue] - grouped[CAN.cost]) / grouped[CAN.revenue] * 100.0,
                        np.nan,
                    )
                if not grouped.empty and "margin_pct" in grouped.columns:
                    top = grouped.dropna(subset=["margin_pct"])
                    top = top[top["margin_pct"] >= 99.9]
                    top = top.sort_values(CAN.revenue, ascending=False).head(20)
                    margin_100_sample = [str(pid) for pid in top.index.tolist()]
                    margin_100_count = len(margin_100_sample)
        except Exception:
            margin_100_count = 0
            margin_100_sample = []
        missing_cost_pct = None
        missing_cost_sample: List[Dict[str, Any]] = []
        missing_cost_products: List[Dict[str, Any]] = []
        try:
            if CAN.product_id in df.columns and CAN.revenue in df.columns and CAN.cost in df.columns:
                grouped = df.groupby(CAN.product_id, observed=True).agg(
                    revenue=(CAN.revenue, "sum"),
                    cost=(CAN.cost, au.sum_cost),
                )
                if not grouped.empty:
                    missing_mask = grouped["cost"].isna() & (grouped["revenue"] > 0)
                    missing_count = int(missing_mask.sum())
                    missing_cost_pct = round((missing_count / max(len(grouped), 1)) * 100.0, 2)
                    top_missing = grouped[missing_mask].sort_values("revenue", ascending=False).head(10)
                    missing_cost_products = [
                        {"product_id": str(idx), "revenue": float(round(row["revenue"], 2))}
                        for idx, row in top_missing.iterrows()
                    ]

                    cost_cols = [c for c in df.columns if "cost" in str(c).lower()]
                    base_cols = [CAN.product_id, CAN.sku, CAN.product_name, CAN.date, CAN.qty, CAN.weight, CAN.cost]
                    sample_cols = [c for c in base_cols if c in df.columns] + [c for c in cost_cols if c not in base_cols]
                    sample_ids = top_missing.index.tolist()
                    if sample_cols and sample_ids:
                        sample_df = df[df[CAN.product_id].isin(sample_ids)][sample_cols].head(30)
                        missing_cost_sample = sample_df.fillna("").to_dict(orient="records")
        except Exception:
            missing_cost_pct = None
            missing_cost_sample = []
            missing_cost_products = []

        cost_cols = [c for c in df.columns if "cost" in str(c).lower()]
        payload = {
            "revenue_sum": float(pd.to_numeric(df.get(CAN.revenue, 0.0), errors="coerce").sum()),
            "qty_basis_sum": float(_qty_basis_series(df).sum()),
            "weight_sum": float(_weight_series(df).sum()),
            "cost_sum": cost_val,
            "margin_100_count": margin_100_count,
            "margin_100_sample": margin_100_sample,
            "cost_columns": cost_cols,
            "missing_cost_pct": missing_cost_pct,
            "missing_cost_products": missing_cost_products,
            "missing_cost_sample": missing_cost_sample,
        }
        logger.info(
            "products.metrics.debug",
            extra={"event": "products_metrics_debug", "columns": list(df.columns), "sample": payload, "filters": filters},
        )
    except Exception:
        logger.debug("products.metrics.debug_failed", exc_info=True)


@lru_cache(maxsize=64)
def _cached_overview_payload(filters_kv: Tuple[Tuple[str, Any], ...], ttl_bucket: int, version: str) -> Dict[str, Any]:
    _ = ttl_bucket
    _ = version
    filters = _tuple_to_filters(filters_kv)
    data_health: Dict[str, Any] = {"defaulted_window": False, "missing_columns": []}

    df = get_fact_df(filters)
    if df.empty:
        try:
            start, end = _recent_window(DEFAULT_MONTHS)
            fallback_filters = {
                **filters,
                "start": start.date().isoformat(),
                "end": end.date().isoformat(),
                "start_date": start.date().isoformat(),
                "end_date": end.date().isoformat(),
            }
            fallback_df = get_fact_df(fallback_filters)
            if not fallback_df.empty:
                df = fallback_df
                data_health["defaulted_window"] = True
                data_health["fallback_window"] = {"start": str(start.date()), "end": str(end.date())}
        except Exception:
            logger.debug("products.overview.fallback_failed", exc_info=True)

    if not df.empty and (CAN.product_id not in df.columns or CAN.date not in df.columns or CAN.revenue not in df.columns):
        try:
            df = _standardize_sales_df(df)
        except Exception:
            logger.debug("products.overview.standardize_failed", exc_info=True)
    df = df.copy()
    qty_label = _qty_label(df)
    has_cost_data = _has_cost_data(df)
    _log_debug_metrics(df, filters)

    try:
        dates = pd.to_datetime(df.get(CAN.date), errors="coerce") if CAN.date in df.columns else pd.Series(dtype="datetime64[ns]")
        logger.debug(
            "product_intel.kpi.input",
            extra={
                "rows": int(len(df)),
                "date_min": str(dates.min()) if not dates.empty else None,
                "date_max": str(dates.max()) if not dates.empty else None,
                "products": int(df[CAN.product_id].nunique()) if CAN.product_id in df.columns else 0,
                "rev": float(pd.to_numeric(df.get(CAN.revenue, 0.0), errors="coerce").sum()) if CAN.revenue in df.columns else 0.0,
                "cost": float(pd.to_numeric(df.get(CAN.cost, 0.0), errors="coerce").sum()) if CAN.cost in df.columns else None,
            },
        )
    except Exception:
        logger.debug("product_intel.kpi.input_failed", exc_info=True)

    trend = _monthly_trend(df)
    price_dist = _price_dist(df)
    top_products = _top_products(df, TOP_N_MAX)
    movers = _top_movers(df)

    base_kpis = _kpis(df)
    kpis = base_kpis | {
        "unit_price_p10": price_dist.get("p10"),
        "unit_price_p50": price_dist.get("p50"),
        "unit_price_p90": price_dist.get("p90"),
    }
    velocity, velocity_health = _build_velocity_pulse(df, kpis, movers)
    if velocity_health.get("missing_columns"):
        data_health["missing_columns"] = sorted(set(data_health.get("missing_columns", []) + velocity_health["missing_columns"]))
    if velocity_health.get("retail_velocity_unavailable"):
        data_health["retail_velocity_unavailable"] = True
    forecast_obj: Dict[str, Any] = {}
    if _coerce_bool(filters.get("forecast"), default=False):
        try:
            forecast_obj = _simple_forecast(
                df,
                periods=6,
                include_current=_coerce_bool(filters.get("include_current_month"), default=False),
            )
            if forecast_obj:
                forecast_obj.setdefault("ds", forecast_obj.get("dates", []))
        except Exception:
            logger.debug("products.overview.forecast_failed", exc_info=True)

    payload = {
        "kpis": kpis,
        "trend": trend,
        "price_dist": price_dist,
        "top_products": top_products,
        "breakdowns": {},  # keep simple; add later if needed
        "top_movers": movers,
        "insights": [],
        "qty_label": qty_label,
        "meta": {"qty_label": qty_label, "has_cost_data": has_cost_data},
        "velocity": velocity,
        "forecast": forecast_obj,
        "data_health": data_health | velocity_health,
    }

    try:
        logger.debug(
            "product_intel.kpi.output",
            extra={
                "kpis": payload.get("kpis"),
                "velocity": payload.get("velocity"),
                "insights": [i.get("metric") for i in payload.get("insights", [])],
            },
        )
    except Exception:
        logger.debug("product_intel.kpi.output_failed", exc_info=True)

    return payload


def build_overview_payload(
    filters: Optional[Dict[str, Any]] = None,
    include_forecast: Optional[bool] = None,
    fallback_months: int = DEFAULT_MONTHS,
) -> Dict[str, Any]:
    if filters is None:
        try:
            filters = parse_filters(fallback_months=fallback_months)
        except TypeError:
            filters = parse_filters()
    else:
        filters = copy.deepcopy(filters) if filters is not None else {}
    if include_forecast is not None:
        try:
            filters["forecast"] = bool(include_forecast)
        except Exception:
            pass
    try:
        if current_app and current_app.config.get("TESTING"):
            return _cached_overview_payload.__wrapped__(
                _filters_to_tuple(filters),
                _ttl_bucket(PAYLOAD_LRU_TTL_SEC),
                _cache_version(),
            )
    except Exception:
        if os.getenv("PYTEST_CURRENT_TEST"):
            return _cached_overview_payload.__wrapped__(
                _filters_to_tuple(filters),
                _ttl_bucket(PAYLOAD_LRU_TTL_SEC),
                _cache_version(),
            )
    return _cached_overview_payload(_filters_to_tuple(filters), _ttl_bucket(PAYLOAD_LRU_TTL_SEC), _cache_version())


@lru_cache(maxsize=128)
def _cached_table_payload(filters_kv: Tuple[Tuple[str, Any], ...], ttl_bucket: int, page: int, per_page: int, sort_by: str, sort_dir: str) -> Dict[str, Any]:
    _ = ttl_bucket
    filters = _tuple_to_filters(filters_kv)
    payload = _build_overview_from_service(filters)
    rows = list(payload.get("top_products") or [])
    qty_label = payload.get("qty_label") or (payload.get("meta") or {}).get("qty_label") or "Quantity"
    has_cost = bool(payload.get("has_cost"))

    sort_by = (sort_by or "revenue").lower()
    sort_dir = (sort_dir or "desc").lower()
    reverse = sort_dir != "asc"

    valid = {
        "revenue",
        "qty",
        "avg_price",
        "unit_price",
        "current_unit_price",
        "revenue_share",
        "qty_share",
        "target_price",
        "target_unit_price",
        "uplift_pct",
        "cost",
        "profit",
        "margin",
        "margin_pct",
    }
    if sort_by not in valid:
        sort_by = "revenue"

    rows.sort(key=lambda r: (r.get(sort_by) or 0.0), reverse=reverse)

    total = len(rows)
    page = max(1, int(page))
    per_page = max(1, min(int(per_page), TABLE_PAGE_SIZE_MAX))
    start = (page - 1) * per_page
    end = start + per_page
    return {
        "rows": rows[start:end],
        "page": page,
        "per_page": per_page,
        "total": total,
        "qty_label": qty_label,
        "has_cost": has_cost,
    }


def build_table_payload(filters: Dict[str, Any], page: int, per_page: int, sort_by: str, sort_dir: str) -> Dict[str, Any]:
    return _cached_table_payload(
        _filters_to_tuple(filters),
        _ttl_bucket(PAYLOAD_LRU_TTL_SEC),
        page,
        per_page,
        sort_by,
        sort_dir,
    )


# -----------------------------
# Recommendations (simple + useful)
# -----------------------------
def _co_purchase_stats(df: pd.DataFrame, target_pid: str) -> List[Dict[str, Any]]:
    if df.empty or df[CAN.order_id].isna().all():
        return []

    orders = df[[CAN.order_id, CAN.product_id]].dropna()
    if orders.empty:
        return []

    # sample orders to keep bounded
    uniq_orders = orders[CAN.order_id].drop_duplicates()
    if len(uniq_orders) > CO_PURCHASE_SAMPLE_ORDERS:
        sample = uniq_orders.sample(CO_PURCHASE_SAMPLE_ORDERS, random_state=42)
        orders = orders[orders[CAN.order_id].isin(sample)]

    basket = orders.groupby(CAN.order_id, observed=True)[CAN.product_id].apply(lambda s: set(s.astype(str).tolist()))
    orders_total = int(basket.shape[0])
    if orders_total == 0:
        return []

    # counts
    with_target = 0
    partner_counts: Dict[str, int] = {}
    partner_orders: Dict[str, int] = {}

    for items in basket:
        has_target = str(target_pid) in items
        if has_target:
            with_target += 1
        for pid in items:
            partner_orders[pid] = partner_orders.get(pid, 0) + 1
            if has_target and pid != str(target_pid):
                partner_counts[pid] = partner_counts.get(pid, 0) + 1

    if with_target == 0:
        return []

    stats = []
    for partner, both in partner_counts.items():
        support = both / orders_total
        confidence = both / with_target
        base = (partner_orders.get(partner, 0) / orders_total) if orders_total else 0
        lift = (confidence / base) if base else None
        stats.append({"partner_id": partner, "support": support, "confidence": confidence, "lift": lift, "orders_with_both": both, "orders_with_target": with_target})

    stats.sort(key=lambda x: x.get("confidence") or 0.0, reverse=True)
    return stats[:10]


def build_product_recommendations(filters: Dict[str, Any], product_id: str) -> Dict[str, Any]:
    df = apply_filters(sales_df(), filters)
    if df.empty:
        return {"product_id": product_id, "recommendations": [], "meta": {}}

    sub = df[df[CAN.product_id].astype(str) == str(product_id)]
    if sub.empty:
        return {"product_id": product_id, "recommendations": [], "meta": {"warning": "no_data_for_product"}}

    recs: List[Dict[str, Any]] = []

    # Price standardization: tighten band around median
    up = _unit_price(sub).dropna()
    qty = _qty_basis_series(sub)
    revenue = pd.to_numeric(sub[CAN.revenue], errors="coerce").fillna(0.0)

    median_price = float(round(up.median(), 2)) if not up.empty else None
    p90 = float(round(up.quantile(0.90), 2)) if not up.empty else None

    if median_price is not None:
        recs.append(
            {
                "title": "Tighten price guardrails",
                "explanation": f"Median price {median_price:.2f}. Keep offers within P10-P90 band to avoid leakage.",
            }
        )
    if p90 is not None and median_price is not None and p90 > median_price:
        uplift = ((p90 - median_price) / median_price) * 100
        recs.append(
            {
                "title": "Consider premium tier",
                "explanation": f"Top decile pricing is {uplift:.1f}% above median; consider premium pack or bundle.",
            }
        )

    if not qty.empty and len(qty) >= 3:
        recent_avg = float(round(qty.tail(3).mean(), 2))
        recs.append(
            {
                "title": "Velocity check",
                "explanation": f"Recent avg qty per order: {recent_avg}. Ensure stock levels match demand.",
            }
        )

    total_rev = float(revenue.sum()) if not revenue.empty else 0.0
    if total_rev > 0:
        recs.append(
            {
                "title": "Revenue concentration",
                "explanation": f"Total revenue {round(total_rev,2)} across {len(sub)} shipments. Target customer follow-ups.",
            }
        )

    # Customer-based insights
    top_cust = _top_customers_for_product(sub, limit=3)
    for cust in top_cust:
        recs.append(
            {
                "title": f"Grow {cust.get('Customer')}",
                "explanation": f"{cust.get('Customer')} contributes {cust.get('RevenueShare', 0):.1f}% of revenue; explore upsell.",
            }
        )

    return {"product_id": product_id, "recommendations": recs, "meta": {"samples": len(sub), "median_price": median_price}}


# Backward-compatible alias for tests/consumers
_build_product_recommendations = build_product_recommendations


# --------- Drilldown helpers ---------

def _simple_forecast(df: pd.DataFrame, periods: int = 6, include_current: bool = False, today: Optional[pd.Timestamp] = None) -> Dict[str, Any]:
    """
    Simple moving average forecast (no Prophet dependency).
    Returns dict with dates, yhat, lower, upper arrays.
    """
    if df.empty or df[CAN.date].isna().all():
        return {}

    tmp = df[[CAN.date, CAN.revenue]].copy(deep=False).dropna(subset=[CAN.date, CAN.revenue])
    if len(tmp) < 2:
        return {}

    tmp = tmp.sort_values(CAN.date)
    tmp["month"] = tmp[CAN.date].dt.to_period("M").dt.to_timestamp()
    monthly = tmp.groupby("month", observed=True)[CAN.revenue].sum().astype(float)

    # Always train on complete months (exclude the current partial unless explicitly overridden)
    try:
        now = pd.to_datetime(today) if today is not None else pd.Timestamp.utcnow()
        now = pd.Timestamp(now).tz_localize(None)
        current_month_start = now.normalize().replace(day=1)
    except Exception:
        current_month_start = pd.Timestamp.utcnow().normalize().replace(day=1)

    if not include_current:
        monthly = monthly[monthly.index < current_month_start]
    if monthly.empty:
        return {}

    if len(monthly) < 3:
        return {}  # not enough data

    vals = monthly.values.astype(float)
    last_date = monthly.index[-1]

    # 3-month moving average
    ma3 = pd.Series(vals).rolling(min(3, len(vals)), center=False).mean()
    last_ma = float(ma3.iloc[-1]) if not pd.isna(ma3.iloc[-1]) else float(vals[-1])

    # Simple linear trend
    x = np.arange(len(vals), dtype=float)
    try:
        # Simple linear regression slope (robust to numpy rcond changes)
        slope, _ = np.linalg.lstsq(np.vstack([x, np.ones(len(x))]).T, vals, rcond=None)[0]
        trend = float(slope)
    except Exception:
        trend = 0.0

    # Generate forecast
    forecast_dates = []
    forecast_vals = []
    forecast_lower = []
    forecast_upper = []

    volatility = float(np.std(vals[-min(3, len(vals)):]))

    for i in range(1, periods + 1):
        next_date = last_date + pd.DateOffset(months=i)
        next_val = max(0.0, last_ma + trend * i)
        lower = max(0.0, next_val - 1.96 * volatility)
        upper = next_val + 1.96 * volatility

        forecast_dates.append(next_date.strftime("%Y-%m"))
        forecast_vals.append(float(round(next_val, 2)))
        forecast_lower.append(float(round(lower, 2)))
        forecast_upper.append(float(round(upper, 2)))

    return {
        "dates": forecast_dates,
        "yhat": forecast_vals,
        "lower": forecast_lower,
        "upper": forecast_upper,
        "model": "simple_ma",
        "confidence": 85,
        "mape": None,
    }


def _lifecycle_stage(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Simple product lifecycle classification based on trend.
    """
    if df.empty or df[CAN.date].isna().all():
        return {"stage": "Unknown", "confidence": 0, "growth_rate": None, "recent_avg": None, "message": ""}

    tmp = df[[CAN.date, CAN.revenue]].copy(deep=False).dropna(subset=[CAN.date, CAN.revenue])
    if len(tmp) < 2:
        return {"stage": "Early", "confidence": 50, "growth_rate": None, "recent_avg": None, "message": "Insufficient data"}

    tmp = tmp.sort_values(CAN.date)
    tmp["month"] = tmp[CAN.date].dt.to_period("M").dt.to_timestamp()
    monthly = tmp.groupby("month", observed=True)[CAN.revenue].sum().astype(float)

    if len(monthly) < 2:
        return {"stage": "Early", "confidence": 60, "growth_rate": None, "recent_avg": None, "message": "Limited history"}

    recent = float(monthly.iloc[-1])
    recent_avg = float(monthly.tail(min(3, len(monthly))).mean())
    prior_avg = float(monthly.iloc[: max(1, len(monthly) - 6)].mean()) if len(monthly) > 6 else float(monthly.iloc[0])

    growth = (recent - prior_avg) / prior_avg * 100.0 if prior_avg > 0 else 0.0

    if growth > 20:
        stage, conf = "Growth", 80
    elif growth > 5:
        stage, conf = "Stable", 75
    elif growth > -10:
        stage, conf = "Mature", 70
    else:
        stage, conf = "Decline", 75

    return {
        "stage": stage,
        "confidence": int(conf),
        "growth_rate": float(round(growth, 1)),
        "recent_avg": float(round(recent_avg, 2)),
        "message": f"{stage} phase with {growth:+.1f}% growth trend",
    }


def _anomalies(df: pd.DataFrame, z_threshold: float = 3.0) -> List[Dict[str, Any]]:
    """
    Detect revenue anomalies using z-score.
    """
    if df.empty or df[CAN.date].isna().all():
        return []

    tmp = df[[CAN.date, CAN.revenue]].copy(deep=False).dropna(subset=[CAN.date, CAN.revenue])
    if len(tmp) < 4:
        return []

    tmp = tmp.sort_values(CAN.date)
    tmp["month"] = tmp[CAN.date].dt.to_period("M").dt.to_timestamp()
    monthly = tmp.groupby("month", observed=True)[CAN.revenue].sum().astype(float)

    if len(monthly) < 4:
        return []

    vals = monthly.values
    mean = float(np.mean(vals))
    std = float(np.std(vals, ddof=1))

    if std == 0:
        return []

    anomalies_out: List[Dict[str, Any]] = []
    for i, (date, val) in enumerate(zip(monthly.index.to_pydatetime(), vals)):
        z = (val - mean) / std
        if abs(z) >= z_threshold:
            expected = mean
            anomalies_out.append({
                "date": date.strftime("%Y-%m"),
                "value": float(round(val, 2)),
                "expected": float(round(expected, 2)),
                "z_score": float(round(z, 2)),
                "severity": "high" if abs(z) > 3.5 else "medium",
                "direction": "spike" if z > 0 else "drop",
            })

    return anomalies_out


def _price_optimization_insights(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Simple price optimization guidance (sales-only heuristics).
    """
    if df.empty:
        return {}

    up = _unit_price(df).dropna()
    if up.empty or len(up) < 3:
        return {}

    current_avg = float(up.mean())
    optimal_min = float(up.quantile(0.25))
    optimal_max = float(up.quantile(0.75))

    return {
        "elasticity": "—",  # would need demand elasticity analysis
        "current_avg_price": float(round(current_avg, 2)),
        "optimal_price_min": float(round(optimal_min, 2)),
        "optimal_price_max": float(round(optimal_max, 2)),
        "recommendation": f"Current pricing is {'within' if optimal_min <= current_avg <= optimal_max else 'outside'} the optimal range. Monitor conversion metrics.",
    }


def _top_customers_for_product(df: pd.DataFrame, limit: int = 10) -> List[Dict[str, Any]]:
    """
    Top customers by revenue.
    """
    if df.empty:
        return []

    g = df.groupby(CAN.customer_id, observed=True)[CAN.revenue].sum().sort_values(ascending=False).head(limit)
    rows = []
    for cid, rev in g.items():
        cust_subset = df[df[CAN.customer_id] == cid]
        if cust_subset[CAN.customer_name].notna().any():
            cname = str(cust_subset[CAN.customer_name].dropna().iloc[0])
        else:
            cname = str(cid)
        rows.append({"customer_id": str(cid), "Customer": cname, "Revenue": float(round(rev, 2))})
    return rows


def _customer_options_for_product(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    List all customers for dropdown.
    """
    if df.empty or CAN.customer_id not in df.columns:
        return []

    uniq = df[[CAN.customer_id, CAN.customer_name]].drop_duplicates().dropna(subset=[CAN.customer_id])
    return [
        {"id": int(cid), "label": str(cname)} if str(cid).isdigit() else {"id": str(cid), "label": str(cname)}
        for cid, cname in zip(uniq[CAN.customer_id], uniq[CAN.customer_name])
    ]


def _price_suggestion_for_customer(df: pd.DataFrame, customer_id: Any) -> Optional[Dict[str, Any]]:
    """
    Suggest a price for this product given a specific customer.
    Uses sales-only heuristics (no cost data).
    """
    if df.empty:
        return None

    # Filter to this customer
    sub = df[df[CAN.customer_id].astype(str) == str(customer_id)]
    if sub.empty:
        return None

    up = _unit_price(sub).dropna()
    if up.empty:
        return None

    current_price = float(up.mean())
    # Simple heuristic: suggest median of upper quartile
    suggested = float(up.quantile(0.75))

    return {
        "current_price": float(round(current_price, 2)),
        "suggested_price": float(round(suggested, 2)),
        "rationale": f"Based on this customer's historical pricing ({len(up)} transactions). Suggested value is 75th percentile.",
    }


def _abc_xyz_classification(df: pd.DataFrame) -> Tuple[str, float]:
    """
    Simple ABC-XYZ classification.
    ABC: A (top 20% revenue), B (next 30%), C (rest)
    XYZ: X (CV < 0.25 stable), Y (0.25-0.5 moderate), Z (> 0.5 variable)
    """
    if df.empty:
        return "—", 0.0

    tmp = df[[CAN.date, CAN.revenue]].copy(deep=False).dropna(subset=[CAN.date, CAN.revenue])
    if len(tmp) < 2:
        return "—", 0.0

    tmp = tmp.sort_values(CAN.date)
    tmp["month"] = tmp[CAN.date].dt.to_period("M").dt.to_timestamp()
    monthly = tmp.groupby("month", observed=True)[CAN.revenue].sum().astype(float)

    total_rev = float(df[CAN.revenue].sum())
    total_all = float(df[CAN.revenue].sum())  # would need all products for true ABC

    # Simplified: if this product has >15% of total, A; >5%, B; else C
    pct = (total_rev / total_all * 100) if total_all else 0
    abc = "A" if pct > 15 else ("B" if pct > 5 else "C")

    vals = monthly.values
    if len(vals) < 2:
        cv = 0.0
        xyz = "X"
    else:
        mean = float(np.mean(vals))
        std = float(np.std(vals, ddof=1))
        cv = (std / mean) if mean > 0 else 0.0
        xyz = "X" if cv < 0.25 else ("Y" if cv < 0.5 else "Z")

    return f"{abc}{xyz}", cv * 100.0


# --------- Routes ---------


@bp.route("/", methods=["GET", "POST"])
@login_required
@requires_roles(*VIEW_ROLES)
def index():
    """
    Overview page.
    - POST stores filters in session (optional)
    - GET reads from query args/session
    """
    # Defensive redirect: if a SKU sneaks onto /products, send to drilldown.
    try:
        sku_hint = request.args.get("sku") or request.args.get("product_id") or request.args.get("id")
    except Exception:
        sku_hint = None
    if sku_hint:
        qs_args = request.args.to_dict(flat=False)
        for key in ("sku", "product_id", "product", "id"):
            qs_args.pop(key, None)
        qs = urlencode(qs_args, doseq=True)
        target = url_for("products.drilldown", product_id=str(sku_hint))
        return redirect(f"{target}?{qs}" if qs else target)

    fallback_months = DEFAULT_MONTHS
    try:
        if current_app.config.get("TESTING"):
            fallback_months = max(DEFAULT_MONTHS, 36)
    except Exception:
        pass
    filters = parse_filters(fallback_months)
    querystring = build_querystring(filters)
    querystring_suffix = querystring[1:] if querystring.startswith("?") else querystring

    if request.method == "POST":
        # Minimal "store what came in" behavior; your template can POST start/end etc.
        session["products_filters"] = filters
        return redirect(url_for("products.index"))

    payload = _build_overview_from_service(filters)
    status = get_products_parquet_status()
    payload = _attach_parquet_warning(payload, status, add_empty_data=not status.available)
    try:
        bubble_url = url_for("products.api_bubble")
    except BuildError:
        logger.warning("products.api_bubble_missing")
        bubble_url = ""
    products_v4 = bool(current_app.config.get("PRODUCTS_V4", False))
    products_v3 = bool(current_app.config.get("PRODUCTS_V3", False)) or products_v4
    products_v2 = bool(current_app.config.get("PRODUCT_INTELLIGENCE_V2", False)) or products_v3
    template_name = "products/index_v4.html" if products_v4 else ("products/index_v3.html" if products_v3 else "products/index.html")
    return render_template(
        template_name,
        filters=filters,
        payload=payload,
        products_warning=status.warning or payload.get("warning"),
        product_intelligence_v2=products_v2,
        products_v3=products_v3,
        products_v4=products_v4,
        currency_code=current_app.config.get("CURRENCY_CODE", "CAD"),
        querystring=querystring,
        querystring_suffix=querystring_suffix,
        bubble_url=bubble_url,
    )


@bp.route("/api/overview")
@login_required
@requires_roles(*VIEW_ROLES)
def api_overview():
    filters = parse_filters(DEFAULT_MONTHS)
    payload = _build_overview_from_service(filters)
    status = get_products_parquet_status()
    payload = _attach_parquet_warning(payload, status, add_empty_data=not status.available)
    ck = _api_cache_key("overview", filters)
    return _json_response(payload, cache_key=ck, ttl=API_TTL_SEC)


@bp.route("/api/table")
@login_required
@requires_roles(*VIEW_ROLES)
def api_table():
    started = time.perf_counter()
    filters = parse_filters(DEFAULT_MONTHS)
    page = _safe_int(request.args.get("page", 1), 1)
    per_page = _safe_int(
        request.args.get("per_page", request.args.get("page_size", TABLE_PAGE_SIZE_DEFAULT)),
        TABLE_PAGE_SIZE_DEFAULT,
    )
    per_page = max(1, min(per_page, TABLE_PAGE_SIZE_MAX))
    sort_by = request.args.get("sort_by", "revenue")
    sort_dir = request.args.get("sort_dir", "desc")

    payload = build_table_payload(
        filters=filters,
        page=page,
        per_page=per_page,
        sort_by=sort_by,
        sort_dir=sort_dir,
    )
    status = get_products_parquet_status()
    payload = _attach_parquet_warning(payload, status, add_empty_data=not status.available)
    ck = _api_cache_key("table", filters, {"p": page, "pp": per_page, "sb": sort_by, "sd": sort_dir})
    resp = _json_response(payload, cache_key=ck, ttl=API_TTL_SEC)
    _log_timing(
        "products.api_table",
        started,
        filters,
        {"rows_in_slice": payload.get("total", 0), "page": page, "per_page": per_page},
    )
    return resp


@bp.route("/api/trend")
@login_required
@requires_roles(*VIEW_ROLES)
def api_trend():
    filters = parse_filters(DEFAULT_MONTHS)
    payload = _build_overview_from_service(filters)
    out = payload.get("trend", {}) or {}
    status = get_products_parquet_status()
    out = _attach_parquet_warning(out, status, add_empty_data=not status.available)
    ck = _api_cache_key("trend", filters)
    return _json_response(out, cache_key=ck, ttl=API_TTL_SEC)


@bp.route("/api/price_dist")
@login_required
@requires_roles(*VIEW_ROLES)
def api_price_dist():
    filters = parse_filters(DEFAULT_MONTHS)
    started = time.perf_counter()
    df, cache_hit = _get_cached_df(filters)
    status = get_products_parquet_status()
    out = _price_dist(df)
    out = _attach_parquet_warning(out, status, add_empty_data=not status.available)
    ck = _api_cache_key("price_dist", filters)
    resp = _json_response(out, cache_key=ck, ttl=API_TTL_SEC)
    _log_timing(
        "products.api_price_dist",
        started,
        filters,
        {
            "rows_in_slice": int(len(df)),
            "cache_hit": cache_hit,
            "products": int(df[CAN.product_id].nunique()) if CAN.product_id in df.columns else 0,
        },
    )
    return resp


@bp.route("/api/price_distribution")
@login_required
@requires_roles(*VIEW_ROLES)
def api_price_distribution():
    """Alias for products.js fallback compatibility."""
    return api_price_dist()


@bp.route("/api/trend_delta")
@login_required
@requires_roles(*VIEW_ROLES)
def api_trend_delta():
    filters = parse_filters(DEFAULT_MONTHS)
    started = time.perf_counter()
    df, cache_hit = _get_cached_df(filters)
    movers = _top_movers(df)
    labels = [m.get("desc") or m.get("sku") or m.get("product_id") for m in movers]
    values = [m.get("delta_revenue") or 0.0 for m in movers]
    payload = {"labels": labels, "values": values, "rows": movers}
    status = get_products_parquet_status()
    payload = _attach_parquet_warning(payload, status, add_empty_data=not status.available)
    ck = _api_cache_key("trend_delta", filters)
    resp = _json_response(payload, cache_key=ck, ttl=API_TTL_SEC)
    _log_timing(
        "products.api_trend_delta",
        started,
        filters,
        {
            "rows_in_slice": int(len(df)),
            "cache_hit": cache_hit,
            "movers": len(movers),
            "products": int(df[CAN.product_id].nunique()) if CAN.product_id in df.columns else 0,
        },
    )
    return resp


@bp.route("/api/segments")
@login_required
@requires_roles(*VIEW_ROLES)
def api_segments():
    filters = parse_filters(DEFAULT_MONTHS)
    started = time.perf_counter()
    page_size = _safe_int(request.args.get("page_size") or request.args.get("per_page") or SEGMENTS_PAGE_SIZE_MAX, SEGMENTS_PAGE_SIZE_MAX)
    page_size = max(1, min(page_size, SEGMENTS_PAGE_SIZE_MAX))
    payload = copy.deepcopy(_cached_sales_segments(_filters_to_tuple(filters), _ttl_bucket(PAYLOAD_LRU_TTL_SEC), _cache_version()))
    total_products = len(payload.get("products") or [])
    payload["products"] = (payload.get("products") or [])[:page_size]
    payload["meta"] = {**payload.get("meta", {}), "total_products": total_products, "page_size": page_size}
    status = get_products_parquet_status()
    payload = _attach_parquet_warning(payload, status, add_empty_data=not status.available)
    ck = _api_cache_key("segments", filters, {"ps": page_size})
    resp = _json_response(payload, cache_key=ck, ttl=API_TTL_SEC)
    _log_timing(
        "products.api_segments",
        started,
        filters,
        {"rows_in_slice": total_products, "page_size": page_size},
    )
    return resp


@bp.route("/api/bundle")
@login_required
@requires_roles(*VIEW_ROLES)
def api_bundle():
    from app.services import bundle_service

    payload = bundle_service.bundle("products", request.args)
    return jsonify(payload)
    bubble_limit = _safe_int(request.args.get("bubble_limit") or request.args.get("limit") or BUBBLE_LIMIT_MAX, BUBBLE_LIMIT_MAX)
    bubble_limit = max(1, min(bubble_limit, BUBBLE_LIMIT_MAX))
    per_page = _safe_int(request.args.get("per_page", request.args.get("page_size", TABLE_PAGE_SIZE_DEFAULT)), TABLE_PAGE_SIZE_DEFAULT)
    per_page = max(1, min(per_page, TABLE_PAGE_SIZE_MAX))
    segments_page_size = _safe_int(request.args.get("segments_page_size") or request.args.get("segments_per_page") or request.args.get("page_size") or SEGMENTS_PAGE_SIZE_MAX, SEGMENTS_PAGE_SIZE_MAX)
    segments_page_size = max(1, min(segments_page_size, SEGMENTS_PAGE_SIZE_MAX))
    df, cache_hit = _get_cached_df(filters)
    rows_in_slice = int(len(df))

    overview = _build_overview_from_service(filters)
    movers = _top_movers(df)
    bundle = {
        "overview": overview,
        "trend_delta": {
            "labels": [m.get("desc") or m.get("sku") or m.get("product_id") for m in movers],
            "values": [m.get("delta_revenue") or 0.0 for m in movers],
            "rows": movers,
        },
        "bubble": _bubble_payload(df, filters, limit=bubble_limit),
    }

    seg_payload = copy.deepcopy(_cached_sales_segments(_filters_to_tuple(filters), _ttl_bucket(PAYLOAD_LRU_TTL_SEC), _cache_version()))
    total_segments = len(seg_payload.get("products") or [])
    seg_payload["products"] = (seg_payload.get("products") or [])[:segments_page_size]
    seg_payload["meta"] = {**seg_payload.get("meta", {}), "total_products": total_segments, "page_size": segments_page_size}
    bundle["segments"] = seg_payload

    table_payload = build_table_payload(
        filters=filters,
        page=1,
        per_page=per_page,
        sort_by=request.args.get("sort_by", "revenue"),
        sort_dir=request.args.get("sort_dir", "desc"),
    )
    bundle["table"] = table_payload
    bundle["meta"] = {
        "duration_ms": int((time.perf_counter() - started) * 1000),
        "rows_in_slice": rows_in_slice,
        "cache_hit": cache_hit,
        "params_hash": _filters_fingerprint(filters),
    }

    status = get_products_parquet_status()
    bundle = _attach_parquet_warning(bundle, status, add_empty_data=not status.available)
    ck = _api_cache_key("bundle", filters, {"bl": bubble_limit, "pp": per_page, "sp": segments_page_size})
    resp = _json_response(bundle, cache_key=ck, ttl=API_TTL_SEC)
    _log_timing(
        "products.api_bundle",
        started,
        filters,
        {
            "rows_in_slice": rows_in_slice,
            "cache_hit": cache_hit,
            "per_page": per_page,
            "bubble_limit": bubble_limit,
            "segments_page_size": segments_page_size,
            "movers": len(movers),
            "products": int(df[CAN.product_id].nunique()) if CAN.product_id in df.columns else 0,
        },
    )
    return resp


@bp.route("/api/recommendations")
@login_required
@requires_roles(*VIEW_ROLES)
def api_recommendations():
    filters = parse_filters(DEFAULT_MONTHS)
    pid = request.args.get("product_id")
    if not pid:
        return _error_response("product_id required", status=400)
    recs = build_product_recommendations(filters, pid)
    status = get_products_parquet_status()
    recs = _attach_parquet_warning(
        recs if isinstance(recs, dict) else {"recommendations": recs},
        status,
        add_empty_data=not status.available,
    )
    ck = _api_cache_key("recommendations", filters, {"pid": str(pid)})
    return _json_response(recs, cache_key=ck, ttl=API_TTL_SEC)


@bp.get("/api/drilldown/bundle")
@login_required
def api_drilldown_bundle():
    from app.services import bundle_service

    payload = bundle_service.drilldown("products", request.args)
    status_code = 200
    if isinstance(payload, dict) and payload.get("error"):
        message = str(payload.get("error", {}).get("message", "")).lower()
        status_code = 404 if "not found" in message else 400
    return jsonify(payload), status_code


@bp.route("/api/also_bought")
@login_required
def api_also_bought():
    filters = parse_filters(DEFAULT_MONTHS)
    pid = request.args.get("product_id") or request.args.get("pid")
    if not pid:
        return _error_response("product_id required", status=400)

    limit = _safe_int(request.args.get("limit", TOP_N_DEFAULT), TOP_N_DEFAULT)
    df = apply_filters(sales_df(), filters)
    payload = _build_also_bought(df, str(pid), limit=limit)
    status = get_products_parquet_status()
    payload = _attach_parquet_warning(payload, status, add_empty_data=not status.available)
    ck = _api_cache_key("also_bought", filters, {"pid": str(pid), "limit": limit})
    return _json_response(payload, cache_key=ck, ttl=API_TTL_SEC)


@bp.route("/api/bubble")
@login_required
@requires_roles(*VIEW_ROLES)
def api_bubble():
    filters = parse_filters(DEFAULT_MONTHS)
    started = time.perf_counter()
    limit = _safe_int(request.args.get("limit") or request.args.get("top_n") or request.args.get("n") or BUBBLE_LIMIT_MAX, BUBBLE_LIMIT_MAX)
    limit = max(1, min(limit, BUBBLE_LIMIT_MAX))
    df, cache_hit = _get_cached_df(filters)
    payload = _bubble_payload(df, filters, limit=limit)
    status = get_products_parquet_status()
    payload = _attach_parquet_warning(payload, status, add_empty_data=not status.available)
    ck = _api_cache_key("bubble", filters, {"limit": limit})
    resp = _json_response(payload, cache_key=ck, ttl=API_TTL_SEC)
    _log_timing(
        "products.api_bubble",
        started,
        filters,
        {
            "rows_in_slice": int(len(df)),
            "cache_hit": cache_hit,
            "limit": limit,
            "products": int(df[CAN.product_id].nunique()) if CAN.product_id in df.columns else 0,
        },
    )
    return resp

def _normalize_overview_export_row(row: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(row or {})
    if out.get("desc") is None:
        out["desc"] = out.get("display_name") or out.get("label") or out.get("product_name")
    if out.get("product_name") is None:
        out["product_name"] = out.get("desc") or out.get("label") or out.get("display_name")
    if out.get("avg_price") is None:
        out["avg_price"] = out.get("unit_price")
    if out.get("current_unit_price") is None:
        out["current_unit_price"] = out.get("unit_price")
    if out.get("target_unit_price") is None:
        out["target_unit_price"] = out.get("target_price")
    if out.get("target_price_27") is None:
        out["target_price_27"] = out.get("target_price")
    if out.get("target_price_21") is None:
        out["target_price_21"] = out.get("target_price")
    if out.get("recommendation_label") is None:
        out["recommendation_label"] = out.get("recommendation") or out.get("quick_rec")
    if out.get("asp_lb") is None:
        out["asp_lb"] = out.get("current_unit_price") if out.get("current_unit_price") is not None else out.get("unit_price")
    if out.get("cost_lb") is None:
        out["cost_lb"] = out.get("unit_cost")
    if out.get("contribution_lb") is None:
        out["contribution_lb"] = out.get("contribution_margin_lb")
    for field in (
        "revenue_current",
        "revenue_prior",
        "revenue_delta",
        "revenue_delta_pct",
        "profit_current",
        "profit_prior",
        "profit_delta",
        "margin_pct_prior",
        "margin_delta_pp",
        "orders_current",
        "orders_prior",
        "supplier",
        "supplier_count",
        "customer_count",
        "region_breadth",
        "top_customer_share",
        "customer_hhi",
        "asp_lb",
        "cost_lb",
        "contribution_lb",
    ):
        out.setdefault(field, out.get(field))
    return out


_DEFAULT_PRODUCTS_EXPORT_FIELDS = [
    "product_id",
    "sku",
    "desc",
    "supplier",
    "customer_count",
    "supplier_count",
    "region_breadth",
    "top_customer_share",
    "customer_hhi",
    "revenue",
    "revenue_current",
    "revenue_prior",
    "revenue_delta",
    "revenue_delta_pct",
    "revenue_share",
    "orders",
    "orders_current",
    "orders_prior",
    "qty",
    "qty_share",
    "unit_price",
    "asp_lb",
    "cost_lb",
    "contribution_lb",
    "avg_price",
    "current_unit_price",
    "cost",
    "profit",
    "profit_current",
    "profit_prior",
    "profit_delta",
    "target_price_27",
    "target_price_21",
    "target_unit_price",
    "uplift_pct",
    "margin_pct",
    "margin_pct_prior",
    "margin_delta_pp",
    "margin",
    "recommendation_label",
]

_PRODUCTS_EXPORT_COLUMN_ALIASES = {
    "product_id": "product_id",
    "sku": "sku",
    "product": "product_name",
    "product_name": "product_name",
    "segment": "segment",
    "revenue": "revenue",
    "revenue_current": "revenue_current",
    "revenue_prior": "revenue_prior",
    "revenue_delta": "revenue_delta",
    "revenue_delta_pct": "revenue_delta_pct",
    "revenue_share": "revenue_share",
    "orders": "orders",
    "orders_current": "orders_current",
    "orders_prior": "orders_prior",
    "qty": "qty",
    "qty_share": "qty_share",
    "weight": "weight",
    "supplier": "supplier",
    "customer_count": "customer_count",
    "supplier_count": "supplier_count",
    "region_breadth": "region_breadth",
    "top_customer_share": "top_customer_share",
    "customer_hhi": "customer_hhi",
    "asp_lb": "asp_lb",
    "cost_lb": "cost_lb",
    "contribution_lb": "contribution_lb",
    "unit_price": "unit_price",
    "current_unit_price": "current_unit_price",
    "target_price": "target_unit_price",
    "target_unit_price": "target_unit_price",
    "uplift_pct": "uplift_pct",
    "cost": "cost",
    "profit": "profit",
    "profit_current": "profit_current",
    "profit_prior": "profit_prior",
    "profit_delta": "profit_delta",
    "profit_share": "profit_share",
    "contribution_margin_lb": "contribution_margin_lb",
    "margin_pct": "margin_pct",
    "margin_pct_prior": "margin_pct_prior",
    "margin_delta_pp": "margin_delta_pp",
    "margin": "margin",
    "price_variance_vs_median": "price_variance_vs_median",
    "volatility_score": "volatility_score",
    "margin_risk": "margin_risk",
    "recommendation": "recommendation_label",
    "recommendation_label": "recommendation_label",
    "quick_rec": "quick_rec",
    "first_sold": "first_sold",
    "last_sold": "last_sold",
}


def _select_export_columns(frame: pd.DataFrame | None, columns: Sequence[str] | None) -> pd.DataFrame:
    """Return a frame with stable ordered columns, filling missing fields safely."""
    if frame is None or not isinstance(frame, pd.DataFrame):
        frame = pd.DataFrame()
    ordered = [str(column).strip() for column in (columns or []) if str(column).strip()]
    if not ordered:
        return frame
    ordered = list(dict.fromkeys(ordered))
    out = frame.copy()
    if out.columns.has_duplicates:
        out = pd.DataFrame.from_records(out.to_dict(orient="records"))
    records = []
    for row in out.to_dict(orient="records"):
        normalized = {column: row.get(column, pd.NA) for column in ordered}
        records.append(normalized)
    return pd.DataFrame.from_records(records, columns=ordered)


def _requested_products_export_fields() -> Optional[List[str]]:
    raw = (request.args.get("columns") or "").strip()
    if not raw:
        return None
    resolved: List[str] = []
    for part in raw.replace(";", ",").split(","):
        key = part.strip().lower()
        if not key:
            continue
        field = _PRODUCTS_EXPORT_COLUMN_ALIASES.get(key)
        if field and field not in resolved:
            resolved.append(field)
    return resolved or None


def _collect_overview_export_rows() -> List[Dict[str, Any]]:
    """
    Export from bundle table pages so exports use the exact same resolved
    filters/scope as KPI+table/drilldown data paths.
    """
    try:
        from app.services import bundle_service

        export_args = MultiDict(request.args)
        try:
            page_size = int(export_args.get("page_size") or export_args.get("per_page") or 200)
        except Exception:
            page_size = 200
        page_size = max(1, min(page_size, 200))
        export_args.setlist("page", ["1"])
        export_args.setlist("page_size", [str(page_size)])
        export_args.setlist("per_page", [str(page_size)])

        first_payload = bundle_service.bundle("products", export_args)
        first_table = first_payload.get("table", {}) if isinstance(first_payload, dict) else {}
        rows = list(first_table.get("rows") or [])
        total = int(first_table.get("total") or len(rows))
        effective_page_size = int(first_table.get("page_size") or page_size or 1)
        total_pages = max(1, math.ceil(total / max(1, effective_page_size)))

        for page in range(2, total_pages + 1):
            export_args.setlist("page", [str(page)])
            payload = bundle_service.bundle("products", export_args)
            table = payload.get("table", {}) if isinstance(payload, dict) else {}
            page_rows = table.get("rows") or []
            if page_rows:
                rows.extend(page_rows)

        normalized = [_normalize_overview_export_row(row) for row in rows if isinstance(row, dict)]
        if normalized:
            return normalized
    except Exception:
        logger.exception("products.overview_export.bundle_failed")

    # Defensive fallback to legacy path if bundle export retrieval fails.
    filters = parse_filters(DEFAULT_MONTHS)
    payload = _build_overview_from_service(filters)
    if not can_view_costs(current_user):
        payload = _sanitize_cost_sensitive_fields(payload)
    legacy_rows = payload.get("top_products") or []
    return [_normalize_overview_export_row(row) for row in legacy_rows if isinstance(row, dict)]


def _safe_float_or_zero(value: Any) -> float:
    try:
        if value is None:
            return 0.0
        val = float(value)
        if math.isnan(val):
            return 0.0
        return val
    except Exception:
        return 0.0


def _mover_status_from_values(current_revenue: float, prior_revenue: float) -> tuple[str, float | None, bool]:
    if prior_revenue <= 0 and current_revenue > 0:
        return "New", None, False
    if current_revenue <= 0 and prior_revenue > 0:
        return "Lost", -100.0, False
    if prior_revenue <= 0:
        return "Stable", None, False
    delta_pct = ((current_revenue - prior_revenue) / prior_revenue) * 100.0
    low_base = prior_revenue < 500.0
    if low_base:
        return "Low base", None, True
    if delta_pct >= 5.0:
        return "Growing", delta_pct, False
    if delta_pct <= -5.0:
        return "Declining", delta_pct, False
    return "Stable", delta_pct, False


def _collect_movers_export_rows() -> List[Dict[str, Any]]:
    rows = _collect_overview_export_rows()
    movers: List[Dict[str, Any]] = []
    for row in rows:
        current = _safe_float_or_zero(row.get("revenue_current") if row.get("revenue_current") is not None else row.get("recent"))
        prior = _safe_float_or_zero(row.get("revenue_prior") if row.get("revenue_prior") is not None else row.get("prior"))
        delta = row.get("revenue_delta")
        delta_val = _safe_float_or_zero(delta) if delta is not None else (current - prior)
        status, delta_pct, low_base = _mover_status_from_values(current, prior)
        movers.append(
            {
                "product_id": row.get("product_id"),
                "sku": row.get("sku") or row.get("product_id"),
                "desc": row.get("desc") or row.get("display_name") or row.get("product_name"),
                "segment": row.get("segment"),
                "revenue_current": current,
                "revenue_prior": prior,
                "delta_revenue": delta_val,
                "delta_revenue_pct": delta_pct,
                "status": status,
                "low_base": bool(low_base),
            }
        )
    movers.sort(key=lambda item: abs(_safe_float_or_zero(item.get("delta_revenue"))), reverse=True)
    return movers


def _safe_float_optional(value: Any) -> float | None:
    try:
        if value is None:
            return None
        parsed = float(value)
        if math.isnan(parsed):
            return None
        return parsed
    except Exception:
        return None


def _quantile(values: List[float], q: float) -> float:
    if not values:
        return 0.0
    return float(np.quantile(np.array(values, dtype=float), q))


def _classify_quadrants(rows: List[Dict[str, Any]]) -> tuple[List[Dict[str, Any]], Dict[str, float]]:
    velocity_vals = [
        _safe_float_optional(r.get("velocity_per_month"))
        for r in rows
    ]
    velocity_vals = [v for v in velocity_vals if v is not None]
    profitability_vals = [
        _safe_float_optional(r.get("margin_pct"))
        if _safe_float_optional(r.get("margin_pct")) is not None
        else _safe_float_optional(r.get("contribution_lb") if r.get("contribution_lb") is not None else r.get("contribution_margin_lb"))
        for r in rows
    ]
    profitability_vals = [v for v in profitability_vals if v is not None]
    v40 = _quantile(velocity_vals, 0.40)
    v50 = _quantile(velocity_vals, 0.50)
    v60 = _quantile(velocity_vals, 0.60)
    p40 = _quantile(profitability_vals, 0.40)
    p50 = _quantile(profitability_vals, 0.50)
    p60 = _quantile(profitability_vals, 0.60)

    def _band(value: float | None, low: float, mid: float, high: float) -> str:
        v = value if value is not None else 0.0
        if v >= high:
            return "high"
        if v <= low:
            return "low"
        return "high" if v >= mid else "low"

    classified: List[Dict[str, Any]] = []
    for row in rows:
        out = dict(row)
        velocity = _safe_float_optional(row.get("velocity_per_month"))
        profitability = _safe_float_optional(row.get("margin_pct"))
        if profitability is None:
            profitability = _safe_float_optional(row.get("contribution_lb") if row.get("contribution_lb") is not None else row.get("contribution_margin_lb"))
        v_band = _band(velocity, v40, v50, v60)
        p_band = _band(profitability, p40, p50, p60)
        if v_band == "high" and p_band == "high":
            quadrant = "protect"
        elif v_band == "high":
            quadrant = "fix_margin"
        elif p_band == "high":
            quadrant = "grow"
        else:
            quadrant = "rationalize"
        out["quadrant"] = quadrant
        classified.append(out)

    return classified, {
        "velocity_p40": v40,
        "velocity_p60": v60,
        "profitability_p40": p40,
        "profitability_p60": p60,
    }


def _build_execution_lists(rows: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    classified, _ = _classify_quadrants(rows)
    pricing_fixes: List[Dict[str, Any]] = []
    cost_fixes: List[Dict[str, Any]] = []
    promote_candidates: List[Dict[str, Any]] = []

    for row in classified:
        margin = _safe_float_optional(row.get("margin_pct"))
        cost = _safe_float_optional(row.get("cost"))
        revenue = _safe_float_or_zero(row.get("revenue"))
        velocity = _safe_float_or_zero(row.get("velocity_per_month"))
        out = {
            "product_id": row.get("product_id"),
            "sku": row.get("sku") or row.get("product_id"),
            "desc": row.get("desc") or row.get("display_name") or row.get("product_name"),
            "segment": row.get("segment"),
            "quadrant": row.get("quadrant"),
            "revenue": revenue,
            "profit": _safe_float_or_zero(row.get("profit")),
            "margin_pct": margin,
            "velocity_per_month": velocity,
        }
        if (row.get("quadrant") == "fix_margin") and (margin is None or margin < 27.0):
            pricing_fixes.append({**out, "action": "Increase / Review cost", "reason": "High velocity with below-target margin and high revenue exposure."})
        if cost is None or cost <= 0:
            cost_fixes.append({**out, "action": "Review cost data", "reason": "Missing or invalid cost blocks reliable margin decisions."})
        if row.get("quadrant") == "grow":
            promote_candidates.append({**out, "action": "Promote / Expand distribution", "reason": "High profitability with lower velocity suggests demand headroom."})

    if not pricing_fixes:
        # Fallback: always surface low-margin candidates when velocity quadrants
        # collapse due sparse/flat distributions.
        for row in classified:
            margin = _safe_float_optional(row.get("margin_pct"))
            if margin is not None and margin >= 27.0:
                continue
            pricing_fixes.append(
                {
                    "product_id": row.get("product_id"),
                    "sku": row.get("sku") or row.get("product_id"),
                    "desc": row.get("desc") or row.get("display_name") or row.get("product_name"),
                    "segment": row.get("segment"),
                    "quadrant": row.get("quadrant"),
                    "revenue": _safe_float_or_zero(row.get("revenue")),
                    "profit": _safe_float_or_zero(row.get("profit")),
                    "margin_pct": margin,
                    "velocity_per_month": _safe_float_or_zero(row.get("velocity_per_month")),
                    "action": "Increase / Review cost",
                    "reason": "Below-target margin candidate selected from fallback rules.",
                }
            )

    pricing_fixes.sort(key=lambda r: (_safe_float_or_zero(r.get("revenue")), _safe_float_or_zero(r.get("velocity_per_month"))), reverse=True)
    cost_fixes.sort(key=lambda r: _safe_float_or_zero(r.get("revenue")), reverse=True)
    promote_candidates.sort(key=lambda r: _safe_float_or_zero(r.get("revenue")), reverse=True)
    return {
        "pricing_fixes": pricing_fixes,
        "cost_fixes": cost_fixes,
        "promote_candidates": promote_candidates,
    }


def _fallback_execution_selection(rows: List[Dict[str, Any]], list_name: str) -> List[Dict[str, Any]]:
    selected: List[Dict[str, Any]] = []
    for row in rows:
        margin = _safe_float_optional(row.get("margin_pct"))
        cost = _safe_float_optional(row.get("cost"))
        revenue = _safe_float_or_zero(row.get("revenue"))
        velocity = _safe_float_or_zero(row.get("velocity_per_month"))
        base = {
            "product_id": row.get("product_id"),
            "sku": row.get("sku") or row.get("product_id"),
            "desc": row.get("desc") or row.get("display_name") or row.get("product_name"),
            "segment": row.get("segment"),
            "quadrant": row.get("quadrant"),
            "revenue": revenue,
            "profit": _safe_float_or_zero(row.get("profit")),
            "margin_pct": margin,
            "velocity_per_month": velocity,
        }
        if list_name == "cost_fixes":
            if cost is None or cost <= 0:
                selected.append({**base, "action": "Review cost data", "reason": "Missing or invalid cost blocks reliable margin decisions."})
            continue
        if list_name == "promote_candidates":
            if margin is not None and margin >= 27.0:
                selected.append({**base, "action": "Promote / Expand distribution", "reason": "Healthy margin profile supports promotion testing."})
            continue
        if margin is None or margin < 27.0:
            selected.append({**base, "action": "Increase / Review cost", "reason": "Below-target margin candidate selected from fallback rules."})

    selected.sort(key=lambda r: (_safe_float_or_zero(r.get("revenue")), _safe_float_or_zero(r.get("velocity_per_month"))), reverse=True)
    return selected


@bp.route("/export/execution.csv")
@login_required
@requires_roles(*VIEW_ROLES)
def export_execution_csv():
    list_name = (request.args.get("list") or "pricing_fixes").strip().lower()
    if list_name not in {"pricing_fixes", "cost_fixes", "promote_candidates"}:
        list_name = "pricing_fixes"

    rows: List[Dict[str, Any]] = []
    selected: List[Dict[str, Any]] = []
    try:
        rows = _collect_overview_export_rows()
        if not isinstance(rows, list):
            rows = []
        lists = _build_execution_lists(rows)
        if isinstance(lists, dict):
            maybe_rows = lists.get(list_name, [])
            if isinstance(maybe_rows, list):
                selected = maybe_rows
    except Exception:
        logger.exception("products.export.execution_failed", extra={"list_name": list_name})
        selected = []

    if not selected and rows:
        selected = _fallback_execution_selection(rows, list_name)
    fields = [
        "product_id",
        "sku",
        "desc",
        "segment",
        "quadrant",
        "revenue",
        "profit",
        "margin_pct",
        "velocity_per_month",
        "action",
        "reason",
    ]
    frame = pd.DataFrame.from_records(selected or []).reindex(columns=fields)
    return dataframe_to_csv_response(frame, filename=f"products_{list_name}.csv")


@bp.route("/export/segment_mix.csv")
@login_required
@requires_roles(*VIEW_ROLES)
def export_segment_mix_csv():
    from app.services import bundle_service

    payload = bundle_service.bundle("products", request.args)
    rows = (((payload or {}).get("charts") or {}).get("segments") or {}).get("mix_shift") or []
    fields = ["segment", "revenue_current", "revenue_prior", "share_current", "share_prior", "share_delta_pp"]
    frame = pd.DataFrame.from_records([row for row in rows if isinstance(row, dict)]).reindex(columns=fields)
    return dataframe_to_csv_response(frame, filename="products_segment_mix.csv")


@bp.route("/export/quadrant.csv")
@login_required
@requires_roles(*VIEW_ROLES)
def export_quadrant_csv():
    quadrant = (request.args.get("quadrant") or "").strip().lower()
    rows = _collect_overview_export_rows()
    try:
        classified, thresholds = _classify_quadrants(rows)
    except Exception:
        logger.exception("products.export.quadrant_failed")
        classified, thresholds = [], {}
    if quadrant:
        classified = [row for row in classified if str(row.get("quadrant") or "").strip().lower() == quadrant]
    fields = [
        "product_id",
        "sku",
        "desc",
        "quadrant",
        "revenue",
        "profit",
        "margin_pct",
        "velocity_per_month",
        "top_customer_share",
        "customer_hhi",
    ]
    quadrant_df = mask_dataframe(pd.DataFrame(classified), current_user, for_export=True)
    if callable(dataframe_to_csv_response):
        resp = dataframe_to_csv_response(quadrant_df, filename="products_quadrant.csv")
    else:
        buf = StringIO()
        writer = csv.DictWriter(buf, fieldnames=list(quadrant_df.columns), extrasaction="ignore")
        writer.writeheader()
        for row in quadrant_df.to_dict(orient="records"):
            writer.writerow(row)
        resp = make_response(buf.getvalue(), 200)
        resp.headers["Content-Type"] = "text/csv; charset=utf-8"
        resp.headers["Content-Disposition"] = "attachment; filename=products_quadrant.csv"
    # Attach thresholds for transparency in response headers for debugging/audit.
    resp.headers["X-Quadrant-Velocity-P40"] = f"{_safe_float_or_zero(thresholds.get('velocity_p40')):.4f}"
    resp.headers["X-Quadrant-Velocity-P60"] = f"{_safe_float_or_zero(thresholds.get('velocity_p60')):.4f}"
    resp.headers["X-Quadrant-Profitability-P40"] = f"{_safe_float_or_zero(thresholds.get('profitability_p40')):.4f}"
    resp.headers["X-Quadrant-Profitability-P60"] = f"{_safe_float_or_zero(thresholds.get('profitability_p60')):.4f}"
    return resp


@bp.route("/export/overview.xlsx")
@login_required
@requires_roles(*VIEW_ROLES)
def export_overview_xlsx():
    rows = _collect_overview_export_rows()
    df_out = mask_dataframe(pd.DataFrame(rows), current_user, for_export=True)
    requested_fields = _requested_products_export_fields()
    if requested_fields:
        df_out = _select_export_columns(df_out, requested_fields)

    buf = BytesIO()
    filename = "products_overview.xlsx"
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    try:
        metadata_rows = [
            {"key": "generated_at", "value": datetime.now(timezone.utc).isoformat()},
            {"key": "dataset_version", "value": str(current_data_version())},
            {"key": "filters_qs", "value": request.query_string.decode("utf-8", errors="ignore")},
            {"key": "row_count", "value": len(df_out.index)},
        ]
        filters = parse_filters(DEFAULT_MONTHS)
        if filters:
            metadata_rows.append({"key": "window_start", "value": str(getattr(filters, "start", "") or "")})
            metadata_rows.append({"key": "window_end", "value": str(getattr(filters, "end", "") or "")})
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name="Top Products")
            pd.DataFrame(metadata_rows).to_excel(writer, index=False, sheet_name="Metadata")
        buf.seek(0)
        content = buf.getvalue()
    except Exception:
        # Fallback to CSV if Excel engine missing
        buf = BytesIO()
        buf.write(df_out.to_csv(index=False).encode("utf-8"))
        buf.seek(0)
        filename = "products_overview.csv"
        mimetype = "text/csv; charset=utf-8"
        content = buf.getvalue()

    resp = make_response(content, 200)
    resp.headers["Content-Type"] = mimetype
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp


@bp.route("/export/overview.csv")
@bp.route("/export/table.csv")
@login_required
@requires_roles(*VIEW_ROLES)
def export_overview_table_csv():
    rows = _collect_overview_export_rows()
    requested_fields = _requested_products_export_fields()
    frame = mask_dataframe(pd.DataFrame.from_records(rows or []), current_user, for_export=True)
    frame = _select_export_columns(frame, requested_fields or _DEFAULT_PRODUCTS_EXPORT_FIELDS)
    buf = StringIO()
    frame.to_csv(buf, index=False)
    resp = make_response(buf.getvalue(), 200)
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = "attachment; filename=products_overview.csv"
    return resp


@bp.route("/export/movers.csv")
@login_required
@requires_roles(*VIEW_ROLES)
def export_movers_csv():
    rows = _collect_movers_export_rows()
    fields = [
        "product_id",
        "sku",
        "desc",
        "segment",
        "revenue_current",
        "revenue_prior",
        "delta_revenue",
        "delta_revenue_pct",
        "status",
        "low_base",
    ]
    frame = pd.DataFrame.from_records(rows or []).reindex(columns=fields)
    return dataframe_to_csv_response(frame, filename="products_movers.csv")


@bp.route("/<product_id>")
@login_required
def product_detail(product_id: str):
    product_id = _normalize_product_id(product_id)
    if product_id.lower().strip("/") == "drilldown" and request.args.get("sku"):
        sku_val = request.args.get("sku")
        qs_args = request.args.to_dict(flat=False)
        for key in ("sku", "product_id", "product", "id"):
            qs_args.pop(key, None)
        qs = urlencode(qs_args, doseq=True)
        target = url_for("products.drilldown", product_id=str(sku_val))
        return redirect(f"{target}?{qs}" if qs else target)
    filters = parse_filters(DEFAULT_MONTHS)
    base_df = sales_df()
    df, filters = _apply_filters_with_recent_fallback(base_df, filters)
    status = get_products_parquet_status()
    sub = df[df[CAN.product_id].astype(str) == str(product_id)]
    try:
        exists_globally = not base_df.empty and str(product_id) in set(base_df[CAN.product_id].astype(str).unique())
    except Exception:
        exists_globally = False
    if sub.empty:
        product_data = {
            "product_id": str(product_id),
            "sku": str(product_id),
            "desc": str(product_id),
            "revenue": 0.0,
            "qty": 0.0,
            "avg_price": None,
            "median_price": None,
        }
        payload = build_overview_payload(filters={**filters, "products": [str(product_id)]})
        payload = _attach_parquet_warning(payload, status, add_empty_data=True)
        recommendations = {"product_id": str(product_id), "recommendations": [], "meta": {"warning": status.warning}}
        return render_template(
            "products/product_detail.html",
            product_id=str(product_id),
            product_data=product_data,
            payload=payload,
            recommendations=recommendations,
            filters=filters,
            products_warning=status.warning or "No data available for this product in the selected window.",
        )

    # Product summary from filtered window
    name = str(sub[CAN.product_name].dropna().iloc[0]) if sub[CAN.product_name].notna().any() else str(product_id)
    revenue = float(sub[CAN.revenue].sum())
    qty_series = _qty_basis_series(sub)
    qty = float(qty_series.sum())
    up = _unit_price(sub).dropna()
    avg_price = float(revenue / qty) if qty else None
    median_price = float(up.median()) if not up.empty else None

    product_data = {
        "product_id": str(product_id),
        "sku": str(product_id),
        "desc": name,
        "revenue": float(round(revenue, 2)),
        "qty": float(round(qty, 2)),
        "avg_price": float(round(avg_price, 2)) if avg_price is not None else None,
        "median_price": float(round(median_price, 2)) if median_price is not None else None,
    }

    # Reuse overview payload (UI often expects it)
    payload = build_overview_payload(filters={**filters, "products": [str(product_id)]})
    recommendations = build_product_recommendations(filters, product_id)

    return render_template(
        "products/product_detail.html",
        product_id=str(product_id),
        product_data=product_data,
        payload=_attach_parquet_warning(payload, status, add_empty_data=not status.available),
        recommendations=_attach_parquet_warning(
            recommendations if isinstance(recommendations, dict) else {"recommendations": recommendations},
            status,
            add_empty_data=not status.available,
        ),
        filters=filters,
        products_warning=status.warning,
    )


@bp.route("/drilldown")
@login_required
def drilldown_query():
    product_id = (
        request.args.get("sku")
        or request.args.get("product_id")
        or request.args.get("product")
        or request.args.get("id")
    )
    if not product_id or not str(product_id).strip():
        abort(404, description="Missing product SKU.")
    qs_args = request.args.to_dict(flat=False)
    for key in ("sku", "product_id", "product", "id"):
        qs_args.pop(key, None)
    qs = urlencode(qs_args, doseq=True)
    target = url_for("products.drilldown", product_id=str(product_id))
    return redirect(f"{target}?{qs}" if qs else target)



@bp.route("/<path:product_id>/drilldown")
@login_required
def drilldown(product_id: str):
    product_id = _normalize_product_id(product_id)
    """
    Product drilldown page with advanced analytics.
    Query params:
      - forecast=1 to include 6-month forecast
      - customer=<id> to focus price suggestion on a customer
    """
    if _product_drilldown_v2_enabled():
        from app.core import access_policy
        from app.services import filters as canonical_filters
        from app.services import product_drilldown_service

        source_payload = request.args.to_dict(flat=False)
        # Legacy drilldown query param used for customer-specific suggestion.
        # It should not narrow the scoped dataset for the full page.
        source_payload.pop("customer", None)
        resolved_filters, _filters_meta = canonical_filters.resolve_filters(
            request,
            current_user_obj=current_user,
            session_obj=session,
            source=source_payload,
            sticky_enabled=True,
            update_session=False,
        )
        context = product_drilldown_service.build_product_drilldown_context(
            str(product_id),
            resolved_filters,
            current_user,
        )
        if not isinstance(context, dict) or context.get("error"):
            abort(404)

        show_costs = can_view_costs(current_user)
        if not show_costs:
            context = _redact_product_drilldown_cost_fields(context)

        legacy_filters = parse_filters(DEFAULT_MONTHS)
        querystring = build_querystring(legacy_filters)
        querystring_suffix = querystring[1:] if querystring.startswith("?") else querystring

        scope_info: Dict[str, Any] = {}
        try:
            candidate = access_policy.get_current_scope(use_cache=True)
            if isinstance(candidate, dict):
                scope_info = candidate
        except Exception:
            scope_info = {}

        meta_payload = context.get("meta") if isinstance(context, dict) else {}
        if not isinstance(meta_payload, dict):
            meta_payload = {}

        return render_template(
            "products/product_drilldown_v2.html",
            product_id=str(product_id),
            product_name=str(meta_payload.get("product_display_label") or meta_payload.get("product_name") or product_id),
            currency_code=current_app.config.get("CURRENCY_CODE", "CAD"),
            qty_title=current_app.config.get("QTY_TITLE", "Quantity"),
            show_costs=show_costs,
            meta=meta_payload,
            kpis=context.get("kpis") or {},
            quality=context.get("quality") or {},
            trend=context.get("trend") or {},
            time_series=context.get("time_series") or {},
            distributions=context.get("distributions") or {},
            customers=context.get("customers") or {},
            regions=context.get("regions") or {},
            suppliers=context.get("suppliers") or {},
            ship_methods=context.get("ship_methods") or {},
            basket=context.get("basket") or {},
            classification=context.get("classification") or {},
            lifecycle=context.get("lifecycle") or {},
            lifecycle_insights=context.get("lifecycle_insights") or {},
            performance_story=context.get("performance_story") or {},
            price_volume=context.get("price_volume") or {},
            margin_risk=context.get("margin_risk") or {},
            weight_analytics=context.get("weight_analytics") or {},
            decision_panel=context.get("decision_panel") or {},
            risk_opportunity=context.get("risk_opportunity") or {},
            header_summary=context.get("header_summary") or {},
            forecast_feature_enabled=bool(context.get("forecast_feature_enabled", _product_forecast_v1_enabled())),
            forecast_enabled=bool(context.get("forecast_enabled")),
            forecast_note=context.get("forecast_note"),
            forecast=context.get("forecast") or {},
            scope_info=scope_info,
            querystring=querystring,
            querystring_suffix=querystring_suffix,
            request_querystring=request.query_string.decode("utf-8", errors="ignore"),
            products_warning=None,
        )

    from app.services import bundle_service
    from werkzeug.datastructures import MultiDict

    filters = parse_filters(DEFAULT_MONTHS)
    querystring = build_querystring(filters)
    querystring_suffix = querystring[1:] if querystring.startswith("?") else querystring

    args = MultiDict(request.args)
    args.setlist("product_id", [str(product_id)])
    show_forecast = str(request.args.get("forecast", "0")).strip().lower() in {"1", "true", "yes", "on"}
    start = filters.get("start") or filters.get("start_date")
    end = filters.get("end") or filters.get("end_date")
    try:
        if current_app.config.get("TESTING") and not (request.args.get("start") or request.args.get("end")):
            start = "2017-01-01"
            end = pd.Timestamp.utcnow().date().isoformat()
    except Exception:
        pass
    if start and not args.get("start"):
        args.setlist("start", [start])
    if end and not args.get("end"):
        args.setlist("end", [end])
    payload = bundle_service.drilldown("products", args)
    if isinstance(payload, dict) and payload.get("error"):
        abort(404)

    kpis = payload.get("kpis") or {}
    rows_count = int(kpis.get("rows") or 0)
    products_warning = None
    if rows_count == 0:
        # Fallback: widen the date window to show the most recent available history.
        try:
            fallback_args = MultiDict(request.args)
            fallback_args.setlist("product_id", [str(product_id)])
            fallback_args.setlist("start", ["2017-01-01"])
            fallback_args.setlist("end", [pd.Timestamp.utcnow().date().isoformat()])
            fallback_args.setlist("date_preset", ["all"])
            payload = bundle_service.drilldown("products", fallback_args)
            kpis = payload.get("kpis") or {}
            rows_count = int(kpis.get("rows") or 0)
            if rows_count > 0:
                products_warning = "No data in the selected window. Showing the most recent available history instead."
            else:
                products_warning = "No data for this SKU in the available history."
        except Exception:
            products_warning = "No data for this SKU in the selected window."

    trend = payload.get("trend") or {}
    months_list = trend.get("labels") or []
    monthly_revenue = trend.get("revenue") or []
    monthly_qty = trend.get("qty") or []

    total_revenue = kpis.get("revenue") or 0.0
    total_qty = kpis.get("qty") or 0.0
    total_weight = kpis.get("weight")
    customers = kpis.get("customers") or 0

    avg_unit_price = (total_revenue / total_qty) if total_qty else None
    asp_recent = None
    if monthly_qty and monthly_revenue and monthly_qty[-1]:
        asp_recent = monthly_revenue[-1] / monthly_qty[-1]

    recent_velocity = None
    if monthly_qty:
        tail = monthly_qty[-min(3, len(monthly_qty)) :]
        recent_velocity = sum(tail) / len(tail) if tail else None

    mom_pct = None
    if len(monthly_revenue) >= 2 and monthly_revenue[-2]:
        mom_pct = (monthly_revenue[-1] - monthly_revenue[-2]) / monthly_revenue[-2] * 100.0

    yoy_pct = None
    if len(monthly_revenue) >= 13 and monthly_revenue[-13]:
        yoy_pct = (monthly_revenue[-1] - monthly_revenue[-13]) / monthly_revenue[-13] * 100.0

    ps = {
        "total_revenue": float(round(total_revenue, 2)),
        "total_qty": float(round(total_qty, 2)),
        "total_weight": float(round(total_weight, 2)) if total_weight is not None else None,
        "customer_count": int(customers),
        "region_count": int(kpis.get("region_count") or 0),
        "supplier_count": int(kpis.get("supplier_count") or 0),
        "avg_unit_price": float(round(avg_unit_price, 2)) if avg_unit_price is not None else None,
        "asp_recent": float(round(asp_recent, 2)) if asp_recent is not None else None,
        "mom_pct": float(round(mom_pct, 1)) if mom_pct is not None else None,
        "yoy_pct": float(round(yoy_pct, 1)) if yoy_pct is not None else None,
        "recent_velocity": float(round(recent_velocity, 2)) if recent_velocity is not None else None,
        "avg_revenue_per_customer": float(round(total_revenue / customers, 2)) if customers else None,
        "first_sold": kpis.get("first_sold"),
        "last_sold": kpis.get("last_sold"),
    }

    price_dist = payload.get("price_distribution") or {}
    unit_price_stats = {
        "p10": price_dist.get("p10"),
        "p50": price_dist.get("p50"),
        "p90": price_dist.get("p90"),
    }
    unit_prices = price_dist.get("samples") or []

    region_labels = [r.get("region") or r.get("Region") for r in (payload.get("top_regions") or []) if isinstance(r, dict)]
    region_values = [float(r.get("revenue") or 0.0) for r in (payload.get("top_regions") or []) if isinstance(r, dict)]

    supplier_labels = [r.get("supplier") or r.get("Supplier") for r in (payload.get("top_suppliers") or []) if isinstance(r, dict)]
    supplier_values = [float(r.get("revenue") or 0.0) for r in (payload.get("top_suppliers") or []) if isinstance(r, dict)]

    top_cust_rows = []
    for r in payload.get("top_customers") or []:
        if not isinstance(r, dict):
            continue
        top_cust_rows.append({
            "CustomerId": r.get("CustomerId") or r.get("customer_id"),
            "Customer": r.get("Customer") or r.get("customer_name") or r.get("label"),
            "Revenue": r.get("Revenue") or r.get("revenue") or 0.0,
            "Qty": r.get("Qty") or r.get("qty") or 0.0,
        })

    cust_options = [
        {"id": r.get("CustomerId") or r.get("customer_id"), "label": r.get("Customer") or r.get("customer_name") or r.get("label")}
        for r in top_cust_rows if (r.get("CustomerId") or r.get("customer_id"))
    ]

    selected_cid = request.args.get("customer")
    price_suggestion = {}
    if selected_cid:
        price_suggestion = {
            "current_price": None,
            "suggested_price": None,
            "rationale": "Insufficient data for customer-specific suggestion",
        }

    forecast_payload = payload.get("forecast") or {}
    if not show_forecast and isinstance(forecast_payload, dict):
        forecast_payload = {**forecast_payload, "forecast": []}
        if not forecast_payload.get("message"):
            forecast_payload["message"] = "Enable forecast to view projection."

    classification = payload.get("classification") or {}
    lifecycle_payload = payload.get("lifecycle") or {}
    bought_together = payload.get("bought_together") or {}

    product_display_name = payload.get("meta", {}).get("entity_display_name") or payload.get("meta", {}).get("entity_label") or str(product_id)
    product_name = payload.get("meta", {}).get("entity_label") or str(product_id)
    currency_code = current_app.config.get("CURRENCY_CODE", "CAD")
    qty_title = current_app.config.get("QTY_TITLE", "Quantity")
    show_costs = can_view_costs(current_user)

    return render_template(
        "products/drilldown.html",
        product_id=str(product_id),
        product_name=product_name,
        product_display_name=product_display_name,
        currency_code=currency_code,
        qty_title=qty_title,
        show_costs=show_costs,
        product_snapshot=ps,
        months=months_list,
        monthly_revenue=monthly_revenue,
        monthly_qty=monthly_qty,
        forecast=forecast_payload,
        unit_price_stats=unit_price_stats,
        unit_prices=unit_prices,
        region_labels=region_labels,
        region_values=region_values,
        supplier_labels=supplier_labels,
        supplier_values=supplier_values,
        top_cust_rows=top_cust_rows,
        cust_options=cust_options,
        selected_cid=selected_cid,
        price_suggestion=price_suggestion,
        classification=classification,
        lifecycle=lifecycle_payload,
        bought_together=bought_together,
        anomalies=[],
        price_insights={},
        recommendations=[],
        products_warning=products_warning,
        querystring=querystring,
        querystring_suffix=querystring_suffix,
    )


@bp.route("/<path:product_id>/drilldown/export")
@login_required
@requires_roles(*VIEW_ROLES)
def export_product_drilldown_v2(product_id: str):
    product_id = _normalize_product_id(product_id)

    from app.services import filters as canonical_filters
    from app.services import product_drilldown_service

    fmt = str(request.args.get("format") or "xlsx").strip().lower()
    if fmt not in {"xlsx", "csv"}:
        fmt = "xlsx"

    kind_raw = str(
        request.args.get("kind")
        or request.args.get("dataset")
        or request.args.get("type")
        or request.args.get("scope")
        or "monthly_series"
    ).strip().lower()
    kind_aliases = {
        "monthly": "monthly_series",
        "series": "monthly_series",
        "customer": "customers",
        "region": "regions",
        "supplier": "suppliers",
        "ship": "ship_methods",
        "forecast_series": "forecast",
    }
    kind = kind_aliases.get(kind_raw, kind_raw)

    resolved_filters, _filters_meta = canonical_filters.resolve_filters(
        request,
        current_user_obj=current_user,
        session_obj=session,
        sticky_enabled=True,
        update_session=False,
    )
    context = product_drilldown_service.build_product_drilldown_context(
        str(product_id),
        resolved_filters,
        current_user,
    )
    if not isinstance(context, dict) or context.get("error"):
        abort(404)

    show_costs = can_view_costs(current_user)
    if not show_costs:
        context = _redact_product_drilldown_cost_fields(context)

    def _csv(df: pd.DataFrame, filename: str):
        if callable(dataframe_to_csv_response):
            return dataframe_to_csv_response(df, filename=filename)
        payload = (df if isinstance(df, pd.DataFrame) else pd.DataFrame()).to_csv(index=False)
        response = make_response(payload, 200)
        response.headers["Content-Type"] = "text/csv; charset=utf-8"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    def _xlsx(sheets: Dict[str, pd.DataFrame], filename: str):
        if callable(dataframes_to_xlsx_response):
            return dataframes_to_xlsx_response(sheets, filename=filename)
        first = next(iter((sheets or {}).values()), pd.DataFrame())
        fallback_name = filename.rsplit(".", 1)[0] + ".csv"
        return _csv(first if isinstance(first, pd.DataFrame) else pd.DataFrame(), fallback_name)

    if kind == "snapshot":
        datasets = context.get("datasets") if isinstance(context, dict) else {}
        if not isinstance(datasets, dict):
            datasets = {}
        meta_payload = context.get("meta") if isinstance(context, dict) else {}
        if not isinstance(meta_payload, dict):
            meta_payload = {}
        metadata_df = pd.DataFrame(
            [
                {"field": "sku", "value": meta_payload.get("sku")},
                {"field": "window_start", "value": meta_payload.get("window_start")},
                {"field": "window_end", "value": meta_payload.get("window_end")},
                {"field": "dataset_version", "value": meta_payload.get("dataset_version")},
                {"field": "generated_at", "value": datetime.now(timezone.utc).isoformat()},
                {"field": "user_id", "value": meta_payload.get("user_id")},
                {"field": "scope_hash", "value": meta_payload.get("scope_hash")},
                {"field": "filters_query", "value": request.query_string.decode("utf-8", errors="ignore")},
                {"field": "export_kind", "value": "snapshot"},
            ]
        )
        sheets = {
            "Summary": pd.DataFrame.from_records(datasets.get("kpis") or []),
            "MonthlySeries": pd.DataFrame.from_records(datasets.get("monthly_series") or []),
            "Customers": pd.DataFrame.from_records(datasets.get("customers") or []),
            "Regions": pd.DataFrame.from_records(datasets.get("regions") or []),
            "Suppliers": pd.DataFrame.from_records(datasets.get("suppliers") or []),
            "ShipMethods": pd.DataFrame.from_records(datasets.get("ship_methods") or []),
            "Basket": pd.DataFrame.from_records(datasets.get("basket") or []),
            "Forecast": pd.DataFrame.from_records(datasets.get("forecast") or []),
            "Seasonality": pd.DataFrame.from_records(datasets.get("seasonality") or []),
            "WeightSummary": pd.DataFrame.from_records([(context.get("weight_analytics") or {}).get("summary") or {}]),
            "Lifecycle": pd.DataFrame.from_records([context.get("lifecycle_insights") or {}]),
            "RiskOpportunity": pd.DataFrame.from_records(
                [{"type": "risk", **row} for row in ((context.get("risk_opportunity") or {}).get("risks") or [])]
                + [{"type": "opportunity", **row} for row in ((context.get("risk_opportunity") or {}).get("opportunities") or [])]
            ),
            "DecisionPanel": pd.DataFrame.from_records((context.get("decision_panel") or {}).get("actions") or []),
            "Metadata": metadata_df,
        }
        stem = sanitize_filename(f"product_{product_id}_drilldown_snapshot")
        if fmt == "csv":
            return _csv(sheets["MonthlySeries"], f"{stem}.csv")
        return _xlsx(sheets, f"{stem}.xlsx")

    data_df, metadata_df, stem = product_drilldown_service.build_export_dataset(context, kind)
    metadata_extra = pd.DataFrame(
        [
            {"field": "filters_query", "value": request.query_string.decode("utf-8", errors="ignore")},
            {"field": "export_kind", "value": kind},
        ]
    )
    metadata_df = pd.concat([metadata_df, metadata_extra], ignore_index=True)
    safe_stem = sanitize_filename(f"{stem}_{product_id}")

    if fmt == "csv":
        return _csv(data_df, f"{safe_stem}.csv")

    sheet_map = {
        "kpis": "Summary",
        "monthly_series": "MonthlySeries",
        "customers": "Customers",
        "regions": "Regions",
        "suppliers": "Suppliers",
        "ship_methods": "ShipMethods",
        "basket": "Basket",
        "forecast": "Forecast",
        "seasonality": "Seasonality",
    }
    sheet_name = sheet_map.get(kind, "Data")
    return _xlsx({sheet_name: data_df, "Metadata": metadata_df}, f"{safe_stem}.xlsx")


@bp.route("/<path:product_id>/export")

@login_required
@requires_roles(*VIEW_ROLES)
def export_product(product_id: str):
    product_id = _normalize_product_id(product_id)
    """
    Export product drilldown data as XLSX or CSV.
    """
    fmt = request.args.get("format", "xlsx").lower()
    if fmt not in ("xlsx", "csv"):
        fmt = "xlsx"
    filters = parse_filters(DEFAULT_MONTHS)
    base_df = sales_df()
    df, filters = _apply_filters_with_recent_fallback(base_df, filters)
    status = get_products_parquet_status()
    sub = df[df[CAN.product_id].astype(str) == str(product_id)]
    qty_series = _qty_basis_series(sub)
    qty_label = _qty_label(sub)

    if sub.empty and not status.available:
        return _json_response(
            _attach_parquet_warning({"data": []}, status, add_empty_data=True),
            status=200,
        )
    if sub.empty:
        abort(404)

    product_name = str(sub[CAN.product_name].dropna().iloc[0]) if sub[CAN.product_name].notna().any() else str(product_id)

    if fmt == "csv":
        sub_export = sub[[CAN.date, CAN.customer_id, CAN.customer_name, CAN.region, CAN.supplier, CAN.revenue]].copy()
        sub_export[CAN.qty] = qty_series.values
        if "missing_packs" in sub.columns:
            sub_export["missing_packs"] = sub["missing_packs"].astype(bool)
        return dataframe_to_csv_response(sub_export, filename=f"product_{product_id}.csv")

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.warning("openpyxl not available for exports")
            return _error_response("Export not available", status=503)

    # XLSX export
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Product Data"

        # Header
        headers = ["Date", "Customer ID", "Customer", "Region", "Supplier", qty_label or "Quantity", "Revenue", "Missing Packs"]
        ws.append(headers)

        # Style header
        header_fill = PatternFill(start_color="7a413a", end_color="7a413a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for _, row in sub.iterrows():
            ws.append([
                row.get(CAN.date, ""),
                row.get(CAN.customer_id, ""),
                row.get(CAN.customer_name, ""),
                row.get(CAN.region, ""),
                row.get(CAN.supplier, ""),
                float(qty_series.get(_, 0.0) if not qty_series.empty else 0.0),
                float(row.get(CAN.revenue, 0)),
                bool(row.get("missing_packs")) if "missing_packs" in sub.columns else None,
            ])

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2.append(["Metric", "Value"])
        ws2.append(["Product ID", str(product_id)])
        ws2.append(["Product Name", product_name])
        ws2.append(["Total Revenue", float(round(sub[CAN.revenue].sum(), 2))])
        ws2.append([f"Total {qty_label}", float(round(qty_series.sum(), 2)) if not qty_series.empty else 0.0])
        ws2.append(["Customer Count", int(sub[CAN.customer_id].nunique())])
        ws2.append(["Record Count", len(sub)])

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        return make_response(out.getvalue(), 200, {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f"attachment; filename=product_{product_id}.xlsx",
        })

    except Exception as e:
        logger.error(f"XLSX export failed for {product_id}: {e}")
        return _error_response("Export generation failed", status=500)


@bp.route("/<path:product_id>/recommendations")
@login_required
@requires_roles(*VIEW_ROLES)
def product_recommendations(product_id: str):
    product_id = _normalize_product_id(product_id)
    filters = parse_filters(DEFAULT_MONTHS)
    out = build_product_recommendations(filters, product_id)
    status = get_products_parquet_status()
    out = _attach_parquet_warning(
        out if isinstance(out, dict) else {"recommendations": out},
        status,
        add_empty_data=not status.available,
    )
    ck = _api_cache_key("product_recs", filters, {"pid": str(product_id)})
    return _json_response(out, cache_key=ck, ttl=API_TTL_SEC)
